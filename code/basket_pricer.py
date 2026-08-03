"""Custom Basket Pricer - price a list of up to 500 cards in one pass.

    .venv/bin/python code/basket_pricer.py --in baskets/my_cards.xlsx
    .venv/bin/python code/basket_pricer.py --seed-set "Pokemon Base Set" \
                                           --variant "1st Edition"

WHY THIS IS NOT THE SCANNER
---------------------------
The scanner reads free-text eBay titles written by strangers, so it must
refuse to guess: "Charizard #4" matches five different Base Set products
priced $370 to $3,099, and picking one would be a fabricated number. That
refusal is correct there and is the whole point of the margin gate.

Here YOU supply the card, so the ambiguity mostly disappears. Give a Set
and a Name and the match is exact and deterministic - no scoring, no
guessing. Free-text names still work, but they run through the SAME
identity resolution the scanner uses and are refused the same way when
they are genuinely ambiguous. An ambiguous row is reported with its
candidates; it is never silently resolved.

Pricing itself is not reimplemented. Grade routing calls the scanner's own
`_guide_cents`, so a PSA 8.5 interpolates between rungs here exactly as it
does in a scan, and a CGC 10 uses the published CGC-10 field rather than
inheriting PSA 10 money.

INPUT
-----
A spreadsheet (.xlsx or .csv) with a column of names and a column of
grades. Header names are matched loosely, so all of these work:

    Card / Name / Product / Item          <- required
    Grade / Condition                     <- optional, blank = ungraded
    Set / Console / Console-Name          <- optional but recommended

COST
----
Local CSVs answer everything they cover at zero API calls. Only rows the
CSVs miss reach the paid API, and never more than --api-cap of them
(default 50), so a 500-row basket cannot quietly burn a run's quota.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import main as scanner                                      # noqa: E402
from valuation.comps import grade_info                      # noqa: E402
from valuation.guide_csv import load_index                  # noqa: E402
from valuation.identity import identity_of                  # noqa: E402
from valuation.price_guide import PriceGuide, _guide_cents  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASKET_DIR = os.path.join(ROOT, "baskets")
MAX_ROWS = 500

NAME_HEADERS = ("card", "name", "product", "product-name", "item", "title")
GRADE_HEADERS = ("grade", "condition", "slab")
SET_HEADERS = ("set", "console", "console-name", "series")
# If your list records what you paid, the report carries it through and
# computes PnL. Absent, the workbook is just a valuation.
COST_HEADERS = ("cost", "paid", "basis", "purchase", "purchase price")

# Priced holdings land here rather than beside the scan reports, so a
# personal valuation is never mistaken for a scan output.
REPORT_DIR = os.path.join(ROOT, "reports", "basket pricer")


def _as_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _money(text) -> float | None:
    """$5,200 / 5200 / (blank) -> 5200.0 / None."""
    cleaned = str(text or "").replace("$", "").replace(",", "").strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    return _as_float(cleaned)


def _norm(text) -> str:
    """Comparison key: case, spacing and punctuation are not identity."""
    return " ".join(str(text or "").split()).casefold()


def _pick_column(headers: list[str], wanted: tuple[str, ...]) -> int | None:
    """Index of the first header matching one of `wanted`.

    Exact match first so a sheet with both "Card" and "Card Number" picks
    "Card", not whichever happened to be leftmost.
    """
    lowered = [_norm(h).replace("_", "-") for h in headers]
    for want in wanted:
        if want in lowered:
            return lowered.index(want)
    for i, h in enumerate(lowered):
        if any(w in h for w in wanted):
            return i
    return None


# ---------------------------------------------------------------- input ---
def read_basket(path: str) -> tuple[list[dict], list[str]]:
    """(rows, warnings). Each row: {name, grade, set, line}."""
    ext = os.path.splitext(path)[1].casefold()
    if ext in (".xlsx", ".xlsm"):
        table = _read_xlsx(path)
    else:
        with open(path, newline="", encoding="utf-8-sig",
                  errors="replace") as fh:
            table = [r for r in csv.reader(fh)]
    table = [r for r in table if any(str(c or "").strip() for c in r)]
    if not table:
        return [], ["the file is empty"]

    headers = [str(c or "") for c in table[0]]
    i_name = _pick_column(headers, NAME_HEADERS)
    i_grade = _pick_column(headers, GRADE_HEADERS)
    i_set = _pick_column(headers, SET_HEADERS)
    i_cost = _pick_column(headers, COST_HEADERS)

    warnings = []
    body = table[1:]
    if i_name is None:
        # No recognisable header: treat column A as names rather than
        # failing outright, but say so - silently discarding a header row
        # would drop a card.
        warnings.append(
            f"no name column found among {headers[:6]} - using the first "
            "column, and treating row 1 as data")
        i_name, body = 0, table
    if i_grade is None:
        warnings.append("no grade column found - pricing everything ungraded")

    rows = []
    for n, raw in enumerate(body, start=2):
        def cell(i):
            return (str(raw[i]).strip()
                    if i is not None and i < len(raw) and raw[i] is not None
                    else "")
        name = cell(i_name)
        if not name:
            continue
        rows.append({"name": name, "grade": cell(i_grade),
                     "set": cell(i_set), "line": n,
                     "cost": _money(cell(i_cost))})
    if len(rows) > MAX_ROWS:
        warnings.append(f"{len(rows)} rows supplied - pricing the first "
                        f"{MAX_ROWS} and ignoring the rest")
        rows = rows[:MAX_ROWS]
    return rows, warnings


def _read_xlsx(path: str) -> list[list]:
    try:
        import openpyxl
    except ImportError:                                     # pragma: no cover
        raise SystemExit("openpyxl is required to read .xlsx baskets:\n"
                         "    .venv/bin/pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True)
    return [list(r) for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]


# -------------------------------------------------------------- matching ---
# People do not write card names the way a price guide stores them. The
# guide says "Alakazam [1st Edition] #1"; a collector's own list says
# "1999 1st Edition Alakazam". Both name one card, and refusing the second
# would make the tool useless for the lists people actually keep.
_YEAR_PREFIX_RE = re.compile(r"^\s*(?:19|20)\d{2}(?:\s*[-/]\s*\d{2,4})?\s+")
_VARIANT_WORDS = (
    "1st edition", "first edition", "shadowless", "no rarity", "no symbol",
    "red cheeks", "yellow cheeks", "ghost stamp", "gray stamp", "grey stamp",
    "1999-2000", "error", "black dot error",
)
# Base Set Pikachu is three products at #58 and the guide labels only the
# unusual ones. "Yellow" is the ordinary card and carries no label, so a
# list saying "Pikachu Yellow" must land on the UNLABELLED product - the
# opposite of what a substring search would do.
_VARIANT_ALIASES = {
    "yellow": "", "yellow cheeks": "", "red": "red cheeks",
    "1st ed": "1st edition", "1st": "1st edition",
}


_BRACKET_RE = re.compile(r"\[([^\]]*)\]")


def _product_variants(product_name: str) -> tuple[str, list[str]]:
    """Fast path for GUIDE names, which have a fixed shape.

    'Charizard [1st Edition] #4' -> ('Charizard', ['1st edition'])

    This runs once per row over a million-row index, so it must not do what
    the input-side parser does (a dozen regex passes per name). Doing so
    took the index build from under a second to minutes.
    """
    name = str(product_name or "")
    core = name.split("[")[0].split("#")[0].strip()
    variants: list[str] = []
    if "[" in name:
        for chunk in _BRACKET_RE.findall(name):
            low = chunk.casefold()
            for word in sorted(_VARIANT_WORDS, key=len, reverse=True):
                if word in low:
                    variants.append(word)
                    low = low.replace(word, " ")
    return core, variants


def _split_variants(name: str) -> tuple[str, list[str]]:
    """('Alakazam', ['1st edition']) from '1999 1st Edition Alakazam'."""
    text = _YEAR_PREFIX_RE.sub("", str(name or "")).strip()
    found = []
    for word in sorted(_VARIANT_WORDS, key=len, reverse=True):
        pattern = re.compile(rf"\[?\b{re.escape(word)}\b\]?", re.I)
        if pattern.search(text):
            found.append(word)
            text = pattern.sub(" ", text)
    # trailing colour words that name a Pikachu variant rather than a card
    words = text.split()
    while words and words[-1].casefold() in _VARIANT_ALIASES:
        mapped = _VARIANT_ALIASES[words.pop().casefold()]
        if mapped:
            found.append(mapped)
    core = re.sub(r"#\s*\d+", " ", " ".join(words))
    core = " ".join(core.replace("[", " ").replace("]", " ").split())
    return core, [f.casefold() for f in found]


class ExactIndex:
    """(set, product-name) -> rows, plus product-name -> rows.

    The scanner searches by relevance because it cannot trust its input.
    A basket names the product, so an exact key is both faster and - more
    importantly - incapable of returning a confident wrong answer.
    """

    def __init__(self, rows):
        self.by_set_name: dict[tuple, list] = defaultdict(list)
        self.by_name: dict[str, list] = defaultdict(list)
        self.by_core: dict[str, list] = defaultdict(list)
        for row in rows:
            name = _norm(row.get("product-name"))
            console = _norm(row.get("console-name"))
            if not name:
                continue
            self.by_set_name[(console, name)].append(row)
            self.by_name[name].append(row)
            core, _ = _product_variants(row.get("product-name") or "")
            if core:
                self.by_core[_norm(core)].append(row)

    def flexible(self, name: str, set_name: str = "") -> tuple[list, str]:
        """Resolve a human-written name to products. Never guesses.

        Requires the variant words to match EXACTLY both ways: a list
        saying "1st Edition" must not match a Shadowless product, and a
        list saying plain "Charizard" must not match the 1st Edition one.
        """
        core, want = _split_variants(name)
        rows = list(self.by_core.get(_norm(core), []))
        if set_name:
            key = _norm(set_name)
            narrowed = [r for r in rows
                        if key in _norm(r.get("console-name"))
                        or _norm(r.get("console-name")) in key]
            rows = narrowed or rows
        hits = []
        for row in rows:
            _, have = _product_variants(row.get("product-name") or "")
            if sorted(set(have)) == sorted(set(want)):
                hits.append(row)
        return hits, ("name + variant" if hits else "")

    def lookup(self, name: str, set_name: str = "") -> tuple[list, str]:
        """(candidate rows, how). Empty list means no exact hit."""
        key = _norm(name)
        if set_name:
            hit = self.by_set_name.get((_norm(set_name), key))
            if hit:
                return hit, "exact set + name"
            # The set was given but did not match: try a looser set match
            # before giving up, since "Base Set" and "Pokemon Base Set" are
            # the same shelf to a human.
            want = _norm(set_name)
            loose = [r for r in self.by_name.get(key, [])
                     if want in _norm(r.get("console-name"))
                     or _norm(r.get("console-name")) in want]
            if loose:
                return loose, "exact name, fuzzy set"
        hit = self.by_name.get(key)
        if hit:
            return hit, "exact name"
        return [], ""


# Words that change WHAT THE OBJECT IS rather than describing it. If your
# list says one of these and the closest product does not, they are not the
# same asset, however well the rest of the name lines up.
#
# Deliberately NARROW. Autograph, relic and parallel are already modelled as
# structured fields on CardIdentity and enforced by conflicts_with(), which
# compares identities rather than guessing from a product name. Listing them
# here too would double-refuse: a product legitimately catalogued as
# "Mike Trout #BCP111" with the autograph expressed elsewhere would be
# thrown out for not having "auto" in its title.
#
# What is left is vocabulary the identity system does not model, where the
# only signal IS the words. "Michael Jordan 1986 Fleer Sticker RC" resolved
# to the base rookie "Michael Jordan #57" and priced at $15,604.
_DISTINGUISHING = (
    "sticker", "promo", "error", "reprint", "proof", "box topper",
    "oversized", "blank back", "puzzle", "wrapper",
)


def _distinguishing_mismatch(supplied: str, matched: str) -> str:
    """The first identity-changing word present in one name and not the
    other, or "" when they agree."""
    a, b = _norm(supplied), _norm(matched)
    for word in _DISTINGUISHING:
        if (word in a) != (word in b):
            return word
    return ""


def price_row(row: dict, index: ExactIndex, guide: PriceGuide,
              api_budget: list[int]) -> dict:
    """Price one basket line. Never guesses between conflicting products."""
    name, set_name, grade_text = row["name"], row["set"], row["grade"]

    # Grade parsing is the scanner's, so "PSA 8.5", "BGS 9", "CGC 10",
    # "SGC 92" and "raw" all mean here exactly what they mean in a scan.
    gi = grade_info(f"{name} {grade_text}" if grade_text else name)
    # grade_info reports its grades as strings; price_guide casts them with
    # float() before use and so must we, or _guide_cents compares str to
    # float and dies on the first graded row.
    grader = gi[0] if gi else None
    printed = _as_float(gi[1]) if gi else None
    eff = _as_float(gi[2]) if gi else None

    out = dict(row, grader=grader or "", printed_grade=printed,
               effective_grade=eff, price=None, source="", how="",
               matched_name="", matched_set="", note="")

    candidates, how_matched = index.lookup(name, set_name)
    if not candidates:
        candidates, how_matched = index.flexible(name, set_name)

    if len(candidates) > 1:
        # Collapse duplicates of the SAME product across files (a set can
        # appear in more than one CSV); genuine rivals are different sets.
        sets = {_norm(c.get("console-name")) for c in candidates}
        if len(sets) > 1:
            shown = sorted({str(c.get("console-name")) for c in candidates})
            out["note"] = ("ambiguous: matches %d sets - add a Set column "
                           "to choose (%s)" % (len(sets), "; ".join(shown[:4])))
            return out
        candidates = candidates[:1]

    if candidates:
        missing = _distinguishing_mismatch(
            name, candidates[0].get("product-name") or "")
        if missing:
            # 2026-08-02: "Michael Jordan 1986 Fleer Sticker RC" resolved to
            # "Michael Jordan #57" - the base rookie - and was priced at
            # $15,604. The sticker is a different product entirely. A word
            # that changes what the object IS must appear on both sides.
            out.update(matched_name=candidates[0].get("product-name") or "",
                       matched_set=candidates[0].get("console-name") or "")
            out["note"] = (f"refused: your list says {missing!r} but the "
                           f"closest product does not - different product")
            return out
        product = candidates[0]
        cents, how = _guide_cents(product, eff, grader=grader,
                                  printed_grade=printed)
        out.update(matched_name=product.get("product-name") or "",
                   matched_set=product.get("console-name") or "",
                   source=f"local CSV ({how_matched})", how=how)
        if cents is None:
            out["note"] = how or "no usable price field for this grade"
        else:
            out["price"] = round(cents / 100.0, 2)
        return out

    # No exact hit: fall back to the scanner's own identity resolution.
    #
    # That path reads the LOCAL CSVs before it reaches for the API, so the
    # budget must gate only the paid part. Gating the whole call made
    # --api-cap 0 mean "resolve nothing" instead of "stay free", and a
    # 37-row list of free-text names priced 0 of 37 with the guide data for
    # most of them sitting on disk.
    if guide is None:
        out["note"] = "not found in local CSVs"
        return out
    query = f"{name} {grade_text}".strip()
    paid_allowed = api_budget[0] > 0
    hosts = getattr(guide, "guide_hosts", [])
    if not paid_allowed:
        guide.guide_hosts = []
    try:
        quote = guide.quote(identity_of(query))
    finally:
        guide.guide_hosts = hosts
    if paid_allowed:
        api_budget[0] -= 1
    if quote.value:
        # Same guard as the exact path. This is the branch the Jordan
        # STICKER actually came through: the identity resolver matched it
        # to the base rookie "Michael Jordan #57" and priced it $15,604.
        missing = _distinguishing_mismatch(name, quote.product_name or "")
        if missing:
            out.update(matched_name=quote.product_name or "",
                       matched_set=quote.console_name or "",
                       note=f"refused: your list says {missing!r} but the "
                            f"closest product does not - different product")
            return out
        out.update(price=round(float(quote.value), 2),
                   source=f"guide lookup ({quote.match})",
                   how=quote.how, matched_name=quote.product_name or "",
                   matched_set=quote.console_name or "")
    else:
        # "no guide host configured" is what the guide says when the local
        # CSVs could not land it and no paid host was allowed. Told to a
        # person staring at a Mike Trout that did not price, that is not an
        # explanation - so say which of the two actually happened.
        note = quote.note or "not found in CSVs or guide"
        if not paid_allowed and "no guide host" in note:
            note = ("no local guide covers this card - re-run without "
                    "--api-cap 0 so the paid guide can answer it")
        out["note"] = note
    return out


# --------------------------------------------------------------- seeding ---
# Sealed product sits in the same catalogue as singles: the 1st Edition
# Base Set listing includes a $14,179 Booster Box and a $10,108 Blister
# Pack. Priced as if they were PSA 10 cards they added ~$32k to a set
# total that was supposed to be "all 100+ cards".
_SEALED_WORDS = r"booster|box|blister|deck|tin|bundle|pack|collection|case"
_SEALED_RE = None
_NUMBER_RE = None


def _is_sealed(name: str) -> bool:
    """Sealed product rather than a single card.

    Word boundaries matter here: a substring test calls 'Dratini' and
    'Fighting Energy' sealed because they contain 'tin'. The decisive
    signal is the card number - singles carry one, sealed product does
    not - with the keyword only used to confirm.
    """
    global _SEALED_RE, _NUMBER_RE
    import re
    if _SEALED_RE is None:
        _SEALED_RE = re.compile(rf"\b({_SEALED_WORDS})\b", re.I)
        _NUMBER_RE = re.compile(r"#\s*\d+")
    if _NUMBER_RE.search(name or ""):
        return False
    return bool(_SEALED_RE.search(name or ""))


def seed_set(index_rows, set_name: str, variant: str = "",
             include_sealed: bool = False) -> tuple[list[dict], list[dict]]:
    """(card rows, sealed rows) for a set, ready for grades.

    Typing 109 card names by hand is how a basket gets typos, and a typo
    here is a silently missing line item rather than an error.
    """
    want = _norm(set_name)
    cards, sealed = [], []
    for row in index_rows:
        console = _norm(row.get("console-name"))
        if console != want and want not in console:
            continue
        name = str(row.get("product-name") or "")
        if variant and _norm(variant) not in _norm(name):
            continue
        entry = {"name": name, "set": str(row.get("console-name") or ""),
                 "grade": ""}
        (sealed if _is_sealed(name) else cards).append(entry)
    if include_sealed:
        cards += sealed
        sealed = []
    cards.sort(key=lambda r: _card_sort_key(r["name"]))
    sealed.sort(key=lambda r: _card_sort_key(r["name"]))
    return cards, sealed


def _card_sort_key(name: str):
    """Sort by card number when there is one, so a set reads in order."""
    import re
    m = re.search(r"#\s*([0-9]+)", name)
    return (0, int(m.group(1)), name) if m else (1, 0, name)


# ---------------------------------------------------------------- output ---
def write_report(path: str, priced: list[dict], warnings: list[str],
                 basket_name: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:                                     # pragma: no cover
        raise SystemExit("openpyxl is required:\n"
                         "    .venv/bin/pip install openpyxl")

    ok = [r for r in priced if r["price"] is not None]
    bad = [r for r in priced if r["price"] is None]
    has_cost = any(r.get("cost") is not None for r in priced)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Priced"
    headers = ["Line", "Card", "Set", "Grade"]
    if has_cost:
        headers += ["Cost", "Value", "PnL"]
    else:
        headers += ["Value"]
    headers += ["Matched Product", "Matched Set", "Source", "Price Field"]
    ws.append(headers)
    for r in ok:
        row = [r["line"], r["name"], r["set"], r["grade"]]
        if has_cost:
            cost = r.get("cost")
            pnl = (r["price"] - cost) if cost is not None else None
            row += [cost, r["price"], pnl]
        else:
            row += [r["price"]]
        row += [r["matched_name"], r["matched_set"], r["source"], r["how"]]
        ws.append(row)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    widths = ([6, 42, 26, 10] + ([14, 14, 14] if has_cost else [14])
              + [42, 26, 26, 34])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    money_first, money_last = 5, (7 if has_cost else 5)
    for row in ws.iter_rows(min_row=2, min_col=money_first,
                            max_col=money_last):
        for c in row:
            c.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'

    ws2 = wb.create_sheet("Not Priced")
    ws2.append(["Line", "Card", "Set", "Grade", "Why"])
    for r in bad:
        ws2.append([r["line"], r["name"], r["set"], r["grade"],
                    r["note"] or "no price"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.freeze_panes = "A2"
    for i, w in enumerate([6, 42, 26, 10, 68], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Summary")
    total = sum(r["price"] for r in ok)
    cost_total = sum(r["cost"] for r in priced if r.get("cost") is not None)
    local = sum(1 for r in ok if r["source"].startswith("local"))
    lines = [
        ("Basket", basket_name),
        ("Line items supplied", len(priced)),
        ("Priced", len(ok)),
        ("Not priced", len(bad)),
        ("", ""),
        ("Total value", total),
        ("Total cost", cost_total) if has_cost else ("", ""),
        ("Total PnL", total - cost_total) if has_cost else ("", ""),
        ("", ""),
        ("Priced from local CSV (no API calls)", local),
        ("Priced via guide lookup", len(ok) - local),
        ("", ""),
        ("NOTE", "A basket total is guide value, not a sale price. It "
                 "ignores fees, shipping and the fact that selling 100 "
                 "cards at once moves the market."),
    ]
    for label, value in lines:
        ws3.append([label, value])
    for w in warnings:
        ws3.append(["WARNING", w])
    for cell in ("B6", "B7", "B8"):
        ws3[cell].number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    for c in ws3["A"]:
        c.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 76

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def write_seed(path: str, rows: list[dict]) -> None:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Basket"
    ws.append(["Card", "Set", "Grade"])
    for r in rows:
        ws.append([r["name"], r["set"], r["grade"]])
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, w in enumerate([46, 28, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


# ------------------------------------------------------------------ main ---
def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Price a basket of cards.")
    p.add_argument("--in", dest="infile", help="basket .xlsx or .csv")
    p.add_argument("-o", "--out", help="where to write the priced workbook")
    p.add_argument("--seed-set", help="write a starter basket for a set")
    p.add_argument("--variant", default="",
                   help="only products whose name contains this "
                        "(e.g. '1st Edition')")
    p.add_argument("--api-cap", type=int, default=50,
                   help="most rows allowed to reach the paid API (default 50)")
    p.add_argument("--include-sealed", action="store_true",
                   help="keep booster boxes/packs in a seeded set "
                        "(default: cards only)")
    p.add_argument("--list-sets", action="store_true",
                   help="show set names the local CSVs cover")
    args = p.parse_args(argv)

    config = scanner.load_config(os.path.join(ROOT, "config.yaml"))
    index = load_index(ROOT)
    if not len(index):
        print("No price-guide CSVs found. Run 'Download Price Guides.command'")
        return 1

    if args.list_sets:
        counts: dict[str, int] = defaultdict(int)
        for row in index.rows:
            counts[str(row.get("console-name") or "?")] += 1
        for name, n in sorted(counts.items(), key=lambda kv: -kv[1])[:60]:
            print(f"  {n:>6}  {name}")
        print(f"\n{len(counts):,} sets covered by {len(index):,} products.")
        return 0

    if args.seed_set:
        rows, sealed = seed_set(index.rows, args.seed_set, args.variant,
                                include_sealed=args.include_sealed)
        if not rows:
            print(f"No products found for set {args.seed_set!r}"
                  + (f" with variant {args.variant!r}" if args.variant else "")
                  + ".\nTry --list-sets to see the exact names.")
            return 1
        label = args.seed_set + (f" {args.variant}" if args.variant else "")
        out = args.out or os.path.join(
            BASKET_DIR, f"{_slug(label)}.xlsx")
        write_seed(out, rows)
        print(f"Wrote {len(rows)} card line items -> {out}")
        if sealed:
            print(f"\nLeft out {len(sealed)} sealed product(s) - they are in "
                  "the same\ncatalogue but are not single cards, and priced "
                  "as slabs they would\ninflate the set total:")
            for s in sealed:
                print(f"    {s['name']}")
            print("  Add them with --include-sealed if you want them.")
        print("\nFill in the Grade column (blank = ungraded), then price it:")
        print(f'  .venv/bin/python code/basket_pricer.py --in "{out}"')
        return 0

    if not args.infile:
        p.error("give --in <basket file>, or --seed-set to make one")

    rows, warnings = read_basket(args.infile)
    if not rows:
        print(f"No card rows found in {args.infile}")
        for w in warnings:
            print(f"  {w}")
        return 1

    guide = PriceGuide(config)
    exact = ExactIndex(index.rows)
    budget = [max(0, args.api_cap)]

    # Cards the local CSVs cannot answer go to the paid guide at one call
    # per second, so a long basket is a long silence. Say what is happening.
    print(f"  pricing {len(rows)} line item(s) from "
          f"{len(index):,} guide products...", flush=True)
    priced = []
    for n, row in enumerate(rows, start=1):
        priced.append(price_row(row, exact, guide, budget))
        if n % 10 == 0 or n == len(rows):
            got = sum(1 for r in priced if r["price"] is not None)
            print(f"    {n}/{len(rows)}  ({got} priced)", flush=True)

    basket_name = os.path.basename(args.infile)
    out = args.out or os.path.join(
        REPORT_DIR, f"{_slug(os.path.splitext(basket_name)[0])}.xlsx")
    write_report(out, priced, warnings, basket_name)

    ok = [r for r in priced if r["price"] is not None]
    total = sum(r["price"] for r in ok)
    local = sum(1 for r in ok if r["source"].startswith("local"))
    print(f"  priced      {len(ok)} of {len(priced)}")
    print(f"  total       ${total:,.2f}")
    print(f"  local CSV   {local} (no API calls)   guide lookups {len(ok)-local}")
    costs = [r["cost"] for r in priced if r.get("cost") is not None]
    if costs:
        print(f"  cost        ${sum(costs):,.2f}")
        print(f"  PnL         ${total - sum(costs):,.2f}")
    if len(priced) - len(ok):
        print(f"  not priced  {len(priced)-len(ok)} - see the 'Not Priced' tab")
    for w in warnings:
        print(f"  note: {w}")
    print(f"\n  -> {out}")
    return 0


def _slug(text: str) -> str:
    import re
    return re.sub(r"[^a-z0-9]+", "-", str(text).casefold()).strip("-") or "basket"


if __name__ == "__main__":
    raise SystemExit(main())
