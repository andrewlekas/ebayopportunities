"""Excel report.

Tabs:
  Action        - what needs a decision now: top scores + ending-soon with
                  positive EV
  Pokemon / Sports / Watches / Games / Pop Culture - the full book by category
  Discovery     - broad theme queries (browse only, soft valuations)
  Movers        - biggest 30-day fair-value moves (needs history)
  About         - column definitions

Formatting: titles are hyperlinks, timing shown as "ends in 3h 40m" (amber
when <6h) or "listed 2h ago" (green when fresh), whole dollars >= $1k,
trend arrows, colored listing types, banded rows, data bars on Edge Now,
audit columns (comps/guide/#comps/expected cost) grouped & hidden.
"""
from __future__ import annotations

import math
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from economics import (best_exit_route, item_price_for_total_cost,
                       sales_tax_rate)
from models import Opportunity
from quality import is_tradeable
from textutil import fold
from valuation.comps import (_subject_tokens, grade_info, title_match_score,
                             variant_conflict)

HEADERS = [
    ("Rank", 6), ("Pri", 5), ("Type", 10), ("Site", 11), ("Title", 56),
    ("Query", 25), ("Price", 11), ("Ship", 9), ("Bids", 6), ("Timing", 20),
    ("Fair Value", 12), ("Trend 30d", 10),
    ("Comps Val", 11), ("Guide Val", 11), ("#Comps", 8), ("Exp Cost", 11),
    ("Expected Value", 13), ("Edge Now", 11),
    ("ROI", 8), ("Sales/mo", 9), ("Ann ROI", 9),
    ("Capture", 8), ("Conf", 7), ("Score", 9),
    ("Best Exit", 14), ("Exit Fee", 9), ("Net Proceeds", 13),
    ("vs eBay", 11), ("Notes", 36),
    ("Model Detail", 34),          # machine diagnostics, hidden
]
MONEY_COLS = (7, 8, 11, 13, 14, 16, 17, 18, 27, 28)
AUDIT_COLS = ("M", "P")           # Comps Val .. Exp Cost, grouped/hidden
MODEL_COL = "AD"                  # Model Detail, grouped/hidden

WATCH_KW = ["rolex", "patek", "audemars", "moonphase", "daytona",
            "submariner", "datejust", "world time", "royal oak"]
# A PLATFORM word means it is definitely a video game, whatever else the
# query says ("Pokemon Red Gameboy" is a game).
GAME_PLATFORM_KW = ["nes", "n64", "snes", "xbox", "gameboy", "game boy",
                    "nintendo", "playstation", "sega", "atari", "wata",
                    "vga", "cartridge"]
# A TITLE word is a game name. Checked AFTER Pokemon, because a card query
# can mention one and "sealed" describes card product just as often as it
# describes a boxed game - it used to live here and sent every sealed
# Pokemon query to the Video Games tab, which also let those rows skip the
# Pokemon grade floor (that floor only applies to the Pokemon category).
GAME_TITLE_KW = ["goldeneye", "halo", "sims", "tycoon", "call of duty",
                 "pinball", "botw", "breath of the wild", "ocarina",
                 "mario", "zelda", "duck hunt"]
POKE_KW = ["pokemon", "charizard", "blastoise", "venusaur", "pikachu",
           "topsun", "carddass", "umbreon", "raichu", "gyarados", "gengar",
           "dragonite", "snorlax", "flareon", "jolteon", "vaporeon",
           "chansey", "poliwrath", "pidgeotto", "trader", "no rarity",
           "jungle", "fossil", "base set"]
POP_KW = ["superman", "batman", "disney", "skywalker", "star wars", "marvel"]


CATEGORIES = ("Pokemon Cards", "Sports Cards", "Video Games", "Watches",
              "Other")


def _category(query: str) -> str:
    q = fold(query).lower()
    if any(k in q for k in WATCH_KW):
        return "Watches"
    if any(k in q for k in GAME_PLATFORM_KW):
        return "Video Games"
    if any(k in q for k in POKE_KW):
        return "Pokemon Cards"
    if any(k in q for k in GAME_TITLE_KW):
        return "Video Games"
    if any(k in q for k in POP_KW):
        return "Other"
    return "Sports Cards"


def _timing(l) -> tuple[str, str]:
    """(text, highlight) where highlight is '', 'amber' or 'green'.

    Auction end times are ABSOLUTE local clock times - a countdown
    ("ends 2h 59m") is true when the file is written and a lie by the
    time it's read at the end of the day."""
    now = datetime.now(timezone.utc)
    if l.listing_type == "auction" and l.end_time:
        secs = (l.end_time - now).total_seconds()
        if secs <= 0:
            return "ended", ""
        end_local = l.end_time.astimezone()          # Mac-local clock
        clock = end_local.strftime("%I:%M %p").lstrip("0")
        if secs <= 24 * 3600:
            day = ("today" if end_local.date() == now.astimezone().date()
                   else "tomorrow")
            text = f"ends {clock} {day}"
        elif secs <= 6 * 86400:
            text = f"ends {end_local.strftime('%a')} {clock}"
        else:
            d, h = int(secs // 86400), int(secs % 86400 // 3600)
            text = f"ends {d}d {h}h"
        return text, ("amber" if secs <= 6 * 3600 else "")
    if l.listing_type == "fixed" and l.created_at:
        secs = (now - l.created_at).total_seconds()
        d, h = int(secs // 86400), int(secs % 86400 // 3600)
        text = (f"listed {d}d ago" if d else f"listed {h}h ago")
        return text, ("green" if secs <= 6 * 3600 else "")
    return "", ""


def _money_fmt(val: float) -> str:
    return '"$"#,##0' if abs(val or 0) >= 1000 else '"$"#,##0.00'


def _exit_name(channel: str) -> str:
    return {"psa_vault": "PSA Vault",
            "fanatics_collect": "Fanatics Collect"}.get(
                channel or "", (channel or "").replace("_", " ").title())


def _fill_sheet(ws, opps: list[Opportunity]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, (name, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    band = PatternFill("solid", fgColor="F5F7FA")
    amber = PatternFill("solid", fgColor="FFF2CC")
    green = PatternFill("solid", fgColor="E2EFDA")
    type_font = {"AUCTION": Font(color="1F4E79"),
                 "BIN": Font(color="2E7D32", bold=True),
                 "BIN+OBO": Font(color="1B5E20", bold=True)}

    for i, o in enumerate(opps, 1):
        l, v = o.listing, o.valuation
        row = i + 1
        ltype = ("BIN+OBO" if l.listing_type == "fixed" and l.best_offer
                 else "BIN" if l.listing_type == "fixed" else "AUCTION")
        timing, hl = _timing(l)
        values = [i, "★" if l.priority else "", ltype, l.site, l.title,
                  l.query, l.current_price, l.shipping, l.bid_count, timing,
                  v.fair_value, v.trend_30d, v.comps_value, v.guide_value,
                  v.n_comps, v.expected_cost, v.expected_value, v.edge_now,
                  v.roi, v.sales_per_month, v.annualized_roi,
                  v.capture, v.confidence, v.opportunity_score,
                  _exit_name(v.resale_channel),
                  v.resale_fee_rate, v.net_proceeds, v.exit_advantage,
                  "; ".join(v.notes + ([getattr(o, "dupe_note")]
                                       if getattr(o, "dupe_note", None)
                                       else [])),
                  "; ".join(getattr(v, "audit_notes", []))]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            if i % 2 == 0:
                c.fill = band
        for col in MONEY_COLS:
            cell = ws.cell(row=row, column=col)
            cell.number_format = _money_fmt(cell.value or 0)
        ws.cell(row=row, column=12).number_format = '"▲"0%;"▼"0%;"-"'
        ws.cell(row=row, column=19).number_format = "0.0%"   # ROI
        ws.cell(row=row, column=20).number_format = "0.0"    # Sales/mo
        ws.cell(row=row, column=21).number_format = "0%"     # Ann ROI
        ws.cell(row=row, column=22).number_format = "0%"     # Capture
        ws.cell(row=row, column=23).number_format = "0%"     # Conf
        ws.cell(row=row, column=26).number_format = "0.0%"   # Exit Fee
        ws.cell(row=row, column=24).number_format = "0.0%"   # Score

        title = ws.cell(row=row, column=5)
        if l.url:
            title.hyperlink = l.url
        title.font = Font(color="0563C1", underline="single")

        tcell = ws.cell(row=row, column=3)
        tcell.font = type_font.get(ltype, Font())
        if l.priority:
            p = ws.cell(row=row, column=2)
            p.font = Font(bold=True, color="B8860B")
            p.alignment = Alignment(horizontal="center")
        if hl:
            tm = ws.cell(row=row, column=10)
            tm.fill = amber if hl == "amber" else green
            tm.font = Font(bold=True)

    ws.column_dimensions.group(*AUDIT_COLS, hidden=True)
    ws.column_dimensions.group(MODEL_COL, MODEL_COL, hidden=True)

    if opps:
        last = len(opps) + 1
        ws.conditional_formatting.add(
            f"Q2:Q{last}",   # Expected Value
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="num", mid_value=0, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))
        ws.conditional_formatting.add(
            f"R2:R{last}",   # Edge Now data bars
            DataBarRule(start_type="min", end_type="max", color="638EC6",
                        showValue=True))
        ws.conditional_formatting.add(
            f"X2:X{last}",   # Score
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color="63BE7B"))
        ws.conditional_formatting.add(
            f"U2:U{last}",   # Ann ROI - the capital-velocity heat
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color="8FAADC"))
        ws.conditional_formatting.add(
            f"L2:L{last}",   # Trend green/red
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=Font(color="006100")))
        ws.conditional_formatting.add(
            f"L2:L{last}",
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(color="9C0006")))


def _sorted(opps):
    return sorted(opps, key=lambda o: (o.listing.priority,
                                       o.valuation.opportunity_score),
                  reverse=True)


def _collapse(opps: list[Opportunity]) -> list[Opportunity]:
    """Same card + same grade from different sellers = ONE row (the best
    expected value; tie-break cheapest), with '+N more from $X' noted.
    Auctions and BINs collapse separately - an auction ending tonight and
    a BIN are different decisions. Titles are noisy, so grouping is fuzzy:
    same query + same effective grade + similar title + no variant
    conflict (holo/edition differences never merge)."""
    buckets: dict[tuple, list[list[Opportunity]]] = {}
    for o in opps:
        gi = grade_info(o.listing.title)
        key = (o.listing.query, o.listing.listing_type,
               gi[2] if gi else "raw")
        clusters = buckets.setdefault(key, [])
        subj_o = _subject_tokens(o.listing.title)
        for cluster in clusters:
            rep = cluster[0].listing.title
            subj_rep = _subject_tokens(rep)
            if subj_o and subj_rep and not (subj_o & subj_rep):
                continue    # different card entirely (Mewtwo vs Gyarados)
            if (title_match_score(rep, o.listing.title) >= 0.6
                    and not variant_conflict(rep, o.listing.title)):
                cluster.append(o)
                break
        else:
            clusters.append([o])
    out = []
    for clusters in buckets.values():
        for rows in clusters:
            if len(rows) == 1:
                out.append(rows[0])
                continue
            rows.sort(key=lambda o: (-o.valuation.expected_value,
                                     o.listing.total_cost_now))
            best = rows[0]
            others = rows[1:]
            costs = [o.listing.total_cost_now for o in others
                     if o.listing.total_cost_now]
            best.dupe_note = (f"+{len(others)} more listing(s) of this card"
                              + (f" from ${min(costs):,.0f}" if costs else ""))
            out.append(best)
    return out


def _grails_tab(wb, grail_rows: list[Opportunity]) -> None:
    """Personal-collection matches, by significance - NOT by profit."""
    ws = wb.create_sheet("Grails")
    cols = [("Grail", 34), ("Sig", 6), ("Type", 9), ("Site", 10),
            ("Title", 56), ("Price", 12), ("Ship", 9), ("Bids", 6),
            ("Timing", 16), ("Fair Value", 12), ("Notes", 34)]
    for col, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7B2D8E")   # grail purple
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    # group by grail, most significant first; within a grail, cheapest
    # first, capped so broad names can't flood the tab
    by_grail: dict[str, list[Opportunity]] = {}
    for o in grail_rows:
        by_grail.setdefault(o.listing.grail, []).append(o)
    ordered = sorted(by_grail.items(),
                     key=lambda kv: -kv[1][0].listing.grail_score)
    band = PatternFill("solid", fgColor="F7F2FA")
    r = 1
    for gname, rows in ordered:
        rows.sort(key=lambda o: o.listing.total_cost_now)
        for o in rows[:5]:
            r += 1
            l, v = o.listing, o.valuation
            ltype = ("BIN+OBO" if l.listing_type == "fixed" and l.best_offer
                     else "BIN" if l.listing_type == "fixed" else "AUCTION")
            timing, _hl = _timing(l)
            vals = [gname, l.grail_score, ltype, l.site, l.title,
                    l.current_price, l.shipping, l.bid_count, timing,
                    v.fair_value or None, "; ".join(v.notes)[:120]]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                if r % 2 == 0:
                    c.fill = band
            for col in (6, 7, 10):
                cell = ws.cell(row=r, column=col)
                if cell.value:
                    cell.number_format = _money_fmt(cell.value)
            t = ws.cell(row=r, column=5)
            if l.url:
                t.hyperlink = l.url
            t.font = Font(color="0563C1", underline="single")


def _bid_levels(o, config: dict) -> tuple[float | None, float | None]:
    """(max_bid, breakeven) - the two numbers a bid decision needs.

    breakeven is the price at which expected value hits exactly zero,
    net of shipping, sales tax and the vault-vs-taxed route. It is a WALL,
    not a target: winning there earns nothing, and fair value carries real
    uncertainty (confidence on these rows runs 26%-90%).

    max_bid is the price that still leaves `output.today.max_bid_target_roi`
    on the table - the number actually worth bidding to.
    """
    l, v = o.listing, o.valuation
    if v.fair_value <= 0:
        return None, None
    algo = config.get("algorithm", {})
    vault = algo.get("psa_vault", {}) or {}
    vault_on = vault.get("enabled", True) and l.site == "ebay"
    vault_min = vault.get("min_price", 500)
    target = (config.get("output", {}).get("today") or {}).get(
        "max_bid_target_roi", 0.15)
    tax = sales_tax_rate(config, l)
    resale = v.fair_value

    def _levels(net_proceeds: float, taxed: bool):
        route_tax = tax if taxed else 0.0
        be = item_price_for_total_cost(l, net_proceeds, route_tax)
        mb = item_price_for_total_cost(
            l, net_proceeds / (1 + max(target, 0.0)), route_tax)
        return mb, be

    normal_route = best_exit_route(config, l, resale, allow_vault=False)
    normal = _levels(normal_route.net_proceeds, taxed=True)
    if not vault_on:
        return tuple(math.floor(x) if x > 0 else None for x in normal)

    vault_route = best_exit_route(config, l, resale, allow_vault=True)
    vault_levels = _levels(vault_route.net_proceeds, taxed=False)

    def _best_valid(normal_value: float, vault_value: float):
        candidates = []
        # Normal checkout is valid strictly below the all-in vault boundary.
        # If its mathematical ceiling crosses the boundary, the highest
        # whole-dollar normal-route bid is the dollar immediately below it.
        normal_cap = l.item_price_for_landed_cost(vault_min)
        if normal_cap > 0:
            candidates.append(min(normal_value, normal_cap - 1e-9))
        # Vault economics are valid only when THIS candidate—not the
        # breakeven from the same route—actually reaches the threshold.
        if l.landed_cost(vault_value) >= vault_min:
            candidates.append(vault_value)
        valid = [value for value in candidates if value > 0]
        return math.floor(max(valid)) if valid else None

    return (_best_valid(normal[0], vault_levels[0]),
            _best_valid(normal[1], vault_levels[1]))


def _today_tab(wb, opps: list[Opportunity], config: dict) -> None:
    """The end-of-day decision list. Auctions ending within the window
    (sorted by END TIME - deadline order, not score) then fresh BINs
    (by score). Quality-floored: this tab is for deciding, not browsing -
    the full book lives in the category tabs."""
    cfg = (config.get("output", {}).get("today") or {})
    hours = cfg.get("hours", 24)
    fresh_h = cfg.get("fresh_hours", 24)
    min_ev = cfg.get("min_expected_value", 75)
    min_conf = cfg.get("min_confidence", 0.25)

    auctions, bins = [], []
    for o in opps:
        l, v = o.listing, o.valuation
        if not is_tradeable(o) or v.confidence < min_conf:
            continue
        if l.listing_type == "auction":
            hrs = l.hours_remaining
            if hrs is not None and 0 < hrs <= hours and v.expected_value >= min_ev:
                auctions.append(o)
        else:
            age = l.age_hours
            if age is not None and age <= fresh_h and v.edge_now >= min_ev:
                bins.append(o)
    auctions.sort(key=lambda o: o.listing.end_time)
    bins.sort(key=lambda o: -o.valuation.opportunity_score)
    rows = auctions + bins
    if not rows:
        return

    ws = wb.create_sheet("Today")
    cols = [("Decide", 9), ("Title", 56), ("Query", 26), ("Price", 11),
            ("Ship", 8), ("Bids", 6), ("Ends / Listed", 20),
            ("Fair Value", 12), ("Best Exit", 14), ("Net Proceeds", 13),
            ("vs eBay", 11), ("Max Bid", 11), ("Breakeven", 11),
            ("Expected Value", 13), ("Edge Now", 11), ("ROI", 8),
            ("Conf", 7), ("Notes", 44)]
    for col, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="B45309")   # decision amber
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    band = PatternFill("solid", fgColor="FDF6EC")
    for r, o in enumerate(rows, 2):
        l, v = o.listing, o.valuation
        ltype = ("BID?" if l.listing_type == "auction"
                 else "BUY?" if not l.best_offer else "OFFER?")
        timing, _hl = _timing(l)
        notes = "; ".join(v.notes + ([getattr(o, "dupe_note")]
                                     if getattr(o, "dupe_note", None) else []))
        max_bid, breakeven = _bid_levels(o, config)
        vals = [ltype, l.title, l.query, l.current_price, l.shipping,
                l.bid_count, timing, v.fair_value,
                _exit_name(v.resale_channel),
                v.net_proceeds, v.exit_advantage, max_bid, breakeven,
                v.expected_value, v.edge_now, v.roi, v.confidence, notes]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if r % 2 == 0:
                c.fill = band
        for col in (4, 5, 8, 10, 11, 12, 13, 14, 15):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.number_format = _money_fmt(cell.value)
        ws.cell(row=r, column=12).font = Font(bold=True)      # Max Bid pops
        ws.cell(row=r, column=13).font = Font(color="808080")  # Breakeven
        ws.cell(row=r, column=16).number_format = "0.0%"
        ws.cell(row=r, column=17).number_format = "0%"
        t = ws.cell(row=r, column=2)
        if l.url:
            t.hyperlink = l.url
            t.font = Font(color="0563C1", underline="single")


DEFAULT_CROSSOVER_TIERS = [[500, 25], [1500, 75], [2500, 150],
                           [10**9, 300]]      # PSA fee by declared value


def _crossover_tab(wb, opps: list[Opportunity], config: dict) -> None:
    """CGC/BGS/SGC/BVG cards priced cheap in PSA-equivalent terms:
    buy -> crack -> regrade to PSA (at the -1 shift grade) -> sell.
    Profit = edge_now (already PSA-equivalent, net of fees/tax) minus the
    PSA grading fee for the card's value tier. Assumes the shift holds
    (CGC 10 comes back PSA 9); a harsher comeback grade is the risk."""
    cfg = (config.get("algorithm", {}).get("crossover") or {})
    if not cfg.get("enabled", True):
        return
    tiers = cfg.get("fee_tiers") or DEFAULT_CROSSOVER_TIERS
    min_profit = cfg.get("min_profit", 100)
    allowed_graders = {str(g).lower() for g in
                       (cfg.get("allowed_graders")
                        or ["cgc", "bgs", "sgc", "bvg"])}
    allowed_categories = {str(cat) for cat in
                          (cfg.get("allowed_categories")
                           or ["Pokemon Cards", "Sports Cards"])}

    def grading_fee(value: float) -> float:
        for cap, fee in tiers:
            if value <= cap:
                return float(fee)
        return float(tiers[-1][1])

    rows = []
    for o in opps:
        l, v = o.listing, o.valuation
        if not is_tradeable(o):
            continue
        gi = grade_info(l.title)
        if not gi or gi[0].lower() not in allowed_graders:
            continue
        if v.fair_value <= 0 or v.disputed:
            continue
        if _category(l.query) not in allowed_categories:
            continue
        fee = grading_fee(v.fair_value)
        profit = v.edge_now - fee
        if profit >= min_profit:
            rows.append((profit, fee, gi, o))
    if not rows:
        return

    ws = wb.create_sheet("Crossover")
    cols = [("Type", 9), ("Site", 8), ("Title", 56), ("Query", 28),
            ("Grade", 9), ("As PSA", 8), ("Price", 12), ("Fair (PSA)", 12),
            ("Edge Now", 12), ("Grading Fee", 11), ("Regrade Profit", 13),
            ("ROI", 8), ("Timing", 15), ("Notes", 40)]
    for col, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0E7C7B")   # crossover teal
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    band = PatternFill("solid", fgColor="EFF7F7")
    rows.sort(key=lambda x: -x[0])
    for r, (profit, fee, gi, o) in enumerate(rows, 2):
        l, v = o.listing, o.valuation
        ltype = ("BIN+OBO" if l.listing_type == "fixed" and l.best_offer
                 else "BIN" if l.listing_type == "fixed" else "AUCTION")
        timing, _hl = _timing(l)
        cost = l.total_cost_now or 0
        roi = profit / cost if cost else None
        vals = [ltype, l.site, l.title, l.query,
                f"{gi[0].upper()} {gi[1]}", f"PSA {gi[2]}",
                l.current_price, v.fair_value, v.edge_now, fee,
                round(profit, 2), roi, timing,
                "; ".join(v.notes)[:120]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            if r % 2 == 0:
                c.fill = band
        for col in (7, 8, 9, 10, 11):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None:
                cell.number_format = _money_fmt(cell.value)
        ws.cell(row=r, column=12).number_format = "0%"
        t = ws.cell(row=r, column=3)
        if l.url:
            t.hyperlink = l.url
            t.font = Font(color="0563C1", underline="single")


def _portfolio_tab(wb, rows: list[dict]) -> None:
    """Positions marked to market - the P&L view."""
    ws = wb.create_sheet("Portfolio")
    cols = [("Status", 8), ("Description", 44), ("Query", 30),
            ("Bought", 11), ("Days", 7), ("Cost Basis", 12),
            ("Value", 12), ("Best Exit", 13), ("P&L", 12),
            ("Return", 9), ("CAGR", 9), ("Notes", 30)]
    for col, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    rows = sorted(rows, key=lambda r: (r["status"] != "OPEN",
                                       -(r["pnl"] or 0)))
    band = PatternFill("solid", fgColor="F5F7FA")
    r_i = 1
    tot_cost_open = tot_val_open = tot_real = 0.0
    for row in rows:
        r_i += 1
        ret = (row["pnl"] / row["cost"]
               if row["pnl"] is not None and row["cost"] else None)
        vals = [row["status"], row["description"], row["query"],
                row["bought"].strftime("%Y-%m-%d") if row["bought"] else "",
                row["days"], row["cost"], row["value"],
                _exit_name(str(row.get("exit_channel") or "")),
                row["pnl"], ret, row["cagr"], row["notes"]]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r_i, column=col, value=val)
            if r_i % 2 == 0:
                c.fill = band
        for col in (6, 7, 9):
            cell = ws.cell(row=r_i, column=col)
            if cell.value is not None:
                cell.number_format = _money_fmt(cell.value)
        for col in (10, 11):
            ws.cell(row=r_i, column=col).number_format = "0.0%"
        pnl = row["pnl"]
        if pnl is not None:
            ws.cell(row=r_i, column=9).font = Font(
                color="006100" if pnl >= 0 else "9C0006", bold=True)
        if row["status"] == "OPEN":
            tot_cost_open += row["cost"]
            if row["value"]:
                tot_val_open += row["value"]
        elif pnl is not None:
            tot_real += pnl
    # totals
    r_i += 2
    for label, val, money in [
            ("Open cost basis", tot_cost_open, True),
            ("Open market value (net)", tot_val_open, True),
            ("Unrealized P&L", tot_val_open - tot_cost_open, True),
            ("Realized P&L", tot_real, True)]:
        ws.cell(row=r_i, column=2, value=label).font = Font(bold=True)
        c = ws.cell(row=r_i, column=6, value=val)
        c.number_format = _money_fmt(val)
        c.font = Font(bold=True)
        r_i += 1


def _source_health_tab(wb, rows: list[dict]) -> None:
    """Per-run source readiness, including deliberately disabled sources."""
    if not rows:
        return
    ws = wb.create_sheet("Source Health")
    cols = [("Source", 28), ("Status", 12), ("OK", 9),
            ("Failed", 9), ("Skipped", 9), ("Freshness (h)", 14),
            ("Mode", 10), ("Checked", 21), ("Detail", 56)]
    for col, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="385723")
        ws.column_dimensions[get_column_letter(col)].width = width
    status_fill = {
        "healthy": "C6EFCE", "idle": "E2F0D9", "disabled": "D9E1F2",
        "degraded": "FFEB9C", "cooling": "F4B183",
        "failing": "FFC7CE", "stale": "FFC7CE", "empty": "FFC7CE",
    }
    for row_i, row in enumerate(rows, 2):
        checked = str(row.get("run_at") or "")
        if "T" in checked:
            checked = checked.replace("T", " ")[:19] + " UTC"
        values = [row.get("source"), row.get("status"), row.get("ok", 0),
                  row.get("failed", 0), row.get("skipped", 0),
                  row.get("freshness_hours"), row.get("mode"), checked,
                  row.get("detail")]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_i, column=col, value=value)
        status = str(row.get("status") or "").lower()
        ws.cell(row=row_i, column=2).fill = PatternFill(
            "solid", fgColor=status_fill.get(status, "E7E6E6"))
        if row.get("freshness_hours") is not None:
            ws.cell(row=row_i, column=6).number_format = "0.0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"


def _filter_waterfall_tab(wb, rows: list[dict]) -> None:
    """Visible count reconciliation from source hits to decision rows."""
    if not rows:
        return
    ws = wb.create_sheet("Filter Waterfall")
    cols = [("Stage", 27), ("Starting", 12), ("Removed", 12),
            ("Remaining", 12), ("Detail", 88)]
    header_fill = PatternFill("solid", fgColor="7F6000")
    for col, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    band = PatternFill("solid", fgColor="FFF8E7")
    warning = PatternFill("solid", fgColor="FCE4D6")
    for row_i, row in enumerate(rows, 2):
        values = [row.get("stage"), row.get("starting"),
                  row.get("removed"), row.get("remaining"),
                  row.get("detail")]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=value)
            if row_i % 2 == 0:
                cell.fill = band
        if (row.get("removed") or 0) > 0:
            ws.cell(row=row_i, column=3).fill = warning
            ws.cell(row=row_i, column=3).font = Font(
                bold=True, color="9C0006")
        ws.cell(row=row_i, column=5).alignment = Alignment(
            wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(rows) + 1}"


def _research_tab(wb, rows: list[dict]) -> None:
    """Every valued row rejected or quarantined, with its exact reason."""
    if not rows:
        return
    ws = wb.create_sheet("Research-Filtered")
    cols = [
        ("Stage", 24), ("Reason", 46), ("Type", 11), ("Site", 12),
        ("Title", 58), ("Query", 40), ("Price", 12), ("Ship", 10),
        ("Bids", 8), ("Timing", 18), ("Fair Value", 13),
        ("Exp Cost", 13), ("Expected Value", 15), ("Edge Now", 13),
        ("ROI", 10), ("Conf", 9), ("#Comps", 9), ("Best Exit", 14),
        ("Notes", 70), ("Model Detail", 70),
    ]
    header_fill = PatternFill("solid", fgColor="9C5700")
    for col, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:T{len(rows) + 1}"

    band = PatternFill("solid", fgColor="FFF4E6")
    quarantine = PatternFill("solid", fgColor="FFF2CC")
    for row_i, record in enumerate(rows, 2):
        o = record["opportunity"]
        l, v = o.listing, o.valuation
        ltype = ("BIN+OBO" if l.listing_type == "fixed" and l.best_offer
                 else "BIN" if l.listing_type == "fixed" else "AUCTION")
        timing, _ = _timing(l)
        values = [
            record.get("stage"), record.get("reason"), ltype, l.site,
            l.title, l.query, l.current_price, l.shipping, l.bid_count,
            timing, v.fair_value, v.expected_cost, v.expected_value,
            v.edge_now, v.roi, v.confidence, v.n_comps,
            _exit_name(v.resale_channel), "; ".join(v.notes),
            "; ".join(getattr(v, "audit_notes", [])),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=value)
            if row_i % 2 == 0:
                cell.fill = band
        if record.get("stage") == "Decision-only quarantine":
            for col in range(1, len(cols) + 1):
                ws.cell(row=row_i, column=col).fill = quarantine
        for col in (7, 8, 11, 12, 13, 14):
            ws.cell(row=row_i, column=col).number_format = _money_fmt(
                ws.cell(row=row_i, column=col).value or 0)
        ws.cell(row=row_i, column=15).number_format = "0.0%"
        ws.cell(row=row_i, column=16).number_format = "0%"
        title = ws.cell(row=row_i, column=5)
        if l.url:
            title.hyperlink = l.url
            title.font = Font(color="0563C1", underline="single")
        for col in (2, 19, 20):
            ws.cell(row=row_i, column=col).alignment = Alignment(
                wrap_text=True, vertical="top")

    last = len(rows) + 1
    ws.conditional_formatting.add(
        f"M2:M{last}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="num", mid_value=0, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))


def write_report(opps: list[Opportunity], path: str,
                 portfolio: list[dict] | None = None,
                 config: dict | None = None,
                 source_health: list[dict] | None = None,
                 research: list[dict] | None = None,
                 filter_waterfall: list[dict] | None = None) -> str:
    grail_rows = [o for o in opps if o.listing.grail]
    main = [o for o in opps if not o.listing.discovery and not o.listing.grail]
    disc = [o for o in opps
            if o.listing.discovery and not o.listing.grail]

    wb = Workbook()

    # Action tab: every positive-EV opportunity, plus the top-scored rows
    # and anything ending/fresh soon. Not artificially capped - if there
    # are 300 real deals, show all 300.
    # Watches are quarantined to their own tab: the valuation data there
    # (modifiers, box/papers, franken-watches) is less reliable than cards,
    # so they never mix into the Action view
    action_pool = [
        o for o in main
        if _category(o.listing.query) != "Watches" and is_tradeable(o)
    ]
    # same card from N sellers = one row ("+N more from $X" in Notes);
    # category tabs keep the full uncollapsed book
    action_pool = _collapse(action_pool)
    ranked = _sorted(action_pool)
    action, seen = [], set()
    for o in action_pool:
        hrs = o.listing.hours_remaining
        fresh = o.listing.age_hours
        soon = (hrs is not None and hrs <= 6) or (
            o.listing.listing_type == "fixed" and fresh is not None and fresh <= 6)
        if o.valuation.expected_value > 0 or (soon and o.valuation.edge_now > 0):
            action.append(o)
            seen.add(id(o))
    for o in ranked[:50]:            # always include the top of the book too
        if id(o) not in seen:
            action.append(o)
            seen.add(id(o))
    ws = wb.active
    ws.title = "Action"
    _fill_sheet(ws, _sorted(action))

    # Today: the end-of-day decision list, first tab in the book
    _today_tab(wb, action_pool, config or {})
    if "Today" in wb.sheetnames:
        wb.move_sheet("Today", offset=-len(wb.sheetnames) + 1)
        wb.active = wb["Today"]

    _filter_waterfall_tab(wb, filter_waterfall or [])
    _research_tab(wb, research or [])

    # category tabs
    by_cat: dict[str, list] = {}
    for o in main:
        by_cat.setdefault(_category(o.listing.query), []).append(o)
    for cat in CATEGORIES:
        if by_cat.get(cat):
            _fill_sheet(wb.create_sheet(cat), _sorted(by_cat[cat]))

    if grail_rows:
        _grails_tab(wb, grail_rows)

    _crossover_tab(wb, main, config or {})

    if portfolio:
        _portfolio_tab(wb, portfolio)

    if source_health:
        _source_health_tab(wb, source_health)

    if disc:
        _fill_sheet(wb.create_sheet("Discovery"), _sorted(disc))

    # Movers: unique queries by |trend|
    movers = {}
    for o in main:
        t = o.valuation.trend_30d
        if t is not None and o.listing.query not in movers:
            movers[o.listing.query] = (t, o.valuation.fair_value,
                                       o.valuation.n_comps)
    if movers:
        mws = wb.create_sheet("Movers")
        for col, (name, width) in enumerate(
                [("Query", 40), ("Trend 30d", 12), ("Fair Value", 14),
                 ("#Comps", 9)], 1):
            c = mws.cell(row=1, column=col, value=name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            mws.column_dimensions[get_column_letter(col)].width = width
        for i, (q, (t, fair, n)) in enumerate(
                sorted(movers.items(), key=lambda kv: abs(kv[1][0]),
                       reverse=True)[:30], 2):
            mws.cell(row=i, column=1, value=q)
            tc = mws.cell(row=i, column=2, value=t)
            tc.number_format = '"▲"0%;"▼"0%;"-"'
            tc.font = Font(color="006100" if t >= 0 else "9C0006")
            fc = mws.cell(row=i, column=3, value=fair)
            fc.number_format = _money_fmt(fair)
            mws.cell(row=i, column=4, value=n)
        mws.freeze_panes = "A2"

    meta = wb.create_sheet("About")
    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Today tab", "the end-of-day decision list: auctions ending within 24h in DEADLINE order, then fresh BINs. Floors: EV >= $75, conf >= 25% (output.today in config). Disputed, suspicious, mixed-pool and ask-based values are excluded"),
        ("Max Bid", "the highest ITEM bid/offer that still leaves your target return (output.today.max_bid_target_roi, default 15%) after adding shipping, buyer/proxy fees, international shipping, duty, FX, insurance, tax and the vault route"),
        ("Breakeven", "the price where profit is exactly zero. A wall, not a target - winning there earns nothing, and fair value is an estimate (see Conf)"),
        ("Action tab", "tradeable top-scored rows + anything ending/fresh within 6h with positive EV (watches excluded - see Watches tab); duplicate cards collapsed to the best listing"),
        ("Expected Value", "fair value net of resale fees, minus expected all-in cost"),
        ("Edge Now", "fair value net of exit-channel fees minus CURRENT landed cost (price, shipping, buyer/proxy fee, international shipping, insurance, duty, FX and tax where applicable)"),
        ("Best Exit", "eligible configured venue with the highest expected net proceeds; a watchlist or portfolio resale_channel can pin a manual override"),
        ("Net Proceeds", "fair value after the chosen exit's percentage/tiered fee and fixed selling costs"),
        ("vs eBay", "extra expected net proceeds from Best Exit compared with the configured eBay route"),
        ("Capture", "how capturable the edge is (auction: time left; BIN: freshness)"),
        ("Score", "ROI x Confidence x Capture - the sort key"),
        ("Targeted comps", "numbered cards discovered by broad searches are repriced from a separate sold pool for that exact card number and listing grade; thin exact pools remain browse-only"),
        ("Sales/mo", "how many of this card sell per month (comp velocity) - the liquidity dimension"),
        ("Ann ROI", "ROI annualized by turnover: same edge on a faster-trading card = higher capital velocity"),
        ("Pri (star)", "query names a specific grade; alerts enabled"),
        ("Crossover tab", "restricted to configured card categories and graders (default CGC/BGS/SGC/BVG Pokemon and sports cards). Profit = edge now minus the PSA grading tier; risk: the card comes back below the modeled shift grade"),
        ("Source Health", "this run's persisted source readiness: request and parser successes/failures, breaker skips, disabled sources and comp-cache freshness"),
        ("Filter Waterfall", "complete count reconciliation from raw marketplace hits through relevance, valuation, value/ROI/collection/economics rules and final kept rows"),
        ("Research-Filtered", "valued listings rejected before the main report plus browsable rows quarantined from Action/Today/alerts; Stage and Reason explain every row"),
        ("Hidden columns", "expand M-P for valuation inputs; Model Detail diagnostics are hidden in AD"),
        ("Grails tab", "personal-collection matches by significance (Sig 40-100), cheapest 5 per grail - NOT a profit view"),
        ("Discovery tab", "broad theme searches; values are mixed medians, NOT bid targets"),
        ("Movers tab", "30-day fair value changes (fills in as history accumulates)"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        meta[f"A{i}"], meta[f"B{i}"] = a, b
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 84

    wb.save(path)
    return path
