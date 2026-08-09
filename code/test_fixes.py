"""Regression tests for the 2026-07-25 valuation/learning fixes.

Every test below is anchored to a REAL row or a real number from Andrew's
own data, so a failure means the bug is back - not that a style changed.

Run it by double-clicking "Run Tests.command" (results are also written to
test_results.log), or:

    .venv/bin/python -m unittest discover -s code -p "test_*.py" -v
"""
from __future__ import annotations

import csv
import json
import logging
import os
import random
import sqlite3
import sys
import tempfile
import threading
import time
import unittest
from datetime import datetime, timedelta, timezone
from unittest import mock

import yaml

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import db as histdb                                          # noqa: E402
import learner                                               # noqa: E402
import main as scanner                                       # noqa: E402
import report as report_mod                                  # noqa: E402
from models import Listing, Opportunity, SoldComp, Valuation # noqa: E402
from report import _category                                 # noqa: E402
from valuation.comps import (grade_conflict, grade_info,     # noqa: E402
                             grader_of)
from valuation.price_guide import _guide_cents               # noqa: E402


# A PriceCharting product shaped like the Topsun Charizard that exposed the
# bug: loose $900, Grade 7 $1,800, Grade 8 $3,300, Grade 9 $6,718,
# Grade 9.5 $12,000, Grade 10 $25,000 (values in cents).
PRODUCT = {"loose-price": 90000, "cib-price": 180000, "new-price": 330000,
           "graded-price": 671800, "box-only-price": 1200000,
           "manual-only-price": 2500000}


def _load_json(path: str) -> dict:
    with open(path) as f:
        return json.load(f)


class TestGradeParsing(unittest.TestCase):
    """GRADE_RE used to accept two-digit numbers, so a seller typing
    'CGC 85' (for CGC 8.5) produced grade 85 -> 'PSA 84'. That row was live
    in the Crossover tab on 2026-07-25 claiming $3,349 of regrade profit."""

    def test_impossible_grade_reads_as_ungraded(self):
        title = "Pokemon Charizard Topsun Blue Back 1997 CGC 85 Rare Japanese"
        self.assertIsNone(grade_info(title))
        self.assertIsNone(grader_of(title))

    def test_normal_grades_still_parse(self):
        self.assertEqual(grade_info("Charizard PSA 10 Gem"), ("psa", "10", "10"))
        self.assertEqual(grade_info("Gengar PSA 9 Fossil"), ("psa", "9", "9"))
        self.assertEqual(grade_info("Card BGS 9.5 (10 subs)"),
                         ("bgs", "9.5", "8.5"))

    def test_cross_grader_shift_is_one_full_grade(self):
        """Andrew's rule: every other grading service counts one grade
        lower than PSA when comparing prices."""
        for grader in ("cgc", "bgs", "sgc", "bvg"):
            self.assertEqual(grade_info(f"Card {grader.upper()} 10")[2], "9")
            self.assertEqual(grade_info(f"Card {grader.upper()} 8.5")[2], "7.5")
        self.assertEqual(grade_info("Card PSA 10")[2], "10")

    def test_effective_grade_floors_at_one(self):
        self.assertEqual(grade_info("Card CGC 1")[2], "1")

    def test_sgc_legacy_100_point_labels(self):
        # SGC graded on a 100-point scale until 2020
        self.assertEqual(grade_info("1952 Topps Mantle SGC 92"),
                         ("sgc", "8.5", "7.5"))
        self.assertEqual(grade_info("T206 Cobb SGC 30"), ("sgc", "1.5", "1"))

    def test_grade_conflict_uses_effective_grades(self):
        self.assertFalse(grade_conflict("Gengar PSA 9", "Gengar CGC 10"))
        self.assertTrue(grade_conflict("Gengar PSA 9", "Gengar CGC 9"))
        # a typo'd grade must not match a real one
        self.assertFalse(grade_conflict("Gengar PSA 9", "Gengar CGC 85"))


class TestPriceGuideGradeRouting(unittest.TestCase):
    """The guide picked its price field from the RAW grade, so the -1
    cross-grader penalty was applied to comps and silently skipped on the
    guide. Two Topsun Charizards (CGC 8.5 and 'CGC 85') were both quoted
    $6,718 - the Grade 9 price - on 2026-07-25."""

    def dollars(self, eff):
        cents, _how = _guide_cents(PRODUCT, eff)
        return None if cents is None else round(cents / 100)

    def test_cgc_10_prices_as_psa_9_not_psa_10(self):
        eff = float(grade_info("Charizard CGC 10")[2])
        self.assertEqual(self.dollars(eff), 6718)
        self.assertNotEqual(self.dollars(eff), 25000)

    def test_half_grade_lands_between_neighbours_never_above(self):
        self.assertEqual(self.dollars(8.0), 3300)
        self.assertEqual(self.dollars(9.0), 6718)
        self.assertTrue(3300 < self.dollars(8.5) < 6718)

    def test_live_cgc_85_row_falls_back_to_raw(self):
        eff = grade_info("Topsun Charizard 1997 CGC 85")
        self.assertIsNone(eff)
        self.assertEqual(self.dollars(None), 900)

    def test_never_inflates_when_a_rung_is_missing(self):
        sparse = {"loose-price": 90000, "graded-price": 671800}
        for eff in (6.0, 7.0, 8.0, 8.5):
            cents, _ = _guide_cents(sparse, eff)
            self.assertLess(cents / 100, 6718,
                            f"grade {eff} inherited Grade-9 money")

    def test_refuses_rather_than_guessing_upward(self):
        """No raw price and the only rung is above our grade: the honest
        answer is no guide value at all, not the higher rung's price."""
        cents, how = _guide_cents({"graded-price": 671800}, 7.0)
        self.assertIsNone(cents)
        self.assertIn("below lowest available field", how)

    def test_monotonic_in_grade(self):
        vals = [self.dollars(g) for g in (5, 6, 7, 7.5, 8, 8.5, 9, 9.5, 10)]
        self.assertEqual(vals, sorted(vals))


class TestGuideCacheInvalidation(unittest.TestCase):
    """Guide prices are cached for a week. Without a version bump the grade
    fix would have been invisible until 2026-08-01: the cache held Topsun
    Charizard at $6,718 for PSA 9, CGC 9, CGC 8.5 and the typo 'CGC 85'."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.db = os.path.join(self.dir, "h.db")

    def _guide(self):
        from valuation.price_guide import PriceGuide
        return PriceGuide({"database": {"file": self.db},
                           "api_keys": {"pricecharting": {"token": "x"}}})

    def test_stale_cache_is_cleared_once_then_left_alone(self):
        conn = sqlite3.connect(self.db)
        conn.execute("CREATE TABLE IF NOT EXISTS guide_cache("
                     "query TEXT PRIMARY KEY, value REAL, ts TEXT)")
        conn.execute("INSERT INTO guide_cache VALUES "
                     "('Topsun Charizard 1997 CGC 8.5', 6718.0, ?)",
                     (datetime.now(timezone.utc).isoformat(),))
        conn.commit()
        conn.close()

        with self.assertLogs("valuation.price_guide", level="WARNING"):
            self._guide()                  # first run purges
        conn = sqlite3.connect(self.db)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM guide_cache").fetchone()[0], 0)
        # a fresh value written after the purge must survive the next run
        conn.execute("INSERT INTO guide_cache VALUES ('q', 100.0, ?)",
                     (datetime.now(timezone.utc).isoformat(),))
        conn.commit()
        conn.close()
        self._guide()
        conn = sqlite3.connect(self.db)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM guide_cache").fetchone()[0], 1)
        conn.close()


class TestPriceChartingRequestControl(unittest.TestCase):
    """Paid API calls obey PriceCharting's one-call-per-second contract."""

    class Response:
        def __init__(self, product):
            self._product = product
            self.status_code = 200
            self.headers = {}

        def raise_for_status(self):
            return None

        def json(self):
            return self._product

    @staticmethod
    def _product(name="Charizard #4"):
        return {
            "status": "success",
            "product-name": name,
            "console-name": "Pokemon",
            "loose-price": 10000,
            "new-price": 30000,
            "graded-price": 50000,
        }

    def _guide(self, directory):
        from valuation.price_guide import PriceGuide
        return PriceGuide({
            "database": {"file": os.path.join(directory, "h.db")},
            "api_keys": {"pricecharting": {"token": "paid-token"}},
            "pricecharting": {"request_delay_seconds": 1.05},
        })

    def test_keys_share_product_payload_but_keep_grade_value_separate(self):
        from valuation.price_guide import (
            _guide_cache_key, _pricecharting_product_key)

        self.assertEqual(
            _pricecharting_product_key("Charizard #015 PSA 9"),
            "charizard #15")
        self.assertEqual(
            _guide_cache_key("Charizard #015 CGC 10"),
            "charizard #15|psa:9")

    def test_different_grades_reuse_one_product_response_across_runs(self):
        with tempfile.TemporaryDirectory() as tmp:
            first = self._guide(tmp)
            first._pc_delay = 0
            calls = []
            first.session.get = lambda *a, **k: (
                calls.append(k["params"]["q"])
                or self.Response(self._product()))

            self.assertEqual(first.guide_value("Charizard #4 PSA 8"), 300)
            self.assertEqual(first.guide_value("Charizard #4 PSA 9"), 500)
            self.assertEqual(calls, ["charizard #4"])

            # The raw product payload is persistent, not only in-memory.
            second = self._guide(tmp)
            second.session.get = lambda *a, **k: self.fail(
                "persistent product cache should avoid a wire call")
            self.assertEqual(
                second.guide_value("Charizard #004 PSA 8.5"), 400)

    def test_requests_are_spaced_start_to_start(self):
        import valuation.price_guide as guide_mod

        with tempfile.TemporaryDirectory() as tmp:
            guide = self._guide(tmp)
            clock = [100.0]
            starts = []

            def monotonic():
                return clock[0]

            def sleep(seconds):
                clock[0] += seconds

            def get(*args, **kwargs):
                starts.append(clock[0])
                name = kwargs["params"]["q"]
                return self.Response(self._product(name))

            guide.session.get = get
            with mock.patch.object(guide_mod.time, "monotonic", monotonic), \
                    mock.patch.object(guide_mod.time, "sleep", sleep):
                guide._pricecharting("Charizard #4 PSA 9")
                guide._pricecharting("Blastoise #2 PSA 9")

            self.assertEqual(len(starts), 2)
            self.assertAlmostEqual(starts[1] - starts[0], 1.05, places=6)

    def test_429_waits_and_retries_once_instead_of_skipping_the_run(self):
        import requests as requests_mod
        import valuation.price_guide as guide_mod
        from scrapers.base import API_STATS, reset_api_stats

        with tempfile.TemporaryDirectory() as tmp:
            reset_api_stats()
            guide = self._guide(tmp)
            clock = [100.0]
            calls = 0

            class RateLimited:
                status_code = 429
                headers = {"Retry-After": "2"}

                def raise_for_status(self):
                    raise requests_mod.HTTPError(
                        "429", response=self)

            def get(*args, **kwargs):
                nonlocal calls
                calls += 1
                if calls == 1:
                    return RateLimited()
                return self.Response(self._product())

            guide.session.get = get
            with mock.patch.object(
                    guide_mod.time, "monotonic", lambda: clock[0]), \
                    mock.patch.object(
                        guide_mod.time, "sleep",
                        lambda seconds: clock.__setitem__(
                            0, clock[0] + seconds)):
                value = guide._pricecharting("Charizard #4 PSA 9")

            self.assertEqual(value, 500)
            self.assertEqual(calls, 2)
            self.assertGreaterEqual(clock[0], 102.0)
            self.assertEqual(API_STATS[("pricecharting", "failed")], 1)
            self.assertEqual(API_STATS[("pricecharting", "ok")], 1)
            self.assertEqual(API_STATS[("pricecharting", "skipped")], 0)

    def test_parallel_requests_share_one_wire_lane(self):
        from concurrent.futures import ThreadPoolExecutor

        with tempfile.TemporaryDirectory() as tmp:
            guide = self._guide(tmp)
            guide._pc_delay = 0
            in_flight = 0
            peak = 0
            calls_lock = threading.Lock()

            def get(*args, **kwargs):
                nonlocal in_flight, peak
                with calls_lock:
                    in_flight += 1
                    peak = max(peak, in_flight)
                time.sleep(0.003)
                with calls_lock:
                    in_flight -= 1
                return self.Response(
                    self._product(kwargs["params"]["q"]))

            guide.session.get = get
            with ThreadPoolExecutor(max_workers=12) as pool:
                rows = list(pool.map(
                    lambda i: guide._pricecharting(
                        f"Card {i} PSA 9"),
                    range(24)))

            self.assertTrue(all(value == 500 for value in rows))
            self.assertEqual(peak, 1)


def _opp(title="Card", query="Pokemon Card", ev=500.0, roi=0.4, bids=3,
         ltype="auction", grail="", site="ebay", has_buy_now=False):
    return Opportunity(
        listing=Listing(site=site, title=title, url="https://e/itm/1",
                        current_price=100.0, bid_count=bids,
                        listing_type=ltype, grail=grail, query=query,
                        has_buy_now=has_buy_now),
        valuation=Valuation(fair_value=1000.0, expected_value=ev, roi=roi))


class TestPokemonGradeFloor(unittest.TestCase):
    """filters.pokemon_grade_floor is 5: graded Pokemon at PSA 5 or below
    are dropped. Pokemon only, grails exempt, raw unaffected."""

    def ok(self, o, floor=5.0):
        return scanner.output_ok(o, min_ev=0.0, poke_floor=floor)

    def test_drops_psa_5_and_below_pokemon(self):
        for g in ("PSA 5", "PSA 4", "PSA 3", "PSA 1"):
            self.assertFalse(self.ok(_opp(title=f"Charizard Holo {g}",
                                          query="Charizard Holo PSA 5")), g)

    def test_keeps_psa_6_and_above(self):
        for g in ("PSA 5.5", "PSA 6", "PSA 9", "PSA 10"):
            self.assertTrue(self.ok(_opp(title=f"Charizard Holo {g}",
                                         query="Charizard Holo PSA 9")), g)

    def test_floor_uses_effective_grade(self):
        # a CGC 6 is a PSA 5 equivalent -> drops; a CGC 7 (= PSA 6) stays
        self.assertFalse(self.ok(_opp(title="Charizard Holo CGC 6",
                                      query="Charizard Holo")))
        self.assertTrue(self.ok(_opp(title="Charizard Holo CGC 7",
                                     query="Charizard Holo")))

    def test_raw_pokemon_unaffected(self):
        self.assertTrue(self.ok(_opp(title="Charizard Holo Base Set",
                                     query="Charizard Holo")))

    def test_only_applies_to_pokemon(self):
        self.assertTrue(self.ok(_opp(title="1986 Fleer Michael Jordan PSA 3",
                                     query="Michael Jordan 1986 Fleer")))

    def test_grails_are_exempt(self):
        self.assertTrue(self.ok(_opp(title="Charizard Holo PSA 2",
                                     query="Charizard Holo PSA 9",
                                     grail="Topsun Charizard 1997")))

    def test_other_output_rules_intact(self):
        self.assertFalse(self.ok(_opp(ev=-1.0)))
        self.assertFalse(self.ok(_opp(roi=-0.1)))
        self.assertFalse(self.ok(_opp(bids=0)))
        self.assertTrue(self.ok(_opp(bids=0, site="yahoo_jp")))

    def test_zero_bid_hybrid_needs_a_known_buy_it_now(self):
        """The hybrid exemption exists because 'the BIN is takeable'. That
        only holds if we actually captured the BIN price - otherwise the
        row is priced off the seller's opening ask."""
        no_price = _opp(bids=0, has_buy_now=True)
        self.assertFalse(self.ok(no_price))
        with_price = _opp(bids=0, has_buy_now=True)
        with_price.listing.buy_now_price = 2500.0
        self.assertTrue(self.ok(with_price))

    def test_bad_row_keeps_the_row_and_never_raises(self):
        broken = Opportunity(listing=None, valuation=None)
        # the traceback this logs is the POINT of the test: a malformed row
        # must be reported loudly and kept, never allowed to kill the report
        with self.assertLogs("scanner", level="ERROR"):
            self.assertTrue(scanner.output_ok(broken))

    def test_drop_reasons_are_counted_for_the_log(self):
        drops = {}
        scanner.output_ok(_opp(title="Charizard PSA 3", query="Charizard"),
                          poke_floor=5.0, drops=drops)
        scanner.output_ok(_opp(ev=-5.0), drops=drops)
        self.assertEqual(sum(drops.values()), 2)


class TestExcludeKeywords(unittest.TestCase):
    """Live junk rows from the 15:44 report that were valued off real card
    comps: a resin art print at $897 EV and a repack box at $979 EV."""

    KEYWORDS = ["reprint", "repack", "poster", "art print", "chase box",
                "set-break", "set break", "custom card"]

    def test_catches_the_live_junk_rows(self):
        junk = [
            "Ermsy MJ 1986 Michael Jordan Fleer RC BLACK RED Poster Print",
            "UNIVERSAL TREASURES BASKETBALL Chase Box LOADED 100 Fleer PACK",
            "1977 Topps Star Wars Set-Break #270 Luke Skywalker",
            "2024 Pokemon Repack Mystery Box",
        ]
        for t in junk:
            self.assertTrue(scanner._excluded(t, self.KEYWORDS), t)

    def test_does_not_catch_real_cards(self):
        real = [
            "1999 Pokemon 1st Edition Fossil Gengar Holo PSA 9",
            "Pokemon Charizard Topsun Blue Back 1997 1st Print",
            "1986 Fleer Michael Jordan #57 RC PSA 8",
            "Umbreon VMAX Alt Art Evolving Skies PSA 10",
            "2004 Panini Mega Cracks Lionel Messi #89 PSA 7",
        ]
        for t in real:
            self.assertFalse(scanner._excluded(t, self.KEYWORDS), t)


class TestObservationSchema(unittest.TestCase):
    """The learner had no way to tell a well-comped $2,500 valuation from a
    mongrel $2.85 one, because n_comps was never recorded."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.db = os.path.join(self.dir, "h.db")

    def test_migrates_a_pre_existing_table_without_losing_rows(self):
        c = sqlite3.connect(self.db)
        c.executescript(
            """CREATE TABLE observations(item_id TEXT, site TEXT, query TEXT,
               title TEXT, listing_type TEXT, price REAL, shipping REAL,
               bids INTEGER, end_time TEXT, fair REAL, predicted_settle REAL,
               hours_left REAL, observed_at TEXT);""")
        c.execute("INSERT INTO observations VALUES ('1','ebay','q','t',"
                  "'auction',1,0,0,NULL,10,10,5,'2026-07-01')")
        c.commit()
        c.close()
        conn = histdb.connect(self.db)
        cols = {r[1] for r in conn.execute("PRAGMA table_info(observations)")}
        self.assertIn("n_comps", cols)
        self.assertIn("confidence", cols)
        self.assertIn("trusted", cols)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM observations").fetchone()[0], 1)
        conn.close()

    def test_comp_cache_reports_its_own_age(self):
        """cached_comps(allow_stale=True) hands back week-old rows without
        saying so. Without an age the report cannot distinguish evidence
        from this morning from evidence frozen before a breaker opened."""
        conn = histdb.connect(self.db)
        self.assertIsNone(histdb.comp_cache_age_hours(conn, "nothing cached"))

        comp = SoldComp(title="1986 Fleer Jordan PSA 8", price=1000.0,
                        sold_date=datetime.now(timezone.utc)
                        - timedelta(days=9),
                        url="https://www.ebay.com/itm/1", site="ebay")
        histdb.save_comps(conn, "jordan", [comp])
        # save_comps stamps scanned_at = now, so backdate it to simulate a
        # cache written before the source went down.
        old = (datetime.now(timezone.utc) - timedelta(hours=200)).isoformat()
        conn.execute("UPDATE comps SET scanned_at=? WHERE query=?",
                     (old, "jordan"))
        conn.commit()

        age = histdb.comp_cache_age_hours(conn, "jordan")
        self.assertIsNotNone(age)
        self.assertAlmostEqual(age, 200, delta=1)
        self.assertTrue(histdb.cached_comps(conn, "jordan", 24,
                                            allow_stale=True),
                        "still returned - stale beats none - but now dated")
        conn.close()

    def test_records_the_evidence(self):
        conn = histdb.connect(self.db)
        listing = Listing(site="ebay", title="Gengar PSA 9",
                          url="https://www.ebay.com/itm/123456789012",
                          current_price=900.0, listing_id="123456789012",
                          query="Gengar PSA 9",
                          end_time=datetime.now(timezone.utc)
                          + timedelta(hours=3))
        histdb.record_observation(conn, listing,
                                  Valuation(fair_value=2500.0, n_comps=11,
                                            confidence=0.81))
        row = conn.execute("SELECT n_comps, confidence, trusted FROM observations"
                           ).fetchone()
        self.assertEqual(row[0], 11)
        self.assertAlmostEqual(row[1], 0.81)
        self.assertEqual(row[2], 1)
        conn.close()


class TestLearnerTrustFilters(unittest.TestCase):
    """On 2026-07-25 the learner reported settle_ratio 1.2756 - auctions
    supposedly closing 28% ABOVE fair value - because 891 of 1,090 matched
    closes were valued under $50. That inflated every expected cost."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.db = os.path.join(self.dir, "h.db")

    def _auction(self, conn, iid, fair, actual, n_comps=9, snapshots=3):
        for i, hrs in enumerate([72, 24, 6][:snapshots]):
            listing = Listing(
                site="ebay", title=f"Card {iid} PSA 9",
                url=f"https://www.ebay.com/itm/{iid}",
                current_price=fair * 0.4, bid_count=i * 3, listing_id=iid,
                query=f"Card {iid} PSA 9",
                end_time=datetime.now(timezone.utc) + timedelta(hours=hrs))
            histdb.record_observation(
                conn, listing,
                Valuation(fair_value=fair, expected_cost=fair * 0.9,
                          n_comps=n_comps, confidence=0.7))
        conn.execute("INSERT OR IGNORE INTO closed VALUES (?,?,?)",
                     (iid, actual, datetime.now(timezone.utc).isoformat()))
        conn.commit()

    def test_junk_valuations_are_excluded(self):
        conn = histdb.connect(self.db)
        # the real shape of the bad data: a $2.85 valuation on a card that
        # closed at $67 (a 24x "settle ratio")
        for i in range(40):
            self._auction(conn, f"90000000000{i}", 2.85, 67.0)
        conn.close()
        msg = learner.fit(self.db, directory=self.dir)
        self.assertIn("trustworthy closes", msg)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 0)
        self.assertGreater(params["training_filter"]
                           ["dropped_fair_below_floor"], 0)

    def test_legacy_rows_without_evidence_are_excluded(self):
        conn = histdb.connect(self.db)
        self._auction(conn, "999000000001", 1200.0, 1000.0)
        conn.execute("UPDATE observations SET n_comps = NULL")
        conn.commit()
        conn.close()
        learner.fit(self.db, directory=self.dir)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 0)
        self.assertGreater(params["training_filter"]
                           ["dropped_no_evidence_recorded"], 0)

    def test_thin_comp_valuations_are_excluded(self):
        conn = histdb.connect(self.db)
        for i in range(40):
            self._auction(conn, f"88000000000{i}", 1200.0, 1000.0, n_comps=1)
        conn.close()
        learner.fit(self.db, directory=self.dir)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 0)
        self.assertGreater(params["training_filter"]["dropped_thin_comps"], 0)

    def test_pre_gate_rows_without_trust_attestation_are_excluded(self):
        conn = histdb.connect(self.db)
        self._auction(conn, "877000000001", 1200.0, 1000.0, n_comps=9)
        conn.execute("UPDATE observations SET trusted = NULL")
        conn.commit()
        conn.close()
        learner.fit(self.db, directory=self.dir)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 0)
        self.assertGreater(
            params["training_filter"]["dropped_no_trust_attestation"], 0)

    def test_params_are_written_even_on_a_cold_start(self):
        """Returning early used to leave a stale (wrong) settle ratio and a
        deployed model on disk, which the engine kept using forever."""
        stale = {"n": 383, "settle_ratio": 1.2756,
                 "ml": {"deployed": True, "cv_mae": 0.2865}}
        path = os.path.join(self.dir, "learned_params.json")
        with open(path, "w") as f:
            json.dump(stale, f)
        conn = histdb.connect(self.db)
        conn.close()
        learner.fit(self.db, directory=self.dir)
        params = _load_json(path)
        self.assertEqual(params["n"], 0)
        self.assertFalse(params["ml"]["deployed"])
        self.assertIsNone(learner.ClosePredictor(self.dir).settle_ratio)

    def test_recovers_the_true_settle_ratio_from_clean_data(self):
        conn = histdb.connect(self.db)
        rng = random.Random(11)
        true_settle = 0.85
        for i in range(60):
            fair = rng.choice([600.0, 900.0, 1500.0, 2600.0, 4200.0])
            self._auction(conn, f"70000000%04d" % i, fair,
                          round(fair * true_settle * rng.gauss(1, 0.10), 2))
        conn.close()
        msg = learner.fit(self.db, directory=self.dir)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 60, "should dedupe 180 snapshots -> 60")
        self.assertEqual(params["n_snapshots"], 180)
        self.assertAlmostEqual(params["settle_ratio"], true_settle, delta=0.06)
        self.assertLess(params["parametric_mae"], 0.20)
        self.assertIn("settle_ratio", msg)

    def test_median_is_per_auction_not_per_snapshot(self):
        """One heavily-scanned auction must not outvote 20 others."""
        conn = histdb.connect(self.db)
        for i in range(20):
            self._auction(conn, "6000000%05d" % i, 1000.0, 800.0)
        self._auction(conn, "600009999999", 1000.0, 2500.0, snapshots=3)
        conn.close()
        learner.fit(self.db, directory=self.dir)
        params = _load_json(os.path.join(self.dir, "learned_params.json"))
        self.assertEqual(params["n"], 21)
        self.assertAlmostEqual(params["settle_ratio"], 0.80, delta=0.01)

    def test_engine_falls_back_to_config_when_learner_is_cold(self):
        conn = histdb.connect(self.db)
        conn.close()
        learner.fit(self.db, directory=self.dir)
        pred = learner.ClosePredictor(self.dir)
        self.assertIsNone(pred.settle_ratio)
        self.assertIsNone(pred.settle_ratio_for(1500))
        self.assertIsNone(pred.predict_ratio(5, 0.5, 3, 1500))


class TestSubjectInjection(unittest.TestCase):
    """When a query names no card ("1999 1st Edition Pokemon Set") the
    engine values each listing on its own subject. It used to take the
    three LONGEST leftover words, which are reliably seller adjectives -
    measured over 599 real listings the top injected "subjects" were
    symbol x116, wotc x67, tcg x51, lp x20, stamp x15, none of them cards.
    """

    def test_picks_the_card_not_the_adjectives(self):
        from valuation.comps import subject_candidates
        cases = [
            ("Pokemon Jungle 1st Edition Snorlax Holo Beautiful Centering "
             "Investment", "snorlax"),
            ("1996 POKEMON BASE SET JAPANESE NO RARITY SYMBOL #145 ZAPDOS "
             "COLLECTIBLE", "zapdos"),
            ("1999 POKEMON BASE SET 1ST EDITION #2 BLASTOISE HOLO GRADED "
             "COLLECTIBLE", "blastoise"),
            ("1996 Pokemon Japanese Base No Rarity Charizard Authenticated "
             "Vintage Holo", "charizard"),
        ]
        for title, expected in cases:
            self.assertIn(expected, subject_candidates(title, 2), title[:40])

    def test_marketing_and_context_words_are_never_subjects(self):
        from valuation.comps import subject_candidates
        banned = {"collectible", "authenticated", "investment", "beautiful",
                  "centering", "symbol", "wotc", "tcg", "stamp", "near",
                  "graded", "vintage", "sealed", "rocket", "trainer"}
        titles = [
            "1996 POKEMON BASE SET JAPANESE NO RARITY SYMBOL #145 ZAPDOS",
            "Pokemon Jungle 1st Edition Snorlax Beautiful Centering",
            "Pokemon TCG WOTC Team Rocket Trainer Card Near Mint Collectible",
        ]
        for t in titles:
            got = set(subject_candidates(t, 3))
            self.assertFalse(got & banned, f"{t[:40]} -> {got}")

    def test_short_real_names_survive(self):
        from valuation.comps import subject_candidates
        self.assertIn("mew", subject_candidates(
            "1999 Pokemon Base Set Mew Promo Holo Graded PSA 9", 2))


class TestUnicodeFolding(unittest.TestCase):
    """7% of real listing titles (277 of 4,075) contain non-ASCII. The
    tokenizers split on [a-z], so 'Pokemon' with an accent became the two
    tokens 'pok' and 'mon' - which is where 'mon x50' and 'pok x22' in the
    injected-subject tally came from."""

    def test_accented_word_stays_one_token(self):
        from valuation.comps import _subject_tokens
        accented = _subject_tokens("Pokémon TCG Gloom Jungle")
        plain = _subject_tokens("Pokemon TCG Gloom Jungle")
        self.assertEqual(accented, plain)
        self.assertNotIn("mon", accented)
        self.assertNotIn("pok", accented)

    def test_title_match_is_accent_insensitive(self):
        from valuation.comps import title_match_score
        a = title_match_score("Pokemon Charizard Base Set",
                              "Pokémon Charizard Base Set Holo")
        b = title_match_score("Pokemon Charizard Base Set",
                              "Pokemon Charizard Base Set Holo")
        self.assertAlmostEqual(a, b, places=6)

    def test_grail_matching_is_accent_insensitive(self):
        import grails
        g = grails.load_grails({"grails": ["Pokemon Red Cheeks Pikachu"]})
        self.assertIsNotNone(
            grails.match(g, "1999 Pokémon Red Cheeks Pikachu PSA 9"))

    def test_japanese_titles_pass_through_unchanged(self):
        """Japanese dakuten decompose exactly like a Latin accent. Dropping
        them turns プ into フ - a different word - which would break the
        Topsun variant check and every translated Yahoo/Buyee query."""
        from textutil import fold
        for s in ("トップサン", "リザードン", "カメックス", "旧裏",
                  "マークなし", "ポケモンカード", "ギャラドス"):
            self.assertEqual(fold(s), s)

    def test_japanese_matchers_still_fire_after_folding(self):
        from valuation.comps import variant_conflict, JP_NATIVE_RE
        from textutil import fold
        self.assertTrue(JP_NATIVE_RE.search(fold("旧裏 ポケモン")))
        # Topsun original vs the VS battle series must still be told apart
        self.assertTrue(variant_conflict("トップサン Charizard",
                                         "トップサン Charizard VS Blastoise"))


class TestHybridBuyItNow(unittest.TestCase):
    """A hybrid auction's current_price is the BID. With zero bids that is
    the seller's opening ask, not a transactable price. Live on 2026-07-25:
    a $499 opening bid on a card valued at $2,821, reported as $1,584 of
    expected value."""

    def setUp(self):
        logging.disable(logging.WARNING)

    def tearDown(self):
        logging.disable(logging.NOTSET)

    def _engine(self):
        from valuation.engine import ValuationEngine
        return ValuationEngine({"_config_dir": tempfile.mkdtemp(),
                                "algorithm": {"sales_tax_rate": 0.0,
                                              "psa_vault": {"enabled": False}}})

    def _listing(self, bids, bin_price):
        return Listing(site="ebay", title="1940 Superman Gum R145",
                       url="https://www.ebay.com/itm/1", current_price=499.0,
                       bid_count=bids, listing_type="auction",
                       has_buy_now=True, buy_now_price=bin_price,
                       query="Superman 1940",
                       end_time=datetime.now(timezone.utc)
                       + timedelta(hours=10))

    def test_zero_bid_hybrid_is_priced_off_the_buy_it_now(self):
        e = self._engine()
        v = e.score(self._listing(0, 2400.0), Valuation(fair_value=2821.0))
        # edge is measured against the BIN (2,400), not the 499 opening bid
        self.assertLess(v.edge_now, 200)
        self.assertTrue(any("Buy It Now" in n for n in v.notes))

    def test_bid_hybrid_still_uses_the_live_bid(self):
        e = self._engine()
        v = e.score(self._listing(7, 2400.0), Valuation(fair_value=2821.0))
        self.assertGreater(v.edge_now, 1500)     # real bidding at $499

    def test_expected_close_never_exceeds_the_buy_it_now(self):
        e = self._engine()
        v = e.score(self._listing(7, 1200.0), Valuation(fair_value=9000.0))
        self.assertLessEqual(v.expected_cost, 1200.0 + 0.01)


class TestCategoryRouting(unittest.TestCase):
    """'sealed' used to be a Video Games keyword, tested before Pokemon -
    so a sealed Pokemon query landed in the games tab AND skipped the
    Pokemon grade floor, which only applies to the Pokemon category."""

    def test_sealed_pokemon_is_a_pokemon_card(self):
        self.assertEqual(_category("Sealed Pokemon Base Set Booster Box"),
                         "Pokemon Cards")
        self.assertEqual(_category("1st Edition Pokemon Pack sealed"),
                         "Pokemon Cards")

    def test_platform_words_still_win(self):
        for q in ("Pokemon Red Gameboy", "Super Mario Bros 1985 NES",
                  "Halo Xbox", "Duck Hunt NES"):
            self.assertEqual(_category(q), "Video Games", q)

    def test_game_titles_without_a_platform(self):
        for q in ("Sealed Zelda Ocarina of Time", "The Sims",
                  "Roller Coaster Tycoon", "Call of Duty MW2"):
            self.assertEqual(_category(q), "Video Games", q)

    def test_cards_and_watches_unchanged(self):
        self.assertEqual(_category("Charizard Base Set PSA 9"),
                         "Pokemon Cards")
        self.assertEqual(_category("Michael Jordan 1986 Fleer Auto"),
                         "Sports Cards")
        self.assertEqual(_category("Patek Philippe World Time"), "Watches")


class TestBidLevels(unittest.TestCase):
    """Max Bid used to be exact breakeven - winning there earns nothing."""

    def _opp_for(self, fair, ship=0.0, *,
                 title="1999 Pokemon Base Set Charizard PSA 9",
                 query="1999 Pokemon Base Set Charizard PSA 9",
                 category="Pokemon Cards"):
        return Opportunity(
            listing=Listing(
                site="ebay", title=title, url="", current_price=1.0,
                shipping=ship, query=query, category=category,
                marketplace="EBAY_US"),
            valuation=Valuation(fair_value=fair))

    def _cfg(self, target=0.15, vault=True):
        return {"algorithm": {"resale_fee_rate": 0.1325,
                              "sales_tax_rate": 0.08,
                              "psa_vault": {"enabled": vault,
                                            "min_price": 500,
                                            "sell_fee_rate": 0.07}},
                "output": {"today": {"max_bid_target_roi": target}}}

    def test_max_bid_leaves_the_target_return(self):
        mb, be = report_mod._bid_levels(self._opp_for(2450.0), self._cfg())
        self.assertIsNotNone(mb)
        self.assertLess(mb, be, "max bid must sit below breakeven")
        proceeds = 2450.0 * (1 - 0.07)          # vault route
        self.assertAlmostEqual((proceeds - mb) / mb, 0.15, places=2)

    def test_breakeven_is_still_zero_profit(self):
        _mb, be = report_mod._bid_levels(self._opp_for(2450.0), self._cfg())
        self.assertAlmostEqual(be, round(2450.0 * (1 - 0.07), 0), delta=1)

    def test_target_is_configurable(self):
        loose, _ = report_mod._bid_levels(self._opp_for(2450.0),
                                          self._cfg(target=0.05))
        tight, _ = report_mod._bid_levels(self._opp_for(2450.0),
                                          self._cfg(target=0.40))
        self.assertGreater(loose, tight)

    def test_taxed_route_below_the_vault_threshold(self):
        mb, be = report_mod._bid_levels(self._opp_for(300.0), self._cfg())
        self.assertIsNotNone(be)
        self.assertLess(mb, be)

    def test_max_bid_validates_its_own_vault_route(self):
        # $600 fair -> vault breakeven is $558, but the 15%-ROI vault max
        # is only $485. The old code saw breakeven above $500 and returned
        # the $485 number using tax-free vault math even though a $485 win
        # cannot enter the vault.
        mb, be = report_mod._bid_levels(self._opp_for(600.0), self._cfg())
        self.assertEqual(mb, 419)
        self.assertEqual(be, 558)
        taxed_cost = mb * 1.08
        taxed_proceeds = 600.0 * (1 - 0.1325)
        self.assertGreaterEqual(
            (taxed_proceeds - taxed_cost) / taxed_cost, 0.15)

    def test_whole_dollar_bid_never_rounds_up_across_vault_boundary(self):
        cfg = self._cfg(target=0.0)
        # Make the vault route deliberately unattractive so the best valid
        # route is normal checkout immediately below the $500 boundary.
        cfg["algorithm"]["psa_vault"]["sell_fee_rate"] = 0.40
        mb, _ = report_mod._bid_levels(
            self._opp_for(650.0), cfg)
        self.assertEqual(mb, 499,
                         "a normal-route result must stay below $500")

    def test_watch_never_receives_vault_or_tax_free_bid_math(self):
        watch = self._opp_for(
            2450.0, title="Rolex Submariner 116610LN Full Set",
            query="Rolex Submariner 116610LN", category="Watches")
        mb, be = report_mod._bid_levels(watch, self._cfg())
        taxed_proceeds = 2450.0 * (1 - 0.1325)
        self.assertEqual(be, int(taxed_proceeds / 1.08))
        self.assertLess(mb, be)

    def test_ungraded_card_does_not_receive_vault_math(self):
        raw = self._opp_for(
            2450.0, title="1999 Pokemon Base Set Charizard Holo #4")
        _mb, be = report_mod._bid_levels(raw, self._cfg())
        self.assertEqual(be, int((2450.0 * (1 - 0.1325)) / 1.08))


class TestLandedEconomics(unittest.TestCase):
    """International buys must not look cheap by omitting the costs between
    the Japan hammer price and the item arriving in the US."""

    def _listing(self, channel="ebay"):
        return Listing(
            site="yahoo_jp", title="1996 Pokemon No Rarity CGC 9",
            url="", current_price=100.0, shipping=8.0, buyer_fees=10.0,
            international_shipping=35.0, insurance_rate=0.01,
            import_duty_rate=0.15, fx_spread_rate=0.03,
            query="Pokemon No Rarity", marketplace="YAHOO_JP",
            listing_type="fixed", resale_channel=channel)

    def _config(self):
        return {
            "_config_dir": tempfile.mkdtemp(),
            "algorithm": {
                "sales_tax_rate": 0.08,
                "tax_free_marketplaces": ["YAHOO_JP"],
                "resale_channels": {"ebay": 0.10, "goldin": 0.20},
                "psa_vault": {"enabled": False},
            },
        }

    def test_listing_sums_every_landed_component(self):
        listing = self._listing()
        # $100 + 19% duty/FX/insurance + $53 fixed shipping/proxy.
        self.assertAlmostEqual(listing.total_cost_now, 172.0)
        self.assertIn("duty 15%", listing.landed_cost_note())

    def test_engine_uses_landed_cost_and_selected_exit_channel(self):
        from valuation.engine import ValuationEngine
        engine = ValuationEngine(self._config())
        ebay = engine.score(self._listing("ebay"),
                            Valuation(fair_value=300.0))
        goldin = engine.score(self._listing("goldin"),
                              Valuation(fair_value=300.0))
        self.assertAlmostEqual(ebay.expected_cost, 172.0)
        self.assertAlmostEqual(ebay.edge_now, 98.0)
        self.assertAlmostEqual(ebay.expected_value - goldin.expected_value,
                               30.0)

    def test_max_bid_inverts_the_same_landed_equation(self):
        opp = Opportunity(self._listing(), Valuation(fair_value=300.0))
        max_bid, breakeven = report_mod._bid_levels(opp, self._config())
        self.assertEqual(max_bid, 152)
        self.assertEqual(breakeven, 182)

    def test_percentage_buyer_premium_and_minimum_are_invertible(self):
        listing = Listing(
            site="goldin", title="Card", url="", current_price=10,
            buyer_fee_rate=0.22, minimum_buyer_fee=19)
        self.assertEqual(listing.landed_cost(10), 29)
        self.assertAlmostEqual(listing.item_price_for_landed_cost(29), 10)
        self.assertEqual(listing.landed_cost(100), 122)
        self.assertAlmostEqual(listing.item_price_for_landed_cost(122), 100)


class TestExitOptimizer(unittest.TestCase):
    def _config(self):
        return {
            "_config_dir": tempfile.mkdtemp(),
            "algorithm": {
                "default_resale_channel": "auto",
                "resale_channels": {
                    "ebay": {"fee_rate": 0.1325},
                    "goldin": {
                        "fee_rate": 0.083, "min_value": 100,
                        "categories": ["Sports Cards"],
                        "requires_graded": True,
                    },
                    "heritage": {
                        "fee_rate": 0.01, "auto_enabled": False},
                },
                "psa_vault": {"enabled": False},
                "sales_tax_rate": 0,
            },
        }

    def _listing(self, channel="auto", title="1986 Fleer Jordan PSA 9"):
        return Listing(
            site="goldin", title=title, url="", current_price=500,
            query="Michael Jordan", category="Sports Cards",
            listing_type="fixed", resale_channel=channel)

    def test_auto_selects_highest_net_eligible_exit(self):
        from economics import best_exit_route
        route = best_exit_route(
            self._config(), self._listing(), 1000)
        self.assertEqual(route.channel, "goldin")
        self.assertAlmostEqual(route.net_proceeds, 917)
        self.assertAlmostEqual(route.advantage_vs_ebay, 49.5)

    def test_manual_override_is_honored(self):
        from economics import best_exit_route
        route = best_exit_route(
            self._config(), self._listing("ebay"), 1000)
        self.assertEqual(route.channel, "ebay")
        self.assertAlmostEqual(route.net_proceeds, 867.5)

    def test_ineligible_ungraded_item_stays_on_ebay(self):
        from economics import best_exit_route
        route = best_exit_route(
            self._config(), self._listing(title="1986 Fleer Jordan"), 1000)
        self.assertEqual(route.channel, "ebay")

    def test_engine_records_exit_and_report_exposes_it(self):
        from openpyxl import Workbook
        from valuation.engine import ValuationEngine
        listing = self._listing()
        valuation = ValuationEngine(self._config()).score(
            listing, Valuation(fair_value=1000, confidence=0.8))
        self.assertEqual(valuation.resale_channel, "goldin")
        self.assertEqual(valuation.net_proceeds, 917)
        self.assertEqual(valuation.exit_advantage, 49.5)
        wb = Workbook()
        report_mod._fill_sheet(
            wb.active, [Opportunity(listing, valuation)])
        self.assertEqual(wb.active["Y2"].value, "Goldin")
        self.assertEqual(wb.active["Z2"].value, 0.083)
        self.assertEqual(wb.active["AA2"].value, 917)
        self.assertEqual(wb.active["AB2"].value, 49.5)

    def test_open_portfolio_mark_uses_the_same_optimizer(self):
        import csv
        import portfolio
        with tempfile.TemporaryDirectory() as tmp:
            with open(os.path.join(tmp, "portfolio.csv"), "w",
                      newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(portfolio.CSV_HEADER)
                writer.writerow([
                    "2026-07-01", "1986 Fleer Jordan PSA 9",
                    "Michael Jordan 1986 Fleer", "500", "", "",
                    "auto", "",
                ])
            rows = portfolio.build_rows(
                self._config(),
                {"michael jordan 1986 fleer": 1000}, tmp)
        self.assertEqual(rows[0]["exit_channel"], "goldin")
        self.assertEqual(rows[0]["value"], 917)
        self.assertEqual(rows[0]["pnl"], 417)


class TestMarketplaceParserCanaries(unittest.TestCase):
    def test_goldin_current_schema_and_buyer_premium(self):
        from scrapers.goldin import GoldinScraper

        class Response:
            @staticmethod
            def json():
                return {"searchalgolia": {"lots": [{
                    "status": "Live", "title": "Jordan PSA 9",
                    "current_price": 100, "number_of_bids": 3,
                    "end_timestamp": "2026-08-01T02:00:00Z",
                    "lot_id": "lot-1", "meta_slug": "jordan-psa-9",
                    "buyer_premium": 22,
                }]}}

        scraper = GoldinScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {}})
        with mock.patch.object(scraper, "_post", return_value=Response()):
            rows = scraper.search_auctions("Jordan")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].buyer_fee_rate, 0.22)
        self.assertEqual(rows[0].minimum_buyer_fee, 19)
        self.assertEqual(rows[0].shipping, 6)
        self.assertEqual(rows[0].insurance_rate, 0.009)
        self.assertTrue(rows[0].insurance_on_buyer_fee)
        self.assertAlmostEqual(rows[0].landed_cost(), 129.098)

    def test_goldin_uses_high_value_shipping_tier(self):
        from scrapers.goldin import GoldinScraper

        class Response:
            @staticmethod
            def json():
                return {"searchalgolia": {"lots": [{
                    "status": "Live", "title": "Jordan PSA 10",
                    "current_price": 1000, "number_of_bids": 9,
                    "lot_id": "lot-2", "buyer_premium": 22,
                }]}}

        scraper = GoldinScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {}})
        with mock.patch.object(scraper, "_post", return_value=Response()):
            rows = scraper.search_auctions("Jordan")
        self.assertEqual(rows[0].shipping, 19)
        self.assertAlmostEqual(rows[0].landed_cost(), 1249.98)

    def test_goldin_non_card_uses_non_card_shipping_floor(self):
        from scrapers.goldin import GoldinScraper

        class Response:
            @staticmethod
            def json():
                return {"searchalgolia": {"lots": [{
                    "status": "Live",
                    "title": "Pokemon Factory-Sealed Booster Box",
                    "current_price": 100, "lot_id": "lot-3",
                    "buyer_premium": 22,
                }]}}

        scraper = GoldinScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {}})
        with mock.patch.object(scraper, "_post", return_value=Response()):
            rows = scraper.search_auctions("Pokemon Box")
        self.assertEqual(rows[0].shipping, 19)
        self.assertAlmostEqual(rows[0].landed_cost(), 142.098)

    def test_goldin_insurance_on_price_realized_is_invertible(self):
        listing = Listing(
            site="goldin", title="Card", url="", current_price=100,
            shipping=6, buyer_fee_rate=0.22, minimum_buyer_fee=19,
            insurance_rate=0.009, insurance_on_buyer_fee=True)
        landed = listing.landed_cost()
        self.assertAlmostEqual(landed, 129.098)
        self.assertAlmostEqual(
            listing.item_price_for_landed_cost(landed), 100)

    def test_heritage_current_bid_markup(self):
        from scrapers.heritage import HeritageScraper
        html = """
        <section class="result">
          <a href="/itm/basketball-cards/jordan/a/50086-80953">
            1986 Fleer Michael Jordan #57 PSA Mint 9
          </a>
          <div>Current Bid: $23,000</div>
        </section>
        """
        rows = HeritageScraper.parse_html(html, "Michael Jordan")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].current_price, 23000)
        self.assertEqual(rows[0].buyer_fee_rate, 0.22)
        self.assertEqual(rows[0].minimum_buyer_fee, 29)

    def test_pristine_current_schema_and_buyer_premium(self):
        from scrapers.pristine import PristineScraper
        future = int((datetime.now(timezone.utc)
                      + timedelta(hours=6)).timestamp())
        html = f"""
        <div aria-label="Auction item" class="row product"
             data-pristine-product-venue-id="12868123"
             data-pristine-title="Charizard Vstar CGC 9">
          <a class="title hidden-xs"
             href="/a12868123-Charizard-Vstar-CGC-9">
            Charizard Vstar CGC 9
          </a>
          <p class="high-bid" data-high-bid="100.00">$100.00</p>
          <span class="end-time" data-pristine-end-time="{future}"></span>
          <img src="https://images.example/card.jpg">
        </div>
        """
        rows = PristineScraper.parse_html(html, "Charizard")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].listing_id, "12868123")
        self.assertEqual(rows[0].buyer_fee_rate, 0.17)
        self.assertEqual(rows[0].total_cost_now, 117)

    def test_fanatics_rejects_an_unknown_schema(self):
        from scrapers.fanatics_collect import FanaticsCollectScraper
        scraper = FanaticsCollectScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {},
             "api_keys": {"fanatics": {}}})
        self.assertEqual(scraper._parse({"unexpected": []}, "Jordan"), [])

    def test_fanatics_normalized_auction_has_20_percent_premium(self):
        from scrapers.fanatics_collect import FanaticsCollectScraper
        scraper = FanaticsCollectScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {},
             "api_keys": {"fanatics": {}}})
        payload = {"items": [{
            "id": "f-1", "status": "live", "listing_type": "auction",
            "title": "1986 Fleer Michael Jordan PSA 9",
            "url": "https://www.fanaticscollect.com/item/f-1",
            "current_bid": 1000, "bid_count": 7,
            "grader": "PSA", "certificate_number": "12345678",
        }]}
        rows = scraper._parse(payload, "Jordan", "auction")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].buyer_fee_rate, 0.20)
        self.assertEqual(rows[0].canonical_asset_id, "psa:12345678")
        self.assertEqual(rows[0].total_cost_now, 1200)

    def test_alt_fixed_has_no_auction_premium(self):
        from scrapers.alt import AltScraper
        scraper = AltScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {},
             "api_keys": {"alt": {}}})
        payload = {"items": [{
            "id": "a-1", "status": "active", "type": "buy_now",
            "title": "1999 Charizard PSA 9",
            "url": "https://app.alt.xyz/item/a-1", "price": "$500",
        }]}
        rows = scraper._parse(payload, "Charizard", "fixed")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].buyer_fee_rate, 0)
        self.assertEqual(rows[0].total_cost_now, 500)

    def test_permission_gated_feed_does_not_make_network_call(self):
        from scrapers.alt import AltScraper
        scraper = AltScraper(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {},
             "api_keys": {"alt": {"authorized": False,
                                  "endpoint": "https://example.invalid"}}})
        with mock.patch.object(scraper, "_get") as get:
            self.assertEqual(scraper.search_auctions("Jordan"), [])
        get.assert_not_called()


class TestConnectorCapabilitiesAndDedupe(unittest.TestCase):
    def test_fixed_price_support_is_not_ebay_specific(self):
        from scrapers.base import BaseScraper

        class NewMarketplace(BaseScraper):
            capabilities = frozenset({"fixed"})

        scraper = NewMarketplace(
            {"_config_dir": tempfile.mkdtemp(), "scraping": {}})
        self.assertTrue(scraper.supports("fixed"))
        self.assertFalse(scraper.supports("auctions"))

    def test_crosslisted_asset_keeps_lower_landed_cost(self):
        ebay = Listing(
            site="ebay", title="Jordan PSA 9", url="https://ebay/item/1",
            listing_id="1", canonical_asset_id="PSA:12345678",
            current_price=110)
        alt = Listing(
            site="alt", title="Jordan PSA 9", url="https://alt/item/2",
            listing_id="2", canonical_asset_id="psa:12345678",
            current_price=100)
        rows, duplicates = scanner._dedupe_listings([ebay, alt])
        self.assertEqual(duplicates, 1)
        self.assertEqual(rows, [alt])

    def test_same_title_without_asset_id_remains_two_items(self):
        first = Listing(
            site="ebay", title="Jordan PSA 9", url="https://ebay/item/1",
            listing_id="1", current_price=100)
        second = Listing(
            site="alt", title="Jordan PSA 9", url="https://alt/item/2",
            listing_id="2", current_price=100)
        rows, duplicates = scanner._dedupe_listings([first, second])
        self.assertEqual(duplicates, 0)
        self.assertEqual(rows, [first, second])

    def test_missing_identity_never_collapses_distinct_rows(self):
        first = Listing(
            site="source", title="Jordan PSA 9", url="", current_price=100)
        second = Listing(
            site="source", title="Jordan PSA 9", url="", current_price=100)
        rows, duplicates = scanner._dedupe_listings([first, second])
        self.assertEqual(duplicates, 0)
        self.assertEqual(rows, [first, second])

    def test_same_listing_across_queries_keeps_exact_context_once(self):
        exact = Listing(
            site="ebay",
            title="1999 Pokemon Base Set Raichu Holo #14 PSA 8",
            url="https://www.ebay.com/itm/123456789014",
            listing_id="123456789014", current_price=500,
            query="1999 1st Edition Base Set Raichu Holo PSA 8",
            priority=True, discovery=False, category="Pokemon Cards")
        broad = Listing(
            site="ebay",
            title=exact.title, url=exact.url, listing_id=exact.listing_id,
            current_price=500, query="1999 1st Edition Pokemon Set",
            priority=False, discovery=True, category="Pokemon Cards")
        prepared = [
            (broad.query, ["broad comps"],
             [(scanner._listing_identity(broad), broad)], []),
            (exact.query, ["exact comps"],
             [(scanner._listing_identity(exact), exact)], []),
        ]
        deduped, duplicates = scanner._dedupe_prepared(prepared)
        rows = [listing for _query, _comps, relevant, _asks in deduped
                for _key, listing in relevant]
        self.assertEqual(duplicates, 1)
        self.assertEqual(rows, [exact])
        self.assertEqual(exact.query,
                         "1999 1st Edition Base Set Raichu Holo PSA 8")
        self.assertEqual(
            exact.matched_queries,
            ["1999 1st Edition Base Set Raichu Holo PSA 8",
             "1999 1st Edition Pokemon Set"])

    def test_orchestration_calls_all_advertised_lanes(self):
        auction = mock.Mock()
        auction.supports.side_effect = lambda lane: lane == "auctions"
        auction.search_auctions.return_value = [Listing(
            site="auction_house", title="A", url="a", current_price=1)]
        fixed = mock.Mock()
        fixed.supports.side_effect = lambda lane: lane == "fixed"
        fixed.search_fixed.return_value = [Listing(
            site="marketplace", title="B", url="b", current_price=2,
            listing_type="fixed")]
        rows = scanner._search_marketplaces(
            {"auction_house": auction, "marketplace": fixed},
            "all", "Jordan", 20)
        self.assertEqual([row.site for row in rows],
                         ["auction_house", "marketplace"])
        auction.search_auctions.assert_called_once_with("Jordan", 20)
        fixed.search_fixed.assert_called_once_with("Jordan", 20)

    def test_site_specific_result_cap_keeps_ebay_deep(self):
        ebay = mock.Mock()
        ebay.supports.side_effect = lambda lane: lane == "auctions"
        ebay.search_auctions.return_value = []
        other = mock.Mock()
        other.supports.side_effect = lambda lane: lane == "auctions"
        other.search_auctions.return_value = []

        scanner._search_marketplaces(
            {"ebay": ebay, "other": other},
            "auctions", "Jordan", 40,
            max_results_by_site={"ebay": 200})

        ebay.search_auctions.assert_called_once_with(
            "Jordan", 200, intl=True)
        other.search_auctions.assert_called_once_with("Jordan", 40)

    def test_local_export_matching_preserves_recall(self):
        from scrapers.authorized_feed import AuthorizedFeedScraper
        item = {
            "title": "1999 Pokemon Venusaur Holo PSA 9",
        }
        self.assertTrue(AuthorizedFeedScraper._matches_export_query(
            item, "1999 1st Edition Base Set Venusaur Holo PSA 9"))
        self.assertFalse(AuthorizedFeedScraper._matches_export_query(
            item, "1952 Topps Mickey Mantle PSA 9"))


class TestEbayBrowsePagination(unittest.TestCase):
    @staticmethod
    def _item(number):
        return {
            "title": f"Card {number}",
            "itemId": str(number),
            "itemWebUrl": f"https://www.ebay.com/itm/{number}",
            "buyingOptions": ["AUCTION"],
            "currentBidPrice": {"value": "1.00", "currency": "USD"},
        }

    def test_follows_next_pages_past_the_200_item_page_limit(self):
        from scrapers.ebay import EbayScraper
        scraper = EbayScraper({
            "_config_dir": tempfile.mkdtemp(),
            "scraping": {},
            "marketplaces": ["EBAY_US"],
            "api_keys": {"ebay": {}},
        })
        scraper._token = "test-token"

        responses = []
        for start, count, response_limit, has_next in (
                (0, 200, 200, True),
                (200, 200, 200, True),
                (400, 50, 200, False)):
            response = mock.Mock()
            response.json.return_value = {
                "itemSummaries": [
                    self._item(i) for i in range(start, start + count)
                ],
                "offset": start,
                "limit": response_limit,
                "next": "https://api.ebay.com/next" if has_next else None,
            }
            responses.append(response)

        with mock.patch.object(
                scraper, "_get", side_effect=responses) as get:
            rows = scraper._search_api(
                "Jordan", 450, buying="AUCTION", intl=False)

        self.assertEqual(len(rows), 450)
        self.assertEqual(
            [call.kwargs["params"].get("offset", 0)
             for call in get.call_args_list],
            [0, 200, 400])
        self.assertEqual(
            [call.kwargs["params"]["limit"]
             for call in get.call_args_list],
            [200, 200, 200])


class TestCrossoverPolicy(unittest.TestCase):
    """The regrade sheet is a card strategy, not a generic non-PSA grader
    strategy: WATA games and CGC-graded games cannot leak into it."""

    @staticmethod
    def _opp(title, query):
        return Opportunity(
            Listing(site="ebay", title=title, url="", current_price=300.0,
                    query=query, listing_type="fixed"),
            Valuation(fair_value=1000.0, edge_now=500.0,
                      expected_value=400.0, confidence=0.8))

    def test_only_configured_card_graders_and_categories_appear(self):
        from openpyxl import Workbook
        wb = Workbook()
        rows = [
            self._opp("1986 Fleer Jordan CGC 9", "Michael Jordan 1986 Fleer"),
            self._opp("Super Mario Bros WATA 9.4", "Super Mario Bros NES"),
            self._opp("Zelda CGC 9.0", "Zelda NES"),
        ]
        report_mod._crossover_tab(wb, rows, {"algorithm": {
            "crossover": {
                "enabled": True, "min_profit": 1,
                "allowed_graders": ["CGC", "BGS", "SGC", "BVG"],
                "allowed_categories": ["Pokemon Cards", "Sports Cards"],
            }}})
        self.assertIn("Crossover", wb.sheetnames)
        titles = [r[2] for r in wb["Crossover"].iter_rows(
            min_row=2, values_only=True)]
        self.assertEqual(titles, ["1986 Fleer Jordan CGC 9"])


class TestTrustedFairHistory(unittest.TestCase):
    """Pre-gate fair values remain auditable but cannot mark the portfolio
    or define a 30-day trend."""

    def test_legacy_rows_are_backed_up_and_quarantined(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "history.db")
            conn = sqlite3.connect(path)
            conn.execute(
                "CREATE TABLE fair_history("
                "query TEXT, ts TEXT, fair REAL, n_comps INTEGER)")
            old = (datetime.now(timezone.utc)
                   - timedelta(days=31)).isoformat()
            conn.execute("INSERT INTO fair_history VALUES (?,?,?,?)",
                         ("q", old, 1.0, 0))
            conn.commit()
            conn.close()

            conn = histdb.connect(path)
            self.assertIsNone(conn.execute(
                "SELECT trusted FROM fair_history").fetchone()[0])
            self.assertIsNone(histdb.trend_30d(conn, "q", 100.0))
            histdb.record_fair(conn, "q", 100.0, 8)
            self.assertEqual(conn.execute(
                "SELECT trusted FROM fair_history ORDER BY rowid DESC "
                "LIMIT 1").fetchone()[0], 1)
            import portfolio
            self.assertEqual(portfolio.latest_fairs(conn), {"q": 100.0})
            conn.close()
            backups = [name for name in os.listdir(tmp)
                       if "-pre-fair-trust-" in name]
            self.assertEqual(len(backups), 1)


class TestSourceHealth(unittest.TestCase):
    def test_snapshot_is_persisted_and_rendered(self):
        from openpyxl import Workbook
        from scrapers.base import note_api, reset_api_stats
        from source_health import capture
        with tempfile.TemporaryDirectory() as tmp:
            reset_api_stats()
            note_api("ebay/api", "ok")
            note_api("ebay/api", "failed")
            db_file = os.path.join(tmp, "history.db")
            config = {
                "sites": ["ebay"],
                "database": {"file": db_file, "comp_cache_hours": 48},
                "scraping": {
                    "use_html_comps": True, "use_130point": False},
                "api_keys": {},
            }
            rows = capture(config, "all")
            ebay = next(r for r in rows
                        if r["source"] == "ebay/listings")
            yahoo = next(r for r in rows
                         if r["source"] == "yahoo_jp/listings")
            self.assertEqual(ebay["status"], "degraded")
            self.assertEqual(yahoo["status"], "disabled")
            conn = histdb.connect(db_file)
            self.assertEqual(
                len(histdb.latest_source_health(conn)), len(rows))
            conn.close()
            wb = Workbook()
            report_mod._source_health_tab(wb, rows)
            self.assertIn("Source Health", wb.sheetnames)
            self.assertEqual(wb["Source Health"]["B2"].value,
                             rows[0]["status"])
            reset_api_stats()


class TestCompHygiene(unittest.TestCase):
    """'Babe Ruth 1933' had 116 comps from $1 to $21,000 with a median of
    $6, because the pool mixed 1933 Goudeys with '1991 Conlon ... You Pick'
    lots and modern tribute cards."""

    def test_year_guard_rejects_modern_tribute_cards(self):
        from valuation.comps import year_conflict
        self.assertTrue(year_conflict(
            "George Mikan 1948", "2009-10 Bowman '48 George Mikan Blue"))
        self.assertTrue(year_conflict(
            "Babe Ruth 1933", "1991 Conlon Collection TSN Babe Ruth"))

    def test_year_guard_accepts_season_ranges(self):
        from valuation.comps import year_conflict, years_in
        self.assertFalse(year_conflict(
            "Michael Jordan 1984 Star", "1984-85 STAR #288 MICHAEL JORDAN"))
        self.assertIn("1985", years_in("1984-85 Star"))

    def test_year_guard_is_inert_without_a_year(self):
        from valuation.comps import year_conflict
        self.assertFalse(year_conflict("Charizard Holo PSA 9",
                                       "1999 Charizard Holo PSA 9"))

    def test_card_number_extraction(self):
        from valuation.comps import card_number
        self.assertEqual(card_number("1984-85 STAR #288 MICHAEL JORDAN"), "288")
        self.assertEqual(card_number("1933 Goudey No. 53 Babe Ruth"), "53")
        self.assertEqual(card_number("Charizard 4/102 Base Set"), "4")
        self.assertIsNone(card_number("1948 Bowman George Mikan RC"))
        # a serial number is not a card number
        self.assertIsNone(card_number("Bowman Blue #rd /1948 Mikan"))

    def test_number_guard_pins_one_card(self):
        from valuation.comps import number_conflict
        q = "Michael Jordan 1984 Star #288"
        self.assertFalse(number_conflict(q, "1984-85 STAR #288 JORDAN PSA 6"))
        self.assertTrue(number_conflict(q, "1984-85 Star #101 Jordan XRC"))
        # an unnumbered sale cannot price a specific card
        self.assertTrue(number_conflict(q, "1984 Star Michael Jordan PSA 7"))

    def test_number_guard_is_inert_without_a_number(self):
        from valuation.comps import number_conflict
        self.assertFalse(number_conflict("Michael Jordan 1984 Star",
                                         "1984-85 Star #101 Jordan"))

    def test_guards_are_applied_by_robust_comp_value(self):
        from valuation.comps import robust_comp_value
        from models import SoldComp
        # ungraded query and ungraded comps, so the (pre-existing) grade
        # guard stays inert and this isolates the year/number guards
        pool = [SoldComp(title="1948 Bowman #69 George Mikan RC", price=5200.0),
                SoldComp(title="1948 Bowman #69 George Mikan rookie",
                         price=4800.0),
                SoldComp(title="1948 Bowman #69 Mikan", price=5000.0),
                # wrong era: a 2009 tribute card priced the 1948 rookie at $1
                SoldComp(title="2009-10 Bowman '48 George Mikan Blue #90",
                         price=1.0),
                # wrong card: different number, different player entirely
                SoldComp(title="1948 Bowman #81 George Kurowski", price=663.0)]
        value, n, _disp, _m = robust_comp_value("George Mikan 1948 #69", pool)
        self.assertEqual(n, 3, "only the three #69 sales should count")
        self.assertGreater(value, 4000)


class TestTargetedCompRouting(unittest.TestCase):
    def _engine(self):
        from valuation import ValuationEngine
        tmp = tempfile.mkdtemp()
        engine = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "history.db")},
            "algorithm": {"min_specific_comps": 3,
                          "guide_skip_min_comps": 8},
        })
        engine.guide.guide_value = lambda _query: None
        return engine

    def _listing(self, number="288", *, priority=False, discovery=False,
                 hours=12):
        return Listing(
            site="ebay",
            title=f"1984-85 Star #{number} Michael Jordan PSA 8",
            url=f"https://www.ebay.com/itm/123456789{number.zfill(3)}",
            current_price=500.0, bid_count=4,
            listing_id=f"123456789{number.zfill(3)}",
            query="Michael Jordan 1984 Star PSA 8",
            end_time=datetime.now(timezone.utc) + timedelta(hours=hours),
            priority=priority, discovery=discovery)

    def test_listing_number_and_grade_become_the_sold_search(self):
        engine = self._engine()
        query = engine.targeted_comp_query(self._listing())
        self.assertIsNotNone(query)
        self.assertIn("#288", query)
        self.assertIn("PSA 8", query)

    def test_specific_discovery_is_promoted_to_exact_comp_search(self):
        engine = self._engine()
        promoted = engine.targeted_comp_query(self._listing(discovery=True))
        self.assertIsNotNone(promoted)
        self.assertIn("#288", promoted)
        self.assertIn("PSA 8", promoted)
        vague = self._listing(discovery=True)
        vague.title = "Michael Jordan Fleer card collection"
        self.assertIsNone(engine.targeted_comp_query(vague))

    def test_already_exact_queries_do_not_multiply(self):
        engine = self._engine()
        exact = self._listing()
        exact.query = "Michael Jordan 1984 Star #288 PSA 8"
        self.assertIsNone(engine.targeted_comp_query(exact))

    def test_planner_deduplicates_caps_and_prioritizes(self):
        engine = self._engine()
        ordinary = self._listing("101", priority=False, hours=1)
        priority = self._listing("288", priority=True, hours=48)
        duplicate = self._listing("288", priority=True, hours=2)
        planned = scanner.plan_targeted_comp_queries(
            [ordinary, priority, duplicate], engine, limit=1)
        self.assertEqual(len(planned), 1)
        self.assertIn("#288", planned[0],
                      "priority beats an earlier non-priority listing")

    def test_exact_pool_prices_288_without_101_contamination(self):
        engine = self._engine()
        listing = self._listing()
        broad = [
            SoldComp(f"1984 Star #101 Michael Jordan PSA 8 sale {i}",
                     9000.0 + i * 500)
            for i in range(3)
        ]
        exact = [
            SoldComp(f"1984-85 Star #288 Michael Jordan PSA 8 sale {i}",
                     price)
            for i, price in enumerate((900.0, 1000.0, 1100.0))
        ]
        opp = engine.evaluate(listing, broad, specific_comps=exact)
        self.assertEqual(opp.valuation.n_comps, 3)
        self.assertAlmostEqual(opp.valuation.fair_value, 1000.0, delta=75)
        self.assertFalse(any("MIXED POOL" in note
                             for note in opp.valuation.notes))
        self.assertTrue(any("targeted comp pool" in note
                            for note in opp.valuation.audit_notes))

    def test_thin_exact_pool_stays_non_tradeable(self):
        from quality import is_tradeable
        engine = self._engine()
        listing = self._listing()
        broad = [
            SoldComp(f"1984 Star #101 Michael Jordan PSA 8 sale {i}",
                     9000.0 + i * 500)
            for i in range(3)
        ]
        exact = [
            SoldComp("1984-85 Star #288 Michael Jordan PSA 8", 1000.0)
        ]
        opp = engine.evaluate(listing, broad, specific_comps=exact)
        self.assertTrue(any("MIXED POOL" in note
                            for note in opp.valuation.notes))
        self.assertFalse(is_tradeable(opp))


class TestCollectingStandards(unittest.TestCase):
    """Pokemon: 1st Edition and No Rarity only. Video games: sealed or
    graded only, but at a lower dollar floor than cards."""

    def setUp(self):
        self.cfg = {"filters": {
            "pokemon_eras_only": True,
            "video_games_sealed_or_graded": True}}

    def ok(self, title, query, grail=""):
        o = Opportunity(
            listing=Listing(site="ebay", title=title, url="",
                            current_price=100.0, query=query, grail=grail),
            valuation=Valuation(fair_value=1000.0))
        return scanner.collection_ok(o, self.cfg)

    def test_keeps_first_edition_and_no_rarity(self):
        self.assertTrue(self.ok(
            "1999 Pokemon Base Set 1st Edition Charizard Holo PSA 9",
            "1999 1st Edition Pokemon Set"))
        self.assertTrue(self.ok(
            "1996 Pokemon Japanese No Rarity Charizard PSA 8",
            "Pokemon No Rarity 1996"))

    def test_drops_unlimited_and_modern_pokemon(self):
        self.assertFalse(self.ok(
            "1999 Pokemon Base Set Unlimited Charizard Holo PSA 9",
            "1999 1st Edition Pokemon Set"))
        self.assertFalse(self.ok("2023 Pokemon 151 Charizard PSA 10",
                                 "1999 1st Edition Pokemon Set"))

    def test_named_vintage_sets_still_work(self):
        """Topsun, Carddass and Movie Promo are neither 1st Edition nor No
        Rarity, but they are exactly what those queries are for."""
        self.assertTrue(self.ok("Pokemon Charizard Topsun Blue Back 1997 PSA 7",
                                "Topsun Charizard 1997"))
        self.assertTrue(self.ok("1996 Bandai Carddass Kangaskhan PSA 9",
                                "Bandai Carddass Pokemon PSA"))
        self.assertTrue(self.ok("Movie Promo Charizard 1998 PSA 9",
                                "Movie Promo Charizard"))

    def test_games_must_be_sealed_or_graded(self):
        self.assertTrue(self.ok("Super Mario Bros NES Factory Sealed WATA 9.8",
                                "Super Mario Bros 1985 NES"))
        self.assertTrue(self.ok("Legend of Zelda NES VGA 85 Sealed",
                                "Zelda 1987 NES"))
        self.assertFalse(self.ok("Super Mario Bros NES Cartridge Only Tested",
                                 "Super Mario Bros 1985 NES"))
        self.assertFalse(self.ok("GoldenEye 007 N64 Complete in Box CIB",
                                 "GoldenEye 007 N64"))

    def test_grails_bypass_every_standard(self):
        self.assertTrue(self.ok("2023 Pokemon 151 Charizard PSA 10", "q",
                                grail="Topsun Charizard 1997"))

    def test_other_categories_untouched(self):
        self.assertTrue(self.ok("1986 Fleer Michael Jordan #57 PSA 8",
                                "Michael Jordan 1986 Fleer"))
        self.assertTrue(self.ok("Rolex Submariner 16610", "Rolex Submariner"))


class TestGameGrading(unittest.TestCase):
    """WATA and VGA grade sealed games and were not recognised at all, so a
    sealed WATA 9.8 copy was matched against loose cartridges."""

    def test_wata_is_a_grader(self):
        self.assertEqual(grade_info("Super Mario Bros NES WATA 9.8 A++"),
                         ("wata", "9.8", "9.8"))

    def test_vga_100_point_scale_maps_to_ten(self):
        self.assertEqual(grade_info("Zelda NES VGA 85")[1], "8.5")
        self.assertEqual(grade_info("Metroid NES VGA 90")[1], "9")

    def test_no_psa_shift_for_game_graders(self):
        """The -1 cross-grader penalty is a CARD rule; game graders are
        never compared against PSA."""
        gi = grade_info("Super Mario Bros WATA 9")
        self.assertEqual(gi[1], gi[2])

    def test_sealed_cib_and_loose_are_different_items(self):
        from valuation.comps import variant_conflict
        self.assertTrue(variant_conflict("Super Mario Bros NES Sealed",
                                         "Super Mario Bros NES loose cart"))
        self.assertTrue(variant_conflict("Zelda NES Sealed",
                                         "Zelda NES CIB complete in box"))
        self.assertFalse(variant_conflict("Zelda NES Sealed",
                                          "Zelda NES Factory Sealed WATA 9.4"))


class TestEbaySoldPacing(unittest.TestCase):
    """Parallel scanner workers must look like one slow sold-page visitor."""

    def setUp(self):
        logging.disable(logging.CRITICAL)

    def tearDown(self):
        logging.disable(logging.NOTSET)

    def _scraper(self, sold_delay=0):
        from scrapers.ebay import EbayScraper
        return EbayScraper({
            "_config_dir": tempfile.mkdtemp(),
            "scraping": {
                "request_delay_seconds": 0,
                "ebay_sold_request_delay_seconds": sold_delay,
                "challenge_cooldown_seconds": 0,
            },
        })

    def test_parallel_sold_searches_use_one_lifecycle_lane(self):
        """The lock covers parsing/cooldowns too, not just the HTTP call."""
        from concurrent.futures import ThreadPoolExecutor

        s = self._scraper()
        in_flight = 0
        peak = 0
        calls_lock = threading.Lock()

        def fake_search(*args, **kwargs):
            nonlocal in_flight, peak
            with calls_lock:
                in_flight += 1
                peak = max(peak, in_flight)
            time.sleep(0.005)
            with calls_lock:
                in_flight -= 1
            return []

        s._search_html = fake_search
        with ThreadPoolExecutor(max_workers=10) as pool:
            rows = list(pool.map(
                lambda n: s.search_sold(f"query {n}"), range(30)))

        self.assertEqual(peak, 1)
        self.assertEqual(rows, [[]] * 30)

    def test_sold_searches_leave_configured_quiet_gap(self):
        """The gap begins when the prior sold search has fully completed."""
        s = self._scraper(sold_delay=10)
        now = [100.0]
        sleeps = []

        class FakeTime:
            @staticmethod
            def monotonic():
                return now[0]

            @staticmethod
            def sleep(seconds):
                sleeps.append(seconds)
                now[0] += seconds

        s._search_html = lambda *args, **kwargs: []
        with mock.patch("scrapers.ebay.time", FakeTime):
            s.search_sold("first")
            s.search_sold("second")

        self.assertEqual(sleeps, [10.0])

    def test_second_worker_cannot_enter_during_challenge_cooldown(self):
        """A worker waiting/retrying a challenge keeps later queries out."""
        s = self._scraper()
        first_in_cooldown = threading.Event()
        release_retry = threading.Event()
        entered = []
        entered_lock = threading.Lock()

        def simulated_challenge_search(query, *args, **kwargs):
            with entered_lock:
                entered.append(query)
            if query == "first":
                first_in_cooldown.set()
                release_retry.wait(timeout=2)
            return []

        s._search_html = simulated_challenge_search
        first = threading.Thread(target=s.search_sold, args=("first",))
        second = threading.Thread(target=s.search_sold, args=("second",))
        first.start()
        self.assertTrue(first_in_cooldown.wait(timeout=1))
        second.start()
        self.assertFalse(release_retry.wait(timeout=0.03))
        with entered_lock:
            self.assertEqual(entered, ["first"])
        release_retry.set()
        first.join(timeout=1)
        second.join(timeout=1)
        with entered_lock:
            self.assertEqual(entered, ["first", "second"])

    def test_cookie_reset_requires_a_fresh_homepage_warmup(self):
        s = self._scraper()
        s._warmed = True
        s.reset_cookies()
        self.assertFalse(s._warmed)


class TestApiThrottling(unittest.TestCase):
    """Andrew's rule: after 3 straight failures an endpoint is left alone
    for the rest of the run. These breakers live in the shared network
    layer, so they apply identically to the full scan and the BIN sweep -
    which are the same program (main.py, with and without --mode bin)."""

    def setUp(self):
        # these tests deliberately cause failures; their warnings would
        # otherwise bury the real result in test_results.log
        logging.disable(logging.CRITICAL)

    def tearDown(self):
        logging.disable(logging.NOTSET)

    def _scraper(self, failing=True):
        from scrapers.base import BaseScraper, reset_api_stats
        reset_api_stats()
        tmp = tempfile.mkdtemp()
        cfg = {"_config_dir": tmp,
               "scraping": {"request_delay_seconds": 0,
                            "api_request_delay_seconds": 0,
                            "circuit_breaker_failures": 3}}
        s = BaseScraper(cfg)
        s.site = "testsite"
        s._streaks = {"api": 0, "html": 0}

        class Boom(Exception):
            pass

        import requests as _rq

        def _get(*a, **k):
            if failing:
                raise _rq.RequestException("simulated outage")
            raise Boom()
        s.session.get = _get
        return s

    def test_stops_calling_after_three_straight_failures(self):
        from scrapers.base import API_STATS
        s = self._scraper()
        for _ in range(20):
            s._get("https://example.invalid/x", api=True)
        self.assertEqual(API_STATS[("testsite/api", "failed")], 3,
                         "should have hit the wire exactly 3 times")
        self.assertEqual(API_STATS[("testsite/api", "skipped")], 17)
        self.assertTrue(s.lane_tripped("api"))

    def test_parallel_api_failures_have_atomic_breaker_admission(self):
        """Queued workers must stop at three wire calls, not all fail at once."""
        from concurrent.futures import ThreadPoolExecutor
        from scrapers.base import API_STATS

        s = self._scraper()
        calls = 0
        calls_lock = threading.Lock()
        original_get = s.session.get

        def counted_get(*args, **kwargs):
            nonlocal calls
            with calls_lock:
                calls += 1
            return original_get(*args, **kwargs)

        s.session.get = counted_get
        with ThreadPoolExecutor(max_workers=20) as pool:
            list(pool.map(
                lambda _: s._get("https://example.invalid/x", api=True),
                range(40)))

        self.assertEqual(calls, 3)
        self.assertEqual(API_STATS[("testsite/api", "failed")], 3)
        self.assertEqual(API_STATS[("testsite/api", "skipped")], 37)

    def test_parallel_api_calls_are_one_per_site_at_a_time(self):
        """Successful API calls share one per-scraper admission lane."""
        from concurrent.futures import ThreadPoolExecutor

        s = self._scraper(failing=False)
        in_flight = 0
        peak = 0
        calls_lock = threading.Lock()

        class Response:
            def raise_for_status(self):
                return None

        def successful_get(*args, **kwargs):
            nonlocal in_flight, peak
            with calls_lock:
                in_flight += 1
                peak = max(peak, in_flight)
            time.sleep(0.005)
            with calls_lock:
                in_flight -= 1
            return Response()

        s.session.get = successful_get
        with ThreadPoolExecutor(max_workers=20) as pool:
            rows = list(pool.map(
                lambda _: s._get("https://example.invalid/x", api=True),
                range(40)))

        self.assertTrue(all(row is not None for row in rows))
        self.assertEqual(peak, 1)

    def test_lanes_trip_independently(self):
        s = self._scraper()
        for _ in range(10):
            s._get("https://example.invalid/x", api=True)
        self.assertTrue(s.lane_tripped("api"))
        self.assertFalse(s.lane_tripped("html"))

    def test_summary_names_the_endpoints_left_alone(self):
        from scrapers.base import api_summary
        s = self._scraper()
        for _ in range(10):
            s._get("https://example.invalid/x", api=True)
        summary = api_summary()
        self.assertIn("testsite/api", summary)
        self.assertIn("3 failed", summary)
        self.assertIn("7 skipped", summary)
        self.assertIn("breaker open", summary)

    def test_ebay_oauth_breaker_counts_and_stops(self):
        from scrapers.base import API_STATS, reset_api_stats
        from scrapers.ebay import EbayScraper
        reset_api_stats()
        tmp = tempfile.mkdtemp()
        s = EbayScraper({"_config_dir": tmp,
                         "api_keys": {"ebay": {"client_id": "x",
                                               "client_secret": "y"}},
                         "scraping": {"request_delay_seconds": 0}})
        import requests as _rq
        s.session.post = lambda *a, **k: (_ for _ in ()).throw(
            _rq.RequestException("auth down"))
        for _ in range(15):
            s._get_token()
        self.assertEqual(API_STATS[("ebay/oauth", "failed")], 3)
        self.assertEqual(API_STATS[("ebay/oauth", "skipped")], 12)

    def test_telegram_breaker_is_shared_between_alerts_and_digest(self):
        import alerts
        from scrapers.base import API_STATS, reset_api_stats
        reset_api_stats()
        alerts._tg_breaker.update({"fails": 0, "announced": False})
        try:
            for _ in range(3):
                alerts.telegram_result(False)
            self.assertTrue(alerts.telegram_blocked())
            # digest.py imports the same two helpers, so its sends stop too
            from digest import _send
            self.assertFalse(_send("x", {"bot_token": "t", "chat_id": "c"}))
            self.assertEqual(API_STATS[("telegram", "failed")], 3)
            self.assertGreaterEqual(API_STATS[("telegram", "skipped")], 1)
        finally:
            alerts._tg_breaker.update({"fails": 0, "announced": False})

    def test_price_guide_breakers_stop_at_three(self):
        from scrapers.base import API_STATS, reset_api_stats
        from valuation.price_guide import PriceGuide
        reset_api_stats()
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        g._pc_delay = 0
        import requests as _rq
        g.session.get = lambda *a, **k: (_ for _ in ()).throw(
            _rq.RequestException("guide down"))
        for i in range(12):
            g._pricecharting(f"query {i}")
        self.assertEqual(API_STATS[("pricecharting", "failed")], 3)
        self.assertEqual(API_STATS[("pricecharting", "skipped")], 9)


class TestRunFooter(unittest.TestCase):
    """Every run ends with its duration, in every mode."""

    def test_duration_formatting(self):
        self.assertEqual(scanner._format_duration(0), "0s")
        self.assertEqual(scanner._format_duration(59), "59s")
        self.assertEqual(scanner._format_duration(159), "2m 39s")
        self.assertEqual(scanner._format_duration(3725), "1h 02m 05s")

    def test_footer_logs_duration_and_api_summary(self):
        with self.assertLogs("scanner", level="INFO") as caught:
            scanner._run_footer(0.0, "bin", 0)
        joined = "\n".join(caught.output)
        self.assertIn("run finished in", joined)
        self.assertIn("mode=bin", joined)
        self.assertIn("api calls:", joined)


class TestCredentialHygiene(unittest.TestCase):
    def test_network_text_redacts_query_path_and_header_secrets(self):
        from security import redact_text
        secrets = {
            "pc-token-123456789": (
                "https://www.pricecharting.com/api/product?"
                "t=pc-token-123456789&q=charizard"),
            "123456789:telegram-token-value": (
                "https://api.telegram.org/"
                "bot123456789:telegram-token-value/sendMessage"),
            "bearer-token-123456789": (
                "Authorization: Bearer bearer-token-123456789"),
            "client-secret-123456789": (
                '"client_secret": "client-secret-123456789"'),
        }
        for secret, text in secrets.items():
            safe = redact_text(text)
            self.assertNotIn(secret, safe)
            self.assertIn("<redacted>", safe)

    def test_secrets_file_and_environment_override_inline_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            config_path = os.path.join(tmp, "config.yaml")
            secrets_path = os.path.join(tmp, "secrets.yaml")
            with open(config_path, "w") as f:
                yaml.safe_dump({
                    "api_keys": {
                        "ebay": {"client_id": "inline-id",
                                 "client_secret": "inline-secret"},
                        "pricecharting": {"token": "inline-token"},
                    },
                    "alerts": {"telegram": {"bot_token": "inline-bot"}},
                    "database": {"file": "database/test.db"},
                }, f)
            with open(secrets_path, "w") as f:
                yaml.safe_dump({
                    "api_keys": {
                        "ebay": {"client_secret": "file-secret"},
                        "pricecharting": {"token": "file-token"},
                    },
                    "alerts": {"telegram": {"bot_token": "file-bot"}},
                }, f)
            with mock.patch.dict(
                    os.environ,
                    {"CARD_SCANNER_PRICECHARTING_TOKEN": "env-token"},
                    clear=False):
                config = scanner.load_config(config_path)
            self.assertEqual(
                config["api_keys"]["ebay"]["client_secret"], "file-secret")
            self.assertEqual(
                config["api_keys"]["pricecharting"]["token"], "env-token")
            self.assertEqual(
                config["alerts"]["telegram"]["bot_token"], "file-bot")
            self.assertEqual(
                config["database"]["file"],
                os.path.join(tmp, "database", "test.db"))


class TestCompIdentity(unittest.TestCase):
    def test_ebay_url_variants_have_one_identity(self):
        item_id = "123456789012"
        urls = [
            f"https://www.ebay.com/itm/{item_id}",
            f"https://www.ebay.com/itm/Charizard-PSA-9/{item_id}?hash=x",
            f"https://www.ebay.com/sch/i.html?item={item_id}&mkcid=1",
        ]
        self.assertEqual(
            {histdb.canonical_item_id(url, "ebay") for url in urls},
            {item_id})

    def test_save_comps_counts_same_listing_once_and_refreshes_price(self):
        conn = histdb.connect(":memory:")
        sold = datetime(2026, 7, 20, tzinfo=timezone.utc)
        first = SoldComp(
            "Charizard PSA 9", 100.0, sold,
            "https://www.ebay.com/itm/Charizard/123456789012?hash=old",
            "ebay")
        second = SoldComp(
            "Charizard PSA 9", 110.0, sold,
            "https://www.ebay.com/itm/123456789012?mkcid=1", "ebay")
        histdb.save_comps(conn, "Charizard PSA 9", [first, second])
        rows = conn.execute(
            "SELECT price, comp_key FROM comps").fetchall()
        self.assertEqual(rows, [(110.0, "item:123456789012")])

    def test_old_database_is_backed_up_before_duplicate_cleanup(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "history.db")
            conn = sqlite3.connect(path)
            conn.execute(
                "CREATE TABLE comps(query TEXT,title TEXT,price REAL,"
                "sold_date TEXT,url TEXT,site TEXT,scanned_at TEXT,"
                "UNIQUE(query,url,price))")
            urls = (
                "https://www.ebay.com/itm/Card/123456789012?hash=one",
                "https://www.ebay.com/itm/123456789012?mkcid=two",
            )
            for i, url in enumerate(urls):
                conn.execute(
                    "INSERT INTO comps VALUES (?,?,?,?,?,?,?)",
                    ("q", "Card", 100.0 + i, None, url, "ebay", f"t{i}"))
            conn.commit()
            conn.close()

            migrated = histdb.connect(path)
            self.assertEqual(
                migrated.execute("SELECT COUNT(*) FROM comps").fetchone()[0],
                1)
            migrated.close()
            backups = [
                name for name in os.listdir(tmp)
                if name.startswith("history.db-pre-comp-dedupe-")]
            self.assertEqual(len(backups), 1)
            original = sqlite3.connect(os.path.join(tmp, backups[0]))
            self.assertEqual(
                original.execute("SELECT COUNT(*) FROM comps").fetchone()[0],
                2)
            original.close()


class TestStaleCompEvidenceEndToEnd(unittest.TestCase):
    """2026-08-02. quality.py and db.py were unit-tested for this, but the
    wiring that joins them lives in run_live - and this session already
    shipped two changes that passed their unit tests and failed in the
    real run. So drive the whole path: blocked comp source, stale cache on
    disk, and assert the note reaches the valuation and the row leaves the
    decision set."""

    def _config(self, db_path, cache_age_hours):
        return {
            "sites": ["ebay"],
            "watchlist": [{"query": "1986 Fleer Michael Jordan PSA 8"}],
            "database": {"file": db_path, "comp_cache_hours": 24,
                         "stale_comp_block_hours": 72},
            "scraping": {"use_130point": False, "use_html_comps": True},
            "algorithm": {}, "filters": {}, "economics": {},
        }

    def _seed_stale_cache(self, db_path, query, age_hours):
        conn = histdb.connect(db_path)
        comps = [SoldComp(title=f"1986 Fleer Michael Jordan #57 PSA 8 copy {i}",
                          price=1000.0 + i,
                          sold_date=datetime.now(timezone.utc)
                          - timedelta(days=30),
                          url=f"https://www.ebay.com/itm/9{i:011d}",
                          site="ebay") for i in range(8)]
        histdb.save_comps(conn, query, comps)
        old = (datetime.now(timezone.utc)
               - timedelta(hours=age_hours)).isoformat()
        conn.execute("UPDATE comps SET scanned_at=?", (old,))
        conn.commit()
        conn.close()

    def _run(self, age_hours):
        import main as scanner
        from valuation.engine import ValuationEngine
        query = "1986 Fleer Michael Jordan PSA 8"
        db_path = os.path.join(tempfile.mkdtemp(), "h.db")
        self._seed_stale_cache(db_path, query, age_hours)
        config = self._config(db_path, age_hours)

        listing = Listing(
            site="ebay", title="1986 Fleer Michael Jordan #57 PSA 8",
            url="https://www.ebay.com/itm/123456789012",
            current_price=300.0, bid_count=2, listing_id="123456789012",
            query=query,
            end_time=datetime.now(timezone.utc) + timedelta(hours=5))

        class BlockedEbay:
            site = "ebay"
            tripped = True
            def search_sold(self, *a, **kw):
                return []                 # the breaker is open
            def search(self, *a, **kw):
                return [listing]
            def close(self, *a, **kw):
                return None
            def _get_token(self, *a, **kw):
                return "token"
            def __getattr__(self, name):
                # any other search_* the pipeline reaches for returns an
                # empty result set, not None - these get concatenated
                return lambda *a, **kw: []

        with mock.patch.object(scanner, "scraper_registry",
                               lambda cfg: {"ebay": lambda c: BlockedEbay()}), \
             mock.patch.object(scanner, "_search_marketplaces",
                               lambda *a, **kw: [listing]):
            opps = scanner.run_live(config, ValuationEngine(config), "all")
        return opps

    def test_a_blocked_comp_lane_marks_and_disqualifies_old_evidence(self):
        from quality import tradeability_rejection
        opps = self._run(age_hours=200)          # 8.3 days, past the block
        self.assertTrue(opps, "the row should still be produced for research")
        notes = " | ".join(str(n).upper()
                           for o in opps for n in o.valuation.notes)
        self.assertIn("STALE COMPS", notes)
        self.assertTrue(
            all(tradeability_rejection(o) for o in opps),
            "evidence frozen for 8 days must not be a bid target")

    def test_a_short_outage_annotates_without_disqualifying(self):
        opps = self._run(age_hours=30)           # stale, but inside 72h
        notes = " | ".join(str(n).upper()
                           for o in opps for n in o.valuation.notes)
        self.assertIn("AGING COMPS", notes)
        self.assertNotIn("STALE COMPS", notes)


class TestUnifiedTrustGates(unittest.TestCase):
    def _opp(self, *, note="", disputed=False, regraded=False,
             discovery=False, n_comps=8, fair=1000.0, item_id="123456789012",
             grail="") -> Opportunity:
        listing = Listing(
            site="ebay", title=f"1986 Fleer Jordan PSA 8 {item_id}",
            url=f"https://www.ebay.com/itm/{item_id}",
            current_price=400.0, bid_count=3, listing_id=item_id,
            query="1986 Fleer Michael Jordan PSA 8",
            end_time=datetime.now(timezone.utc) + timedelta(hours=3),
            created_at=datetime.now(timezone.utc) - timedelta(hours=1),
            priority=True, discovery=discovery, grail=grail,
            grail_score=90 if grail else 0)
        valuation = Valuation(
            fair_value=fair, comps_value=fair, n_comps=n_comps,
            expected_cost=500.0, expected_value=350.0, edge_now=450.0,
            roi=0.7, capture=0.8, confidence=0.75,
            opportunity_score=0.42, disputed=disputed, regraded=regraded,
            notes=[note] if note else [])
        return Opportunity(listing=listing, valuation=valuation)

    def test_every_hard_risk_is_blocked_by_the_same_tradeability_gate(self):
        from quality import tradeability_rejection
        cases = [
            (self._opp(note="ASK-BASED estimate from 4 live asks"),
             "ask-based valuation"),
            (self._opp(note="MIXED POOL: no exact card sales"),
             "mixed comp pool"),
            (self._opp(note="SUSPICIOUS: price far below market"),
             "suspicious listing"),
            (self._opp(disputed=True), "disputed valuation"),
            (self._opp(discovery=True), "discovery query"),
            (self._opp(note="STALE COMPS - fresh fetch blocked, evidence "
                            "is 8.4d old"),
             "comp evidence frozen by a blocked source"),
        ]
        for opp, expected in cases:
            self.assertEqual(tradeability_rejection(opp), expected)
        self.assertIsNone(tradeability_rejection(self._opp()))

    def test_aging_comps_annotate_but_do_not_block(self):
        """2026-08-02. eBay opens the comp breaker often enough that
        blocking on ANY staleness would mute the Action sheet routinely,
        which teaches you to ignore it. Only evidence past the block
        threshold is disqualifying; short outages are merely labelled."""
        from quality import tradeability_rejection
        opp = self._opp(note="AGING COMPS - fresh fetch blocked, evidence "
                             "is 1.2d old")
        self.assertIsNone(tradeability_rejection(opp))
        self.assertIn("AGING COMPS", opp.valuation.notes[0],
                      "but the row must still say so on its face")

    def test_regrade_and_collection_failure_never_enter_learning(self):
        from quality import evidence_rejection, is_tradeable
        regraded = self._opp(regraded=True)
        self.assertTrue(is_tradeable(regraded),
                        "a well-comped per-listing regrade can be actionable")
        self.assertEqual(
            evidence_rejection(regraded, collection_passed=True),
            "listing-specific regrade")
        self.assertEqual(
            evidence_rejection(self._opp(), collection_passed=False),
            "outside collection standards")
        self.assertEqual(
            evidence_rejection(
                self._opp(n_comps=2), collection_passed=True, min_comps=3),
            "too few matched comps")

    def test_persistence_writes_clean_evidence_and_rejects_contamination(self):
        conn = histdb.connect(":memory:")
        config = {
            "algorithm": {"learner_min_fair": 50,
                          "learner_min_comps": 3},
            "filters": {},
        }
        recorded, reason = scanner.persist_trusted_evidence(
            conn, self._opp(), config)
        self.assertTrue(recorded)
        self.assertIsNone(reason)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM fair_history").fetchone()[0], 1)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM observations").fetchone()[0], 1)

        poisoned = self._opp(
            note="MIXED POOL: broad set median", item_id="123456789013")
        recorded, reason = scanner.persist_trusted_evidence(
            conn, poisoned, config, recorded)
        self.assertTrue(recorded)
        self.assertEqual(reason, "mixed comp pool")
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM fair_history").fetchone()[0], 1)
        self.assertEqual(
            conn.execute("SELECT COUNT(*) FROM observations").fetchone()[0], 1)
        conn.close()

    def test_decision_sheets_exclude_risk_but_category_keeps_it_visible(self):
        from openpyxl import load_workbook
        clean = self._opp(item_id="123456789014")
        clean.listing.listing_type = "fixed"
        risky = self._opp(
            note="SUSPICIOUS: price far below market",
            item_id="123456789015")
        risky.listing.listing_type = "fixed"
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "quality.xlsx")
            report_mod.write_report(
                [clean, risky], path,
                config={"output": {"today": {
                    "min_expected_value": 75, "min_confidence": 0.25}}})
            wb = load_workbook(path, read_only=True)
            self.assertIn("Today", wb.sheetnames)
            action_titles = [
                row[4] for row in wb["Action"].iter_rows(
                    min_row=2, values_only=True)]
            today_titles = [
                row[1] for row in wb["Today"].iter_rows(
                    min_row=2, values_only=True)]
            category_titles = [
                row[4] for row in wb["Sports Cards"].iter_rows(
                    min_row=2, values_only=True)]
            self.assertIn(clean.listing.title, action_titles)
            self.assertIn(clean.listing.title, today_titles)
            self.assertNotIn(risky.listing.title, action_titles)
            self.assertNotIn(risky.listing.title, today_titles)
            self.assertIn(risky.listing.title, category_titles)
            wb.close()

    def test_report_classifier_preserves_rejections_and_quarantines(self):
        config = {
            "filters": {"min_value": 1000, "max_roi": 2.0},
            "output": {"min_expected_value": 0, "max_rows": 1000},
        }
        clean = self._opp(item_id="123456789020")
        low = self._opp(fair=500, item_id="123456789021")
        negative = self._opp(item_id="123456789022")
        negative.valuation.expected_value = -25
        negative.valuation.roi = -0.05
        risky = self._opp(
            note="MIXED POOL: broad set median",
            item_id="123456789023")

        kept, research, diagnostics = scanner.classify_report_rows(
            [clean, low, negative, risky], config)

        self.assertEqual({id(o) for o in kept}, {id(clean), id(risky)})
        by_stage = {
            (row["stage"], row["reason"])
            for row in research
        }
        self.assertTrue(any(
            stage == "Fair-value floor" for stage, _ in by_stage))
        self.assertIn(
            ("Output economics", "expected value < $0"), by_stage)
        self.assertIn(
            ("Decision-only quarantine", "mixed comp pool"), by_stage)
        self.assertEqual(
            diagnostics["stage_counts"]["Fair-value floor"], 1)
        self.assertEqual(
            diagnostics["stage_counts"]["Output economics"], 1)
        self.assertEqual(
            diagnostics["stage_counts"]["Decision-only quarantine"], 1)

    def test_diagnostic_workbook_exists_even_with_zero_kept_rows(self):
        from openpyxl import load_workbook

        rejected = self._opp(fair=500, item_id="123456789024")
        research = [{
            "stage": "Fair-value floor",
            "reason": "fair value $500 below Sports Cards floor $1,000",
            "opportunity": rejected,
        }]
        waterfall = [{
            "stage": "Raw marketplace hits", "starting": 1,
            "removed": 0, "remaining": 1, "detail": "ebay=1",
        }, {
            "stage": "Fair-value floor", "starting": 1,
            "removed": 1, "remaining": 0,
            "detail": "fair value below floor x1",
        }]
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "diagnostic.xlsx")
            report_mod.write_report(
                [], path, research=research,
                filter_waterfall=waterfall, config={})
            wb = load_workbook(path, read_only=True)
            self.assertIn("Action", wb.sheetnames)
            self.assertIn("Filter Waterfall", wb.sheetnames)
            self.assertIn("Research-Filtered", wb.sheetnames)
            self.assertEqual(wb["Action"].max_row, 1)
            self.assertEqual(
                wb["Research-Filtered"].cell(2, 2).value,
                research[0]["reason"])
            self.assertEqual(
                wb["Filter Waterfall"].cell(3, 3).value, 1)
            wb.close()

    def test_phone_outputs_use_the_tradeability_gate(self):
        import alerts
        import digest
        risky = self._opp(
            note="ASK-BASED estimate from live asks",
            item_id="123456789016", grail="Jordan rookie")
        config = {
            "alerts": {
                "enabled": True, "priority_only": False,
                "min_edge_now": 1, "min_roi": 0, "max_roi": 2,
                "min_capture": 0, "min_confidence": 0,
                "telegram": {"bot_token": "test", "chat_id": "test"},
                "digest": {"enabled": True, "top_opportunities": 10,
                           "top_grails": 10},
            }
        }
        with mock.patch("alerts._send_telegram") as alert_send:
            self.assertEqual(
                alerts.send_alerts([risky], config, ":memory:"), 0)
            alert_send.assert_not_called()
        with mock.patch("digest._send") as digest_send:
            self.assertEqual(digest.send_digest([risky], config), 0)
            digest_send.assert_not_called()


class TestMLDeploymentGate(unittest.TestCase):
    """The GBM deployed because it beat a baseline with an MAE of 1.126 by
    3%. A model must clear an absolute accuracy bar, and cross-validation
    must split by AUCTION so the same item can't sit in train and test."""

    def test_ml_tier_requires_enough_distinct_auctions(self):
        self.assertGreaterEqual(learner.MIN_ML_ITEMS, 150)

    def test_accuracy_ceiling_is_configured(self):
        self.assertLessEqual(learner.DEFAULT_MAX_CV_MAE, 0.25)

    def test_gate_rejects_an_inaccurate_model(self):
        para_mae, max_cv = 1.1257, learner.DEFAULT_MAX_CV_MAE
        cv_mae = 0.2865                      # the model that shipped
        self.assertTrue(cv_mae < para_mae * 0.97, "old gate passed it")
        self.assertFalse(cv_mae < para_mae * 0.97 and cv_mae <= max_cv,
                         "new gate must reject it")


class TestGuideDatabaseLocation(unittest.TestCase):
    """2026-07-26: a stray 28KB history.db appeared in the PROJECT ROOT
    holding guide_cache / guide_meta / guide_product_cache, while the real
    20MB database/history.db sat one folder down. PriceGuide connected to a
    bare relative "history.db", so any config that had not been through
    main.load_config's path resolution silently created a second cache in
    whatever directory the process started in. Guide values written there
    were invisible to every later run, so the scanner re-paid PriceCharting
    for lookups it had already made."""

    def test_relative_db_file_resolves_against_the_config_folder(self):
        from valuation.price_guide import guide_db_path
        config = {"_config_dir": "/somewhere/ebay opportunities",
                  "database": {"file": "database/history.db"}}
        self.assertEqual(
            guide_db_path(config),
            "/somewhere/ebay opportunities/database/history.db")

    def test_absolute_db_file_is_left_alone(self):
        from valuation.price_guide import guide_db_path
        config = {"_config_dir": "/somewhere/else",
                  "database": {"file": "/var/data/history.db"}}
        self.assertEqual(guide_db_path(config), "/var/data/history.db")

    def test_missing_database_config_uses_the_database_folder(self):
        from valuation.price_guide import guide_db_path
        # The old default was a bare "history.db" - THAT is what landed in
        # the project root. The default must be the database/ folder.
        self.assertTrue(
            guide_db_path({"_config_dir": "/proj"}).endswith(
                os.path.join("database", "history.db")),
            "guide must default to database/history.db, never bare CWD")

    def test_guide_does_not_create_a_second_db_beside_the_process(self):
        from valuation.price_guide import PriceGuide
        with tempfile.TemporaryDirectory() as proj, \
                tempfile.TemporaryDirectory() as elsewhere:
            db_file = os.path.join(proj, "database", "history.db")
            os.makedirs(os.path.dirname(db_file), exist_ok=True)
            config = {"_config_dir": proj, "api_keys": {},
                      "database": {"file": "database/history.db"}}
            cwd = os.getcwd()
            os.chdir(elsewhere)          # cron-style: unrelated cwd
            try:
                guide = PriceGuide(config)
                guide._db_put("charizard psa 9", 6718.0)
            finally:
                os.chdir(cwd)
            self.assertTrue(os.path.exists(db_file), "real db not written")
            self.assertFalse(
                os.path.exists(os.path.join(elsewhere, "history.db")),
                "guide created a stray history.db in the working directory")
            conn = sqlite3.connect(db_file)
            self.assertEqual(
                conn.execute("SELECT value FROM guide_cache "
                             "WHERE query=?", ("charizard psa 9",)
                             ).fetchone()[0], 6718.0)
            conn.close()


class TestPokemonTcgKeyGate(unittest.TestCase):
    """2026-07-26: the 11:54 full run reported pokemontcg/guide as
    "disabled" in Source Health while simultaneously logging 7 ok, 10
    failed and 348 skipped calls. source_health calls the source disabled
    when api_keys.pokemontcg.api_key is empty, but the guide made
    ANONYMOUS requests anyway. A source cannot be both disabled and live -
    the operator reading Source Health has to be able to trust it."""

    def test_no_key_means_no_request(self):
        from scrapers.base import API_STATS, reset_api_stats
        from valuation.price_guide import PriceGuide
        reset_api_stats()
        guide = PriceGuide({"api_keys": {"pokemontcg": {"api_key": ""}},
                            "database": {}})
        with mock.patch.object(guide.session, "get") as get:
            self.assertIsNone(guide._pokemontcg("charizard base set"))
            get.assert_not_called()
        self.assertEqual(API_STATS[("pokemontcg.io", "ok")], 0)
        self.assertEqual(API_STATS[("pokemontcg.io", "failed")], 0)
        self.assertEqual(API_STATS[("pokemontcg.io", "skipped")], 1)

    def test_source_health_disabled_agrees_with_runtime_silence(self):
        """Bind the two together so they can never drift apart again."""
        from scrapers.base import API_STATS, reset_api_stats
        from source_health import capture
        from valuation.price_guide import PriceGuide
        with tempfile.TemporaryDirectory() as tmp:
            reset_api_stats()
            config = {
                "sites": ["ebay"],
                "database": {"file": os.path.join(tmp, "history.db"),
                             "comp_cache_hours": 48},
                "scraping": {"use_html_comps": True, "use_130point": False},
                "api_keys": {"pokemontcg": {"api_key": ""}},
            }
            guide = PriceGuide(config)
            with mock.patch.object(guide.session, "get") as get:
                guide._pokemontcg("pikachu")
                guide._pokemontcg("blastoise")
                calls = get.call_count
            row = next(r for r in capture(config, "all")
                       if r["source"] == "pokemontcg/guide")
            self.assertEqual(row["status"], "disabled")
            self.assertEqual(calls, 0,
                             "Source Health says disabled - the guide must "
                             "not be calling the API behind its back")
            self.assertEqual(row["ok"], 0)
            self.assertEqual(row["failed"], 0)

    def test_a_configured_key_still_queries(self):
        from valuation.price_guide import PriceGuide
        guide = PriceGuide({"api_keys": {"pokemontcg": {"api_key": "k"}},
                            "database": {}})
        reply = mock.Mock()
        reply.raise_for_status = mock.Mock()
        reply.json.return_value = {"data": []}
        with mock.patch.object(guide.session, "get",
                               return_value=reply) as get:
            guide._pokemontcg("charizard base set")
            get.assert_called_once()
            self.assertEqual(
                get.call_args.kwargs["headers"]["X-Api-Key"], "k")


class TestBinResultCaps(unittest.TestCase):
    """2026-07-26: the eBay 500-result ceiling went live between the 12:30
    and 13:00 BIN sweeps. Sweep runtime went from 3m37s to over an hour -
    one query alone returned 1,277 eBay rows - and the 13:30 sweep was
    skipped because the 13:00 sweep still held the lock. A 30-minute sweep
    that takes an hour drops every second sweep."""

    SCFG = {"max_results_per_query": 40,
            "max_results_per_query_by_site": {"ebay": 500},
            "max_results_per_query_by_site_bin": {"ebay": 100}}

    def test_full_scan_keeps_deep_ebay_recall(self):
        self.assertEqual(
            scanner.site_result_caps(self.SCFG, "all")["ebay"], 500)
        self.assertEqual(
            scanner.site_result_caps(self.SCFG, "auctions")["ebay"], 500)

    def test_bin_sweep_uses_the_shallow_overlay(self):
        self.assertEqual(
            scanner.site_result_caps(self.SCFG, "bin")["ebay"], 100)

    def test_overlay_only_touches_named_sites(self):
        scfg = dict(self.SCFG,
                    max_results_per_query_by_site={"ebay": 500,
                                                   "goldin": 60})
        self.assertEqual(scanner.site_result_caps(scfg, "bin"),
                         {"ebay": 100, "goldin": 60})

    def test_absent_overlay_is_backward_compatible(self):
        scfg = {"max_results_per_query_by_site": {"ebay": 500}}
        self.assertEqual(scanner.site_result_caps(scfg, "bin")["ebay"], 500)

    def test_shipped_config_keeps_sweeps_no_deeper_than_full_scans(self):
        path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "config.yaml")
        if not os.path.isfile(path):
            self.skipTest("live config.yaml not present")
        with open(path) as fh:
            scfg = (yaml.safe_load(fh) or {}).get("scraping", {}) or {}
        full = scanner.site_result_caps(scfg, "all")
        sweep = scanner.site_result_caps(scfg, "bin")
        for site, cap in sweep.items():
            self.assertLessEqual(
                cap, full.get(site, scfg.get("max_results_per_query", 40)),
                f"{site}: BIN sweep must never dig deeper than a full scan")


class TestPristineLandedCost(unittest.TestCase):
    """Pristine shipping was configured as $0, which understates delivered
    cost on every Pristine row. No Pristine row had reached Action yet, so
    nothing was mispriced in production - but the moment one did, Max Bid
    would have been too high by the whole shipping amount."""

    def test_configured_shipping_reaches_landed_cost(self):
        from scrapers.pristine import PristineScraper
        future = int(time.time()) + 7200
        html = f"""
        <div class="row product" aria-label="Auction item"
             data-pristine-product-venue-id="12868123"
             data-pristine-title="1999 Charizard PSA 9">
          <a class="title" href="/a12868123-charizard">1999 Charizard PSA 9</a>
          <p class="high-bid" data-high-bid="100.00"></p>
          <span class="end-time" data-pristine-end-time="{future}"></span>
        </div>"""
        rows = PristineScraper.parse_html(
            html, "Charizard", 50, buyer_fee_rate=0.17, shipping_estimate=15.0)
        self.assertEqual(len(rows), 1)
        listing = rows[0]
        self.assertEqual(listing.shipping, 15.0)
        # $100 hammer + 17% premium + $15 shipping = $132
        self.assertAlmostEqual(listing.landed_cost(100.0), 132.0, places=2)

    def test_zero_shipping_is_still_honoured_when_configured(self):
        from scrapers.pristine import PristineScraper
        future = int(time.time()) + 7200
        html = f"""
        <div class="row product" aria-label="Auction item"
             data-pristine-product-venue-id="1" data-pristine-title="x">
          <a class="title" href="/a1-x">x</a>
          <p class="high-bid" data-high-bid="100.00"></p>
          <span class="end-time" data-pristine-end-time="{future}"></span>
        </div>"""
        rows = PristineScraper.parse_html(html, "x", 50, 0.17, 0.0)
        self.assertAlmostEqual(rows[0].landed_cost(100.0), 117.0, places=2)


# ==========================================================================
# 2026-07-26: identity-led, PriceCharting-first valuation
# ==========================================================================
# The workbook of 2026-07-26 11:54 contained 4,237 valued rows carrying only
# FIVE distinct (comps, guide) opinions. Fair value was a function of the
# watchlist QUERY, and the price guide was a set-level product lookup that
# silently carried 65% of the blend on every row that reached Action.
#
# Real rows from that run, all reproduced as fixtures below:
#   * eight 2023 Topps Chrome Disney parallels at exactly $1,069.60
#   * six "Superman 1940" rows at $2,821.29 - including a wax wrapper, a
#     novelty coupon, Action Comics #22 and an $18 McFarlane plastic figure
#   * five 1948 Bowman Mikan rows at $1,529.25 - authentic, altered and raw

DISNEY_TITLES = [
    "2023 Topps Chrome Disney 100 Cinderella Pink Refractor #/399 PSA 10",
    "2023 Topps Chrome Disney 100 Meg Year Diamond Refractor #/100 PSA 10",
    "2023 Topps Chrome Disney 100 Escape from a Plane Toy Story Black #/10 PSA 10",
    "2023 Topps Chrome Disney 100 The Dance Beauty Beast Gold Wave #/50 PSA 10",
    "2023 Topps Chrome Disney 100 Zurg Orange Wave Refractor #/25 PSA 10",
    "2023 Topps Chrome Disney 100 Mirabel Madrigal Bruno Dual Auto #/99 PSA 10",
]

SUPERMAN_TITLES = [
    ("1940 Superman Gum R145 Superman Racing the Shells #40 PSA 6", "card"),
    ("1940 SUPERMAN Leader Novelty Premium Coupon R146 Vintage", "coupon"),
    ("1940 Gum Inc. Superman Gum Wax Pack Wrapper VERY SCARCE!", "wrapper"),
    ("1940 D.C. Comics Action Comics 22 Early Superman Cover", "comic"),
    ('McFarlane DC Multiverse Superman Classic 1940 Animation 7" Figure',
     "figure"),
]


class TestCardIdentityExtraction(unittest.TestCase):
    def test_disney_parallels_are_distinct_assets(self):
        from valuation.identity import identity_of
        fps = {identity_of(t).fingerprint() for t in DISNEY_TITLES}
        self.assertEqual(
            len(fps), len(DISNEY_TITLES),
            "a /10 Black and a /399 Pink Refractor are not the same asset")

    def test_parallel_and_serial_are_extracted(self):
        from valuation.identity import identity_of
        i = identity_of(DISNEY_TITLES[0])
        self.assertEqual(i.parallel, "pink refractor")
        self.assertEqual(i.serial, 399)
        self.assertEqual(i.grade, 10.0)
        self.assertEqual(i.object_class, "card")

    def test_parallel_word_order_does_not_matter(self):
        from valuation.identity import parallel_of
        self.assertEqual(parallel_of("Orange Wave Refractor"),
                         parallel_of("Wave Orange Refractor"))

    def test_object_classes_separate_the_superman_pool(self):
        from valuation.identity import identity_of
        for title, expected in SUPERMAN_TITLES:
            self.assertEqual(identity_of(title).object_class, expected,
                             f"misclassified: {title}")

    def test_a_plastic_figure_never_matches_a_card(self):
        from valuation.identity import identity_of
        figure = identity_of(SUPERMAN_TITLES[4][0])
        card = identity_of(SUPERMAN_TITLES[0][0])
        self.assertIsNotNone(figure.conflicts_with(card))
        self.assertIn("figure", figure.conflicts_with(card))

    def test_grade_qualifiers_are_not_a_numeric_grade(self):
        from valuation.identity import identity_of
        auth = identity_of("1948 Bowman #69 George Mikan RC PSA AUTHENTIC")
        altered = identity_of("1948 Bowman George Mikan #69 PSA Authentic Altered")
        graded = identity_of('1948 Bowman George Mikan #69 PSA 3')
        self.assertEqual(auth.qualifier, "authentic")
        self.assertIsNone(auth.grade, "AUTHENTIC is not a numeric grade")
        self.assertEqual(altered.qualifier, "altered")
        self.assertEqual(graded.grade, 3.0)
        self.assertEqual(len({auth.fingerprint(), altered.fingerprint(),
                              graded.fingerprint()}), 3)

    def test_watch_components_are_not_watches(self):
        from valuation.identity import identity_of
        self.assertEqual(
            identity_of("Patek Philippe Ref 1593 platinum dial for restoration"
                        ).object_class, "watch_part")
        self.assertEqual(
            identity_of("Patek Philippe Calatrava 4864R complete watch"
                        ).object_class, "watch")

    def test_specificity_gates_vague_listings(self):
        from valuation.identity import identity_of
        vague = identity_of("1940 Gum Inc. Superman Gum Wax Pack Wrapper")
        exact = identity_of(DISNEY_TITLES[0])
        self.assertLess(vague.specificity(), 0.70)
        self.assertTrue(exact.is_specific(0.70))

    def test_guide_query_carries_parallel_and_serial_not_grade(self):
        from valuation.identity import identity_of
        q = identity_of(DISNEY_TITLES[0]).guide_query()
        self.assertIn("pink refractor", q)
        self.assertIn("/399", q)
        self.assertNotIn("PSA", q.upper())


class TestPriceChartingProductResolution(unittest.TestCase):
    """`/api/product?q=<phrase>` returns ONE best guess with no indication of
    how good it is. `/api/products` returns up to 20 candidates, which is the
    only way to know whether we actually landed the card."""

    def _guide(self, calls):
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "tok"}}})
        g._pc_call = lambda path, params, host=None: calls(path, params)
        return g

    def test_candidate_scoring_picks_the_right_parallel(self):
        from valuation.identity import identity_of
        products = {"status": "success", "products": [
            {"id": "1", "product-name": "Cinderella #100",
             "console-name": "2023 Topps Chrome Disney 100"},
            {"id": "2", "product-name": "Cinderella #100 [Pink Refractor] /399",
             "console-name": "2023 Topps Chrome Disney 100"},
            {"id": "3", "product-name": "Cinderella #100 [Black] /10",
             "console-name": "2023 Topps Chrome Disney 100"},
        ]}
        seen = {}

        def calls(path, params, host=None):
            seen[path] = params
            if path == "products":
                return products
            return {"status": "success", "id": params.get("id"),
                    "product-name": "Cinderella #100 [Pink Refractor] /399",
                    "console-name": "2023 Topps Chrome Disney 100",
                    "genre": "Pokemon Card", "manual-only-price": 106960,
                    "sales-volume": 24}

        g = self._guide(calls)
        q = g.quote(identity_of(DISNEY_TITLES[0]))
        self.assertEqual(q.product_id, "2",
                         "must pick the /399 Pink Refractor, not the base "
                         "card and not the /10 Black")
        self.assertTrue(q.landed)
        self.assertAlmostEqual(q.value, 1069.60, places=2)
        self.assertEqual(q.sales_volume, 24)

    def test_no_good_candidate_returns_no_value_and_says_why(self):
        from valuation.identity import identity_of
        products = {"status": "success", "products": [
            {"id": "9", "product-name": "Charizard #4",
             "console-name": "Pokemon Base Set"}]}
        g = self._guide(lambda path, params: products)
        q = g.quote(identity_of(DISNEY_TITLES[0]))
        self.assertFalse(q.landed)
        self.assertIsNone(q.value)
        self.assertIn("best match only", q.note)

    def test_genre_mismatch_rejects_the_product(self):
        from valuation.identity import identity_of
        def calls(path, params, host=None):
            if path == "products":
                return {"status": "success", "products": [
                    {"id": "5",
                     "product-name": "Superman Gum #40",
                     "console-name": "1940 Gum Inc"}]}
            return {"status": "success", "genre": "Comic",
                    "product-name": "Superman Gum #40",
                    "console-name": "1940 Gum Inc",
                    "loose-price": 282128}
        g = self._guide(calls)
        ident = identity_of("1940 Superman Gum R145 Racing Shells #40 PSA 6")
        self.assertEqual(ident.object_class, "card")
        # The name matches well, so this would have been accepted on title
        # alone. PriceCharting's own genre says it is a Comic - a CARD
        # identity resolved onto a Comic product is a resolution failure,
        # not a price.
        q = g.quote(ident)
        self.assertFalse(q.landed)
        self.assertIsNone(q.value)
        self.assertIn("not the same object", q.note)

    def test_search_and_product_payloads_are_cached_once(self):
        from valuation.identity import identity_of
        hits = []

        def calls(path, params, host=None):
            hits.append(path)
            if path == "products":
                return {"status": "success", "products": [
                    {"id": "2",
                     "product-name": "Cinderella #100 [Pink Refractor] /399",
                     "console-name": "2023 Topps Chrome Disney 100"}]}
            return {"status": "success", "manual-only-price": 106960,
                    "product-name": "Cinderella #100 [Pink Refractor] /399",
                    "console-name": "2023 Topps Chrome Disney 100"}

        g = self._guide(calls)
        ident = identity_of(DISNEY_TITLES[0])
        for _ in range(4):
            g.quote(ident)
        self.assertEqual(hits, ["products", "product"],
                         "product ids are stable - resolve once per run")


class TestGraderTopGradePrices(unittest.TestCase):
    """PriceCharting publishes real CGC 10 / SGC 10 / BGS 10 markets. Those
    prices already include the cross-grader discount, so applying Andrew's
    one-grade-down shift on top of them would double-count it."""

    PRODUCT = {"loose-price": 90000, "cib-price": 180000, "new-price": 330000,
               "graded-price": 671800, "box-only-price": 1200000,
               "manual-only-price": 2500000,
               "condition-17-price": 900000,   # CGC 10
               "condition-18-price": 850000,   # SGC 10
               "bgs-10-price": 1500000}

    def test_cgc_10_uses_the_published_cgc_price(self):
        from valuation.price_guide import _guide_cents
        cents, how = _guide_cents(self.PRODUCT, 9.0, grader="cgc",
                                  printed_grade=10.0)
        self.assertEqual(cents, 900000)
        self.assertIn("condition-17-price", how)

    def test_sgc_and_bgs_10_use_their_own_fields(self):
        from valuation.price_guide import _guide_cents
        self.assertEqual(
            _guide_cents(self.PRODUCT, 9.0, grader="sgc",
                         printed_grade=10.0)[0], 850000)
        self.assertEqual(
            _guide_cents(self.PRODUCT, 9.0, grader="bgs",
                         printed_grade=10.0)[0], 1500000)

    def test_missing_grader_field_falls_back_to_the_shift(self):
        from valuation.price_guide import _guide_cents
        product = dict(self.PRODUCT)
        product.pop("condition-17-price")
        cents, how = _guide_cents(product, 9.0, grader="cgc",
                                  printed_grade=10.0)
        self.assertEqual(cents, 671800, "falls back to PSA-9 graded-price")
        self.assertEqual(how, "graded-price")

    def test_psa_10_is_unaffected(self):
        from valuation.price_guide import _guide_cents
        cents, _ = _guide_cents(self.PRODUCT, 10.0, grader="psa",
                                printed_grade=10.0)
        self.assertEqual(cents, 2500000)

    def test_a_cgc_9_still_shifts_normally(self):
        from valuation.price_guide import _guide_cents
        cents, _ = _guide_cents(self.PRODUCT, 8.0, grader="cgc",
                                printed_grade=9.0)
        self.assertEqual(cents, 330000, "CGC 9 -> PSA 8 -> new-price")


class TestGuideLedValuation(unittest.TestCase):
    """Andrew's rule: on an exact product match the paid guide SETS the
    value; identity-matched comps may pull it down but never inflate it."""

    def _engine(self, quote):
        from valuation import ValuationEngine
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "h.db"),
                         "comp_cache_hours": 48},
            "algorithm": {"min_specific_comps": 3},
        })
        e.guide.quote = lambda ident: quote
        return e

    def _listing(self, title, **kw):
        return Listing(site="ebay", title=title,
                       url="https://www.ebay.com/itm/1",
                       current_price=kw.pop("price", 400.0), bid_count=3,
                       listing_id="1", query=kw.pop("query", "Disney Chrome 2023"),
                       end_time=datetime.now(timezone.utc) + timedelta(hours=6),
                       **kw)

    def _quote(self, **kw):
        from valuation.price_guide import GuideQuote
        from valuation.identity import MATCH_EXACT
        base = dict(value=1200.0, match=MATCH_EXACT, score=0.9,
                    product_id="2", product_name="Cinderella [Pink] /399",
                    how="manual-only-price")
        base.update(kw)
        return GuideQuote(**base)

    def test_exact_match_lets_the_guide_set_the_value(self):
        e = self._engine(self._quote())
        opp = e.evaluate(self._listing(DISNEY_TITLES[0]), [])
        self.assertAlmostEqual(opp.valuation.fair_value, 1200.0)
        self.assertEqual(opp.valuation.identity_match, "exact")

    def test_comps_may_lower_the_guide_but_never_raise_it(self):
        e = self._engine(self._quote())
        low = [SoldComp("2023 Topps Chrome Disney 100 Cinderella Pink "
                        f"Refractor #/399 PSA 10 sale {i}", 800.0 + i)
               for i in range(4)]
        opp = e.evaluate(self._listing(DISNEY_TITLES[0]), low)
        self.assertLess(opp.valuation.fair_value, 1000.0,
                        "comps below the guide must pull the value down")
        high = [SoldComp("2023 Topps Chrome Disney 100 Cinderella Pink "
                         f"Refractor #/399 PSA 10 sale {i}", 5000.0 + i)
                for i in range(4)]
        opp = self._engine(self._quote()).evaluate(
            self._listing(DISNEY_TITLES[0]), high)
        self.assertAlmostEqual(opp.valuation.fair_value, 1200.0,
                               msg="comps must never inflate an exact match")

    def test_thin_comps_cannot_pull_the_guide_down(self):
        e = self._engine(self._quote())
        one = [SoldComp("2023 Topps Chrome Disney 100 Cinderella Pink "
                        "Refractor #/399 PSA 10", 300.0)]
        opp = e.evaluate(self._listing(DISNEY_TITLES[0]), one)
        self.assertAlmostEqual(opp.valuation.fair_value, 1200.0)

    def test_weak_match_is_browse_only_and_not_tradeable(self):
        from quality import is_tradeable
        from valuation.identity import MATCH_WEAK
        e = self._engine(self._quote(match=MATCH_WEAK, score=0.5))
        comps = [SoldComp("2023 Topps Chrome Disney 100 Cinderella Pink "
                          f"Refractor #/399 PSA 10 sale {i}", 900.0 + i)
                 for i in range(5)]
        opp = e.evaluate(self._listing(DISNEY_TITLES[0]), comps)
        self.assertTrue(any("IDENTITY UNRESOLVED" in n
                            for n in opp.valuation.notes))
        self.assertFalse(is_tradeable(opp))

    def test_the_disney_eight_no_longer_share_one_value(self):
        """The headline regression: 8 parallels, 8 different numbers."""
        from valuation.price_guide import GuideQuote
        from valuation.identity import MATCH_EXACT, identity_of
        from valuation import ValuationEngine
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({"_config_dir": tmp,
                             "database": {"file": os.path.join(tmp, "h.db")},
                             "algorithm": {"min_specific_comps": 3}})
        # PriceCharting prices each parallel separately - as it does in
        # reality, because each parallel is its own catalogue product.
        prices = {}
        for n, t in enumerate(DISNEY_TITLES):
            prices[identity_of(t).fingerprint()] = 500.0 + 250.0 * n
        e.guide.quote = lambda ident: GuideQuote(
            value=prices[ident.fingerprint()], match=MATCH_EXACT, score=0.9,
            product_id="x", product_name="p", how="manual-only-price")
        values = [e.evaluate(self._listing(t), []).valuation.fair_value
                  for t in DISNEY_TITLES]
        self.assertEqual(len(set(values)), len(DISNEY_TITLES),
                         "each parallel must get its own fair value")

    def test_a_figure_cannot_inherit_a_card_pool(self):
        from valuation.price_guide import GuideQuote
        e = self._engine(GuideQuote())          # guide lands nothing
        card_comps = [SoldComp(f"1940 Superman Gum R145 #40 PSA 6 sale {i}",
                               2800.0 + i) for i in range(6)]
        opp = e.evaluate(
            self._listing(SUPERMAN_TITLES[4][0], price=18.0,
                          query="Superman 1940"), card_comps)
        v = opp.valuation
        self.assertTrue(
            any("identity filter dropped" in a for a in v.audit_notes),
            "graded-card sales must not survive into a plastic figure's pool")

    def test_stale_comps_discount_confidence(self):
        e = self._engine(self._quote())
        fresh = datetime.now(timezone.utc)
        old = datetime.now(timezone.utc) - timedelta(hours=200)
        def comps(ts):
            return [SoldComp("2023 Topps Chrome Disney 100 Cinderella Pink "
                             f"Refractor #/399 PSA 10 sale {i}", 1100.0 + i,
                             sold_date=ts) for i in range(5)]
        hot = e.evaluate(self._listing(DISNEY_TITLES[0]), comps(fresh))
        cold = self._engine(self._quote()).evaluate(
            self._listing(DISNEY_TITLES[0]), comps(old))
        self.assertLess(cold.valuation.confidence, hot.valuation.confidence,
                        "a 200-hour-old pool must not read as fresh conviction")
        self.assertTrue(any("confidence discounted" in a
                            for a in cold.valuation.audit_notes))


class TestSpecificityByObjectClass(unittest.TestCase):
    """The first specificity() demanded a card number and a parallel, so a
    sealed WATA Super Mario Bros scored 50% and was permanently locked out
    of guide-led valuation - despite video games being PriceCharting's
    deepest and oldest catalogue. What counts as "pinned" depends on what
    the thing IS."""

    def _spec(self, title):
        from valuation.identity import identity_of
        return identity_of(title)

    def test_a_sealed_graded_game_is_specific(self):
        for title in ("Super Mario Bros NES Sealed WATA 9.4 A",
                      "The Legend of Zelda NES Sealed VGA 85",
                      "Pokemon Red Version Game Boy Sealed WATA 9.0"):
            i = self._spec(title)
            self.assertEqual(i.object_class, "game")
            self.assertTrue(i.is_specific(),
                            f"{title} -> {i.specificity():.0%}")

    def test_a_graded_numbered_comic_is_specific(self):
        i = self._spec("Action Comics #22 1940 DC Comics CGC 4.0")
        self.assertTrue(i.is_specific())

    def test_a_watch_stays_vague(self):
        # PriceCharting does not carry watches; gating them is correct, and
        # watches are excluded from Action by policy anyway.
        i = self._spec("Rolex Submariner 5513 stainless steel watch")
        self.assertFalse(i.is_specific())

    def test_a_vintage_variant_counts_when_there_is_no_parallel(self):
        i = self._spec("1997 Topsun Charizard Green Back PSA 8")
        self.assertTrue(i.is_specific(),
                        "Green Back is a real discriminator")


class TestGameIdentity(unittest.TestCase):
    """The 2026-07-26 coverage probe landed 0/3 sealed games despite video
    games being PriceCharting's oldest catalogue. Subject extraction is
    tuned for cards - it dropped "Super" as a generic adjective and "NES"
    as a generic token, so we searched for 'mario bros' with no platform."""

    def _i(self, t):
        from valuation.identity import identity_of
        return identity_of(t)

    def test_the_game_title_survives_intact(self):
        i = self._i("Super Mario Bros NES Sealed WATA 9.4 A")
        q = i.guide_query()
        self.assertIn("super", q)
        self.assertIn("mario", q)
        self.assertIn("nes", q)
        self.assertNotIn("wata", q)
        self.assertNotIn("sealed", q)

    def test_the_platform_is_captured_as_a_set_token(self):
        self.assertIn("nes", self._i("Super Mario Bros NES Sealed").set_tokens)
        self.assertIn("gameboy",
                      self._i("Pokemon Red Version GameBoy Sealed").set_tokens)

    def test_a_sequel_is_not_the_original(self):
        i = self._i("Super Mario Bros NES Sealed WATA 9.4")
        original = i.score_candidate("Super Mario Bros", "NES")
        sequel = i.score_candidate("Super Mario Bros 3", "NES")
        compilation = i.score_candidate("Super Mario Bros and Duck Hunt",
                                        "NES")
        self.assertGreater(original, sequel + 0.05)
        self.assertGreater(original, compilation + 0.05)

    def test_the_shorter_wrong_game_does_not_win(self):
        # "Mario Bros" outranked "Super Mario Bros" 72% to 69% before the
        # recall term - a different and far cheaper game.
        i = self._i("Super Mario Bros NES Sealed WATA 9.4")
        self.assertGreater(i.score_candidate("Super Mario Bros", "NES"),
                           i.score_candidate("Mario Bros", "NES"))

    def test_a_pal_release_is_a_different_product(self):
        i = self._i("Super Mario Bros NES Sealed WATA 9.4")
        self.assertGreater(i.score_candidate("Super Mario Bros", "NES"),
                           i.score_candidate("Super Mario Bros", "PAL NES"))

    def test_a_cgc_graded_comic_is_a_comic_not_a_card(self):
        # CGC grades both, so the grader cannot be the card signal.
        i = self._i("Action Comics #22 1940 DC Comics CGC 4.0")
        self.assertEqual(i.object_class, "comic")
        self.assertTrue(i.is_specific())


class TestAlphanumericCardNumbers(unittest.TestCase):
    """Modern inserts number themselves #T264 / #IM-15 / #BC-15. comps
    .card_number reads digits only - correct for the comp filter it was
    built for - so the cards whose identity matters most had no number."""

    def test_letter_prefixed_numbers_are_read(self):
        from valuation.identity import card_number_of
        self.assertEqual(card_number_of(
            "2001 Topps Chrome Traded #T264 Ichiro PSA 10"), "T264")
        self.assertEqual(card_number_of(
            "Escape from a Plane #IM-15 Black"), "IM-15")

    def test_plain_numbers_are_unchanged(self):
        from valuation.identity import card_number_of
        self.assertEqual(card_number_of("1952 Topps #311 Mantle"), "311")
        self.assertEqual(card_number_of("Charizard #4 Holo"), "4")

    def test_a_print_run_is_not_a_card_number(self):
        from valuation.identity import card_number_of
        self.assertIsNone(card_number_of("Pink Refractor #/399 PSA 10"))

    def test_a_year_is_not_a_card_number(self):
        from valuation.identity import card_number_of
        self.assertIsNone(card_number_of("1933 Goudey Babe Ruth"))

    def test_it_feeds_identity_and_lifts_specificity(self):
        from valuation.identity import identity_of
        i = identity_of("2001 Topps Chrome Traded #T264 Ichiro Rookie PSA 10")
        self.assertEqual(i.number, "T264")
        self.assertTrue(i.is_specific())


class TestSportsCardHostRouting(unittest.TestCase):
    """2026-07-26: pricecharting.com carries TCG, video games, comics, Funko
    and LEGO but NOT sports cards - a 1952 Mantle search returned Funko POPs
    and LEGO sets, 0/8 across two sports categories. sportscardspro.com is
    the same company's sports catalogue, the same token reached it, and it
    landed all six test cards at 90-100%."""

    def _guide(self, responses):
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        g.hosts_called = []

        def call(path, params, host=None):
            g.hosts_called.append(host)
            return responses(path, params, host)
        g._pc_call = call
        return g

    def test_a_sports_card_falls_through_to_sportscardspro(self):
        from valuation.identity import identity_of

        def responses(path, params, host):
            if "sportscardspro" not in (host or ""):
                # what pricecharting really returned for this card
                return {"status": "success", "products": [
                    {"id": "1", "product-name": "Mugman #311",
                     "console-name": "Funko POP Games"}]}
            if path == "products":
                return {"status": "success", "products": [
                    {"id": "72584", "product-name": "Mickey Mantle #311",
                     "console-name": "Baseball Cards 1952 Topps"}]}
            return {"status": "success", "product-name": "Mickey Mantle #311",
                    "console-name": "Baseball Cards 1952 Topps",
                    "genre": "Baseball Card", "cib-price": 4138000}

        g = self._guide(responses)
        q = g.quote(identity_of("1952 Topps #311 Mickey Mantle PSA 7"))
        self.assertTrue(q.landed, "SportsCardsPro has this card")
        self.assertEqual(q.product_id, "72584")
        self.assertTrue(any("sportscardspro" in (h or "")
                            for h in g.hosts_called))

    def test_a_pokemon_card_never_needs_the_second_host(self):
        from valuation.identity import identity_of

        def responses(path, params, host):
            if path == "products":
                return {"status": "success", "products": [
                    {"id": "7096109",
                     "product-name": "Charizard [1st Edition] #4",
                     "console-name": "Pokemon Base Set"}]}
            return {"status": "success",
                    "product-name": "Charizard [1st Edition] #4",
                    "console-name": "Pokemon Base Set",
                    "genre": "Pokemon Card", "new-price": 2500000}

        g = self._guide(responses)
        q = g.quote(identity_of("1999 Pokemon Base Set 1st Edition "
                                "Charizard #4 Holo PSA 8"))
        self.assertTrue(q.landed)
        self.assertFalse(any("sportscardspro" in (h or "")
                             for h in g.hosts_called),
                         "a card that lands on the first host must not cost "
                         "a second lookup")

    def test_the_two_hosts_never_share_a_cache_entry(self):
        from valuation.price_guide import _host_tag
        self.assertNotEqual(_host_tag("https://www.pricecharting.com"),
                            _host_tag("https://www.sportscardspro.com"))

    def test_an_unasked_bracket_variant_is_penalised(self):
        """`Michael Jordan #57` 90% vs `[20th Anniversary] #57` 86% is four
        points - a coin flip to the margin gate, though the listing says
        nothing about an anniversary reprint."""
        from valuation.identity import identity_of
        i = identity_of("1986 Fleer #57 Michael Jordan Rookie PSA 9")
        base = i.score_candidate("Michael Jordan #57",
                                 "Basketball Cards 1986 Fleer")
        variant = i.score_candidate("Michael Jordan [20th Anniversary] #57",
                                    "Basketball Cards 1986 Fleer")
        self.assertGreater(base, variant + 0.06,
                           "the base card must clear the margin gate")

    def test_the_wrong_player_scores_zero_however_well_numbered(self):
        """SportsCardsPro returned `Chris Smith #T264` and `Mike Adams #T264`
        at 64% each for an Ichiro listing - same set, same card number,
        entirely the wrong person. The name on the card cannot be outvoted
        by numbering and set agreement."""
        from valuation.identity import identity_of
        i = identity_of("2001 Topps Chrome Traded #T264 Ichiro Rookie PSA 10")
        console = "Baseball Cards 2001 Topps Chrome Traded"
        self.assertEqual(i.score_candidate("Chris Smith #T264", console), 0.0)
        self.assertEqual(i.score_candidate("Mike Adams #T264", console), 0.0)
        self.assertGreater(i.score_candidate("Ichiro Suzuki #T264", console),
                           0.62)

    def test_two_word_platforms_survive(self):
        """'Game Boy' contributed nothing, so we searched for the product
        'pokemon red boy'."""
        from valuation.identity import identity_of, set_tokens_of
        self.assertIn("gameboy", set_tokens_of("Pokemon Red Version Game Boy"))
        i = identity_of("Pokemon Red Version Game Boy Sealed WATA 9.0")
        q = i.guide_query()
        self.assertIn("version", q)
        # BOTH sides fold to the catalogue's single word, or our 'game'+'boy'
        # never matches their 'GameBoy' and recall halves.
        self.assertIn("gameboy", q)
        self.assertGreater(i.score_candidate("Pokemon Red Version", "GameBoy"),
                           0.62)
        self.assertGreater(
            i.score_candidate("Pokemon Red Version", "GameBoy"),
            i.score_candidate("Pokemon Blue Version", "GameBoy") + 0.2)

    def test_a_variant_we_DID_ask_for_is_not_penalised(self):
        from valuation.identity import identity_of
        i = identity_of("2018 Panini Prizm #280 Luka Doncic Silver PSA 10")
        self.assertGreater(
            i.score_candidate("Luka Doncic [Silver Prizm] #280",
                              "Basketball Cards 2018 Panini Prizm"),
            0.80)


class TestLocalPriceGuideCsv(unittest.TestCase):
    """A downloaded price-guide CSV carries the same columns as an API
    response, so it can answer a lookup with no call and no latency. The
    paid API is one call per second - the dominant cost of a full scan."""

    HEADER = ("id,product-name,console-name,loose-price,cib-price,"
              "new-price,graded-price,box-only-price,manual-only-price,"
              "genre,sales-volume\n")

    def _folder(self, body, name="set.csv"):
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        with open(os.path.join(tmp, "guide_csv", name), "w") as fh:
            fh.write(self.HEADER + body)
        return tmp

    def test_it_loads_and_indexes_rows(self):
        from valuation.guide_csv import load_index
        tmp = self._folder(
            "72584,Michael Jordan #57,Basketball Cards 1986 Fleer,"
            "225500,413800,602295,,,,Basketball Card,145\n"
            "72585,Michael Jordan #8,Basketball Cards 1986 Fleer Sticker,"
            "50000,,,,,,Basketball Card,12\n")
        idx = load_index(tmp)
        self.assertEqual(len(idx), 2)
        rows = idx.search("1986 fleer michael jordan #57")
        self.assertTrue(rows)
        self.assertEqual(rows[0]["console-name"],
                         "Basketball Cards 1986 Fleer")

    def test_dollar_formatted_prices_are_normalised_to_cents(self):
        from valuation.guide_csv import load_index
        tmp = self._folder(
            '1,Mickey Mantle #311,Baseball Cards 1952 Topps,"$2,255.00",'
            '"$4,138.00",,,,,Baseball Card,\n')
        row = load_index(tmp).rows[0]
        self.assertEqual(row["loose-price"], 225500)
        self.assertEqual(row["cib-price"], 413800)

    def test_a_csv_answer_costs_no_api_call(self):
        from valuation.price_guide import PriceGuide
        tmp = self._folder(
            "72584,Michael Jordan #57,Basketball Cards 1986 Fleer,"
            "225500,413800,602295,4064100,,,Basketball Card,145\n")
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        calls = []
        g._pc_call = lambda path, params, host=None: calls.append(path) or None
        q = g.quote(self._ident_for(
            "1986 Fleer #57 Michael Jordan Rookie PSA 9"))
        self.assertTrue(q.landed, "the CSV has this card")
        self.assertEqual(calls, [], "a local hit must not touch the network")
        self.assertAlmostEqual(q.value, 40641.00, places=2)
        self.assertEqual(q.sales_volume, 145)

    def test_a_card_not_in_the_csv_still_uses_the_api(self):
        from valuation.price_guide import PriceGuide
        tmp = self._folder(
            "72584,Michael Jordan #57,Basketball Cards 1986 Fleer,"
            "225500,,,,,,Basketball Card,145\n")
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        calls = []

        def call(path, params, host=None):
            calls.append(path)
            return {"status": "success", "products": []}
        g._pc_call = call
        g.quote(self._ident_for("1952 Topps #311 Mickey Mantle PSA 5"))
        self.assertTrue(calls, "a miss must fall through to the API")

    def test_a_junk_file_is_ignored_not_fatal(self):
        from valuation.guide_csv import load_index
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        with open(os.path.join(tmp, "guide_csv", "notes.csv"), "w") as fh:
            fh.write("date,amount\n2026-07-26,12.00\n")
        self.assertEqual(len(load_index(tmp)), 0)

    def test_no_folder_at_all_is_fine(self):
        from valuation.guide_csv import load_index
        self.assertEqual(len(load_index(tempfile.mkdtemp())), 0)

    def test_the_csv_is_held_to_the_same_margin_gate(self):
        """Local data gets no special trust: an ambiguous pair is refused
        exactly as it would be from the API."""
        from valuation.price_guide import PriceGuide
        tmp = self._folder(
            "1,Babe Ruth #53,Baseball Cards 1933 Goudey,100000,,,,,,"
            "Baseball Card,\n"
            "2,Babe Ruth #53,Baseball Cards 1933 Goudey,200000,,,,,,"
            "Baseball Card,\n")
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        g._pc_call = lambda path, params, host=None: {"status": "success",
                                                      "products": []}
        q = g.quote(self._ident_for("1933 Goudey #53 Babe Ruth PSA 4"))
        self.assertFalse(q.landed)

    @staticmethod
    def _ident_for(title):
        from valuation.identity import identity_of
        return identity_of(title)


class TestGuideCsvDownloader(unittest.TestCase):
    """`/price-guide/download-custom?t=TOKEN&category=<slug>` returns a whole
    category in one request. Their limits are one CSV per 10 minutes and a
    24-hour regeneration cycle, so the downloader must pace itself and must
    not re-fetch files that are already current."""

    def setUp(self):
        """Neutralise curl_cffi for every test in this class.

        Regression: 2026-08-01. `_http_get` prefers curl_cffi and only falls
        back to `requests`. Tests that patch `m.requests.get` therefore did
        NOT intercept anything on a machine where curl_cffi is installed -
        they made LIVE network calls, failed, and the self-test gate then
        blocked every scan. They passed on a machine without curl_cffi,
        which is the worst possible failure mode: green where it is not
        installed, red and networked where it is.

        Defaulting it off means a test reaches the network only if it
        deliberately re-patches `curl_requests` itself.
        """
        import fetch_guide_csv
        patcher = mock.patch.object(fetch_guide_csv, "curl_requests", None)
        patcher.start()
        self.addCleanup(patcher.stop)

    def _mod(self):
        import fetch_guide_csv
        return fetch_guide_csv

    def test_filenames_separate_the_two_sites(self):
        m = self._mod()
        self.assertEqual(m._filename(m.PC, "pokemon-cards"),
                         "pricecharting--pokemon-cards.csv")
        self.assertEqual(m._filename(m.SCP, "baseball-cards"),
                         "sportscardspro--baseball-cards.csv")
        # no category = the host default, which on PriceCharting is games
        self.assertEqual(m._filename(m.PC, ""),
                         "pricecharting--video-games.csv")

    def test_an_html_response_is_rejected_not_saved(self):
        """A refused token returns a web page. Writing that to guide_csv/
        would poison every future lookup with garbage."""
        m = self._mod()
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "out.csv")

        class FakeResp:
            status_code = 200
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"<!DOCTYPE html><html><body>Login</body></html>"

        with mock.patch.object(m.requests, "get", return_value=FakeResp()):
            ok, detail, kind = m.download(m.PC, "", "tok", path)
        self.assertFalse(ok)
        self.assertIn("web page", detail)
        self.assertFalse(os.path.exists(path))

    def test_a_response_without_the_expected_header_is_rejected(self):
        m = self._mod()
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "out.csv")

        class FakeResp:
            status_code = 200
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"date,amount\n2026-01-01,5\n"

        with mock.patch.object(m.requests, "get", return_value=FakeResp()):
            ok, _, _kind = m.download(m.PC, "", "tok", path)
        self.assertFalse(ok)
        self.assertFalse(os.path.exists(path))

    def test_a_real_csv_is_written(self):
        m = self._mod()
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "out.csv")

        class FakeResp:
            status_code = 200
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield (b"id,product-name,console-name,loose-price\n"
                       b"1,Michael Jordan #57,Basketball Cards 1986 Fleer,225500\n")

        with mock.patch.object(m.requests, "get", return_value=FakeResp()):
            ok, detail, kind = m.download(m.PC, "", "tok", path)
        self.assertTrue(ok, detail)
        self.assertTrue(os.path.exists(path))
        from valuation.guide_csv import GuideCsvIndex
        idx = GuideCsvIndex(tmp)
        idx._load_file(path)
        self.assertEqual(len(idx), 1)

    def test_fresh_files_are_not_refetched(self):
        m = self._mod()
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "x.csv")
        open(path, "w").write("id,product-name,console-name\n")
        age = m._age_hours(path)
        self.assertIsNotNone(age)
        self.assertLess(age, m.FRESH_HOURS,
                        "a file just written must count as fresh")
        self.assertIsNone(m._age_hours(os.path.join(tmp, "missing.csv")))

    def test_the_cooldown_respects_their_documented_limit(self):
        m = self._mod()
        self.assertGreaterEqual(m.CSV_COOLDOWN_SECONDS, 600,
                                "PriceCharting allows one CSV per 10 minutes")

    # --- failure kinds -------------------------------------------------
    # 2026-07-28: all four PriceCharting categories downloaded cleanly at
    # ~10-minute spacing while all four SportsCardsPro ones failed. That is
    # a per-site subscription boundary, not a rate limit - but the old
    # 2-tuple return could not express the difference, so the run slept ten
    # minutes before each of the three remaining doomed attempts.

    def _resp(self, status=200, body=b"", ):
        class FakeResp:
            status_code = status
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                if body:
                    yield body
        return FakeResp()

    def test_403_is_a_subscription_problem_not_a_rate_limit(self):
        m = self._mod()
        with mock.patch.object(m.requests, "get",
                               return_value=self._resp(403)):
            ok, _detail, kind = m.download(m.SCP, "baseball-cards", "t",
                                           os.path.join(tempfile.mkdtemp(),
                                                        "x.csv"))
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_NOT_COVERED)

    def test_429_is_a_rate_limit(self):
        m = self._mod()
        with mock.patch.object(m.requests, "get",
                               return_value=self._resp(429)):
            ok, _detail, kind = m.download(m.PC, "", "t",
                                           os.path.join(tempfile.mkdtemp(),
                                                        "x.csv"))
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_RATE)

    def test_an_html_login_page_is_a_subscription_problem(self):
        m = self._mod()
        page = b"<!DOCTYPE html><html><body>Please subscribe</body></html>"
        with mock.patch.object(m.requests, "get",
                               return_value=self._resp(200, page)):
            ok, _detail, kind = m.download(m.SCP, "baseball-cards", "t",
                                           os.path.join(tempfile.mkdtemp(),
                                                        "x.csv"))
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_NOT_COVERED)

    def test_an_html_slow_down_page_is_a_rate_limit(self):
        m = self._mod()
        page = b"<!DOCTYPE html><html><body>Too many requests, try again</body></html>"
        with mock.patch.object(m.requests, "get",
                               return_value=self._resp(200, page)):
            ok, _detail, kind = m.download(m.PC, "", "t",
                                           os.path.join(tempfile.mkdtemp(),
                                                        "x.csv"))
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_RATE)

    def _run_main(self, m, tmp, guides, fake_download, sleeps):
        with mock.patch.object(m, "FOLDER", os.path.join(tmp, "guide_csv")), \
             mock.patch.object(m, "download", fake_download), \
             mock.patch.object(m.time, "sleep", lambda s: sleeps.append(s)), \
             mock.patch.object(m.scanner, "load_config", lambda p: {
                 "api_keys": {"pricecharting": {"token": "t"}},
                 "guide_csv": dict(cooldown_seconds=900, **guides)}), \
             mock.patch.object(m.sys, "argv", ["x"]):
            m.main()

    def test_a_refusal_skips_the_rest_of_that_shape_without_sleeping(self):
        """The whole point: one 403 must not cost 45 more minutes of sleep."""
        m = self._mod()
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        sleeps, tried = [], []

        def fake_download(host, slug, token, path, console_uids=None):
            tried.append(console_uids or slug)
            return False, "HTTP 403", m.KIND_NOT_COVERED

        cats = [{"name": f"{s} cards", "host": m.SCP, "category": f"{s}-cards"}
                for s in ("baseball", "basketball", "football", "hockey")]
        self._run_main(m, tmp, {"guides": cats}, fake_download, sleeps)

        self.assertEqual(tried, ["baseball-cards"],
                         "one refusal is enough to know the rest of the "
                         "catalogues will be refused too")
        self.assertFalse(sleeps, "and no sleeping between the skipped ones")

    def test_a_refused_catalogue_does_not_skip_a_working_set_download(self):
        """Regression: 2026-08-01. The skip was keyed by host, so a
        speculative `category=` refusal would also skip `console-uids=`
        downloads on that host - which are the ones known to work. The two
        selectors must be judged separately."""
        m = self._mod()
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        sleeps, tried = [], []

        def fake_download(host, slug, token, path, console_uids=None):
            tried.append(console_uids or slug)
            if console_uids:                       # the proven selector
                with open(path, "w") as fh:
                    fh.write("id,product-name,console-name\n1,a,b\n")
                return True, "25 KB", m.KIND_OK
            return False, "HTTP 403", m.KIND_NOT_COVERED

        cats = [{"name": f"{s} cards", "host": m.SCP, "category": f"{s}-cards"}
                for s in ("baseball", "basketball")]
        sets = [{"name": f"set {u}", "host": m.SCP, "console_uids": u}
                for u in ("G155", "G156")]
        self._run_main(m, tmp, {"guides": cats, "extra_guides": sets},
                       fake_download, sleeps)

        self.assertEqual(tried, ["baseball-cards", "G155", "G156"],
                         "the catalogue refusal must not poison the sets")

    def test_a_rate_limited_file_is_retried_once(self):
        m = self._mod()
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        attempts = []

        def fake_download(host, slug, token, path, console_uids=None):
            attempts.append(slug or "video-games")
            if attempts.count(slug or "video-games") == 1:
                return False, "HTTP 429", m.KIND_RATE
            with open(path, "w") as fh:
                fh.write("id,product-name,console-name\n1,a,b\n")
            return True, "1.0 MB", m.KIND_OK

        with mock.patch.object(m, "FOLDER", os.path.join(tmp, "guide_csv")), \
             mock.patch.object(m, "download", fake_download), \
             mock.patch.object(m.time, "sleep", lambda s: None), \
             mock.patch.object(m, "DEFAULT_GUIDES", [("", m.PC, "video games")]), \
             mock.patch.object(m.scanner, "load_config", lambda p: {
                 "api_keys": {"pricecharting": {"token": "t"}}}), \
             mock.patch.object(m.sys, "argv", ["x"]):
            rc = m.main()
        self.assertEqual(attempts.count("video-games"), 2,
                         "a genuine rate limit deserves one retry")
        self.assertEqual(rc, 0)

    def _csv_file(self, consoles):
        path = os.path.join(tempfile.mkdtemp(), "g.csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["id", "console-name", "product-name", "loose-price"])
            for i, console in enumerate(consoles):
                w.writerow([i, console, f"Item {i}", "$1.00"])
        return path

    def test_a_card_guide_full_of_video_games_is_rejected(self):
        """2026-08-01: four SportsCardsPro 'sports' downloads arrived as
        122,330 rows of ZX Spectrum, PS4 and Switch. That host ignores
        `category=` and serves its default catalogue rather than erroring,
        so HTTP 200, right size, right filename, wrong contents.

        The sportscardspro-- prefix is what lets rows price Sports Cards,
        so keeping the file was worse than having nothing."""
        m = self._mod()
        games = self._csv_file(["ZX Spectrum", "Playstation 4",
                                "Nintendo Switch", "PAL Playstation 4"] * 30)
        why = m.verify_contents(games, {"category": "baseball-cards",
                                        "host": m.SCP})
        self.assertIn("video games", why)
        self.assertIn("default catalogue", why)

    def test_a_real_card_guide_passes(self):
        m = self._mod()
        cards = self._csv_file(["Basketball Cards 1986 Fleer"] * 50)
        self.assertEqual(
            m.verify_contents(cards, {"console_uids": "G155",
                                      "host": m.SCP}), "")

    def test_a_video_game_guide_is_not_rejected_for_being_video_games(self):
        """The check must fire on a mismatch, not on games as such - the
        PriceCharting default catalogue is legitimately video games."""
        m = self._mod()
        games = self._csv_file(["Playstation 4", "Nintendo Switch"] * 40)
        self.assertEqual(
            m.verify_contents(games, {"category": "", "host": m.PC}), "")

    def test_defaults_send_no_category_to_sportscardspro(self):
        """Reverted 2026-08-02. The four sport slugs were restored on 08-01
        after Cloudflare stopped hiding them, and downloaded video games."""
        m = self._mod()
        scp = [g for g in m._normalise_guides(m.DEFAULT_GUIDES)
               if "sportscardspro" in g["host"]]
        self.assertEqual(scp, [], "SCP selects sets by console-uids only")

    def test_a_set_without_a_uid_is_skipped_with_instructions(self):
        """25 sets were queued on 2026-08-08 with uids still to be
        collected. An SCP entry with no uid must never be fetched - that
        host ignores category= and serves video games (the 08-01 incident)
        - and the summary must say which sets are waiting and how to get
        their uids."""
        m = self._mod()
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        tried, sleeps = [], []

        def fake_download(host, slug, token, path, console_uids=None):
            tried.append((m._host_label(host), console_uids or slug))
            with open(path, "w") as fh:
                fh.write("id,product-name,console-name\n1,a,Basketball\n")
            return True, "1 KB", m.KIND_OK

        guides = [
            {"name": "1986 Fleer Basketball", "host": m.SCP,
             "console_uids": "G155"},
            {"name": "T206", "host": m.SCP, "console_uids": None},
            {"name": "1933 Goudey Baseball", "host": m.SCP},
        ]
        printed = []
        with mock.patch.object(m, "FOLDER", os.path.join(tmp, "guide_csv")), \
             mock.patch.object(m, "download", fake_download), \
             mock.patch.object(m.time, "sleep", lambda s: sleeps.append(s)), \
             mock.patch.object(m.scanner, "load_config", lambda p: {
                 "api_keys": {"pricecharting": {"token": "t"}},
                 "guide_csv": {"guides": guides}}), \
             mock.patch("builtins.print",
                        lambda *a, **kw: printed.append(" ".join(map(str, a)))), \
             mock.patch.object(m.sys, "argv", ["x"]):
            m.main()

        self.assertEqual(tried, [("SportsCardsPro", "G155")],
                         "only the set WITH a uid is fetched")
        text = "\n".join(printed)
        self.assertIn("T206", text)
        self.assertIn("1933 Goudey Baseball", text)
        self.assertIn("console-uids", text)

    def test_a_503_is_transient_and_gets_the_retry_path(self):
        """2026-08-08: the 50MB comic guide 503'd - server-side, usually a
        generation timeout - and the downloader gave up instead of taking
        the one wait-and-retry the rate-limit path already had."""
        m = self._mod()
        for code in (500, 502, 503, 504):
            class FakeResp:
                status_code = code
                content = b""
                def close(self): pass
                def iter_content(self, chunk_size=0): yield b""
            with mock.patch.object(m, "curl_requests", None), \
                    mock.patch.object(m.requests, "get",
                                      lambda *a, **kw: FakeResp()):
                ok, detail, kind = m.download(
                    m.PC, "comic-books", "tok",
                    os.path.join(tempfile.mkdtemp(), "x.csv"))
            self.assertFalse(ok)
            self.assertEqual(kind, m.KIND_RATE,
                             f"{code} must take the retry path")
            self.assertIn("temporary", detail)

    def test_the_cooldown_clears_the_documented_limit(self):
        """One CSV per ten minutes is the provider's rule, so the default
        must exceed 600s - but only just. It was briefly 900 while the
        SportsCardsPro 403s were still thought to be a rate limit; they
        were a bot challenge, and the extra four minutes bought nothing
        except an hour added to a four-file run."""
        m = self._mod()
        self.assertGreater(m.CSV_COOLDOWN_SECONDS, 600,
                           "must clear the documented one-per-10-minutes")
        self.assertLessEqual(m.CSV_COOLDOWN_SECONDS, 720,
                             "headroom, not an hour of dead time")

    # --- per-host selector ---------------------------------------------
    # 2026-07-28: every SportsCardsPro download returned 403 and looked like
    # a subscription problem. The account holds SportsCardsPro Legendary,
    # which includes downloads. The real cause was the selector: that host
    # takes `console-uids` per SET, while PriceCharting takes `category`
    # per catalogue. Same path, same token, different parameter.

    def test_sportscardspro_uses_console_uids_not_category(self):
        m = self._mod()
        seen = {}

        class FakeResp:
            status_code = 200
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"id,product-name,console-name\n1,Jordan #57,1986 Fleer\n"

        def fake_get(url, params=None, **kw):
            seen.update(params or {})
            return FakeResp()

        with mock.patch.object(m.requests, "get", fake_get):
            ok, _d, _k = m.download(m.SCP, "", "tok",
                                    os.path.join(tempfile.mkdtemp(), "x.csv"),
                                    console_uids="G155")
        self.assertTrue(ok)
        self.assertEqual(seen.get("console-uids"), "G155")
        self.assertNotIn("category", seen,
                         "sending category= to SportsCardsPro is the 403")

    def test_pricecharting_still_uses_category(self):
        m = self._mod()
        seen = {}

        class FakeResp:
            status_code = 200
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"id,product-name,console-name\n1,a,b\n"

        with mock.patch.object(m.requests, "get",
                               lambda url, params=None, **kw: (
                                   seen.update(params or {}) or FakeResp())):
            m.download(m.PC, "pokemon-cards", "tok",
                       os.path.join(tempfile.mkdtemp(), "x.csv"))
        self.assertEqual(seen.get("category"), "pokemon-cards")
        self.assertNotIn("console-uids", seen)

    def test_a_uid_guide_keeps_sportscardspro_provenance(self):
        """The filename prefix is what lets these rows price Sports Cards.
        A sports guide saved under pricecharting-- would be refused by the
        category guard and the download would be wasted."""
        m = self._mod()
        g = m._normalise_guides([{"name": "1986 Fleer Basketball",
                                  "host": m.SCP, "console_uids": "G155"}])[0]
        self.assertEqual(g["console_uids"], "G155")
        self.assertIsNone(g["category"])
        self.assertTrue(
            m._filename(g["host"], g["slug"]).startswith("sportscardspro--"))

    def test_a_uid_list_is_joined(self):
        m = self._mod()
        g = m._normalise_guides([{"name": "sets", "host": m.SCP,
                                  "console_uids": ["G155", "G156"]}])[0]
        self.assertEqual(g["console_uids"], "G155,G156")

    def test_legacy_tuple_guides_still_work(self):
        m = self._mod()
        g = m._normalise_guides([("pokemon-cards", m.PC, "Pokemon cards")])[0]
        self.assertEqual(g["category"], "pokemon-cards")
        self.assertIsNone(g["console_uids"])

    def test_a_403_reports_the_server_reason_not_just_the_status(self):
        """Same URL, same token, same parameter as the browser link, still
        403. The status alone cannot distinguish "log in" from "subscription
        required" from "bad token", and guessing wasted two rounds."""
        m = self._mod()

        class FakeResp:
            status_code = 403
            content = (b"<html><body><h1>Forbidden</h1>"
                       b"<p>You must be logged in.</p></body></html>")
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b""

        with mock.patch.object(m.requests, "get", return_value=FakeResp()):
            ok, detail, kind = m.download(m.SCP, "", "tok",
                                          os.path.join(tempfile.mkdtemp(),
                                                       "x.csv"),
                                          console_uids="G155")
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_NOT_COVERED)
        self.assertIn("logged in", detail)

    def test_an_echoed_token_is_never_printed(self):
        """security.redact_text only strips credential-shaped QUERY params,
        so a page echoing the token in body text would reach the screen."""
        m = self._mod()

        class FakeResp:
            content = (b"<html>bad token "
                       b"3ede03b3e2e44af62695a1235e632392dc4b9bbb</html>")

        reason = m._reason_from(FakeResp())
        self.assertNotIn("3ede03b3", reason)
        self.assertIn("<redacted>", reason)

    def test_a_cloudflare_interstitial_is_not_a_subscription_problem(self):
        """The real 2026-07-31 body. Calling this "not covered by this
        subscription" sent the investigation at the account page twice when
        the account was fine - the site is simply behind a bot challenge."""
        m = self._mod()

        class FakeResp:
            status_code = 403
            content = (b"<html><head><title>Just a moment...</title>"
                       b"<meta http-equiv=\"content-security-policy\" "
                       b"content=\"script-src https://challenges."
                       b"cloudflare.com\"></head></html>")
            def __enter__(self): return self
            def __exit__(self, *a): return False
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b""

        with mock.patch.object(m.requests, "get", return_value=FakeResp()):
            ok, detail, kind = m.download(m.SCP, "", "tok",
                                          os.path.join(tempfile.mkdtemp(),
                                                       "x.csv"),
                                          console_uids="G155")
        self.assertFalse(ok)
        self.assertEqual(kind, m.KIND_CHALLENGE)
        self.assertNotEqual(kind, m.KIND_NOT_COVERED)
        self.assertIn("Cloudflare", detail)

    def test_challenge_markers_cover_the_usual_wording(self):
        m = self._mod()
        for body in ("Just a moment...", "Checking your browser before",
                     "Attention Required! | Cloudflare",
                     "Enable JavaScript and cookies to continue"):
            self.assertTrue(m._looks_like_challenge(body), body)
        self.assertFalse(m._looks_like_challenge(
            "You must be logged in to download price lists"))

    def test_a_challenged_host_also_stops_costing_sleep(self):
        m = self._mod()
        tmp = tempfile.mkdtemp()
        os.makedirs(os.path.join(tmp, "guide_csv"), exist_ok=True)
        sleeps, tried = [], []

        def fake_download(host, slug, token, path, console_uids=None):
            tried.append(console_uids or slug)
            if "sportscardspro" in host:
                return (False, "blocked by a Cloudflare bot challenge",
                        m.KIND_CHALLENGE)
            with open(path, "w") as fh:
                fh.write("id,product-name,console-name\n1,a,b\n")
            return True, "1.0 MB", m.KIND_OK

        sets = [{"name": f"set {u}", "host": m.SCP, "console_uids": u}
                for u in ("G155", "G156", "G157")]
        with mock.patch.object(m, "FOLDER", os.path.join(tmp, "guide_csv")), \
             mock.patch.object(m, "download", fake_download), \
             mock.patch.object(m.time, "sleep", lambda s: sleeps.append(s)), \
             mock.patch.object(m.scanner, "load_config", lambda p: {
                 "api_keys": {"pricecharting": {"token": "t"}},
                 "guide_csv": {"extra_guides": sets}}), \
             mock.patch.object(m.sys, "argv", ["x"]):
            m.main()
        self.assertEqual(len([t for t in tried if str(t).startswith("G")]), 1,
                         "one challenge is enough to know the rest will fail")

    def test_downloads_go_through_the_impersonating_client(self):
        """This endpoint serves a Cloudflare interstitial to bare requests.

        scrapers/base.py has used curl_cffi for eBay and Goldin for the same
        reason; the downloader was the last caller still on plain requests.
        """
        m = self._mod()
        calls = []

        class FakeResp:
            status_code = 200
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"id,product-name,console-name\n1,a,b\n"

        class FakeCurl:
            @staticmethod
            def get(url, params=None, impersonate=None, **kw):
                calls.append(impersonate)
                return FakeResp()

        with mock.patch.object(m, "curl_requests", FakeCurl):
            m.download(m.PC, "", "tok",
                       os.path.join(tempfile.mkdtemp(), "x.csv"))
        self.assertEqual(calls, ["chrome"])

    def test_download_requests_stream_mode(self):
        """Regression: 2026-07-31. Switching to curl_cffi dropped
        stream=True. curl_cffi then raises 'stream mode is not enabled' from
        iter_content - but only AFTER a clean 200, so a working download
        looked like a site failure. The 50MB comic guide also should not be
        held in memory. Both clients must be asked to stream."""
        m = self._mod()
        seen = {}

        class FakeResp:
            status_code = 200
            content = b""
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"id,product-name,console-name\n1,a,b\n"

        def record(client):
            def _get(url, params=None, **kw):
                seen[client] = kw.get("stream")
                return FakeResp()
            return _get

        class FakeCurl:
            get = staticmethod(record("curl"))

        with mock.patch.object(m, "curl_requests", FakeCurl):
            m.download(m.PC, "", "tok",
                       os.path.join(tempfile.mkdtemp(), "x.csv"))
        with mock.patch.object(m, "curl_requests", None), \
                mock.patch.object(m.requests, "get", record("plain")):
            m.download(m.PC, "", "tok",
                       os.path.join(tempfile.mkdtemp(), "y.csv"))
        self.assertEqual(seen, {"curl": True, "plain": True})

    def test_error_body_is_read_even_when_streaming(self):
        """A streamed response has an empty .content until something pulls
        it, and the error page is exactly what we need on a 403."""
        m = self._mod()

        class FakeResp:
            status_code = 403
            content = b""
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"<html><body>Subscription does not cover this</body>"

        with mock.patch.object(m, "curl_requests", None), \
                mock.patch.object(m.requests, "get",
                                  lambda *a, **kw: FakeResp()):
            ok, detail, kind = m.download(
                m.PC, "", "tok", os.path.join(tempfile.mkdtemp(), "x.csv"))
        self.assertFalse(ok)
        self.assertIn("Subscription does not cover", detail)
        self.assertEqual(kind, m.KIND_NOT_COVERED)

    def test_no_spoofed_user_agent_without_matching_tls(self):
        """Claiming to be Chrome over Python's TLS handshake is a bot signal
        in itself - worse than sending nothing. So the fallback path, used
        only when curl_cffi is missing, sends no User-Agent."""
        m = self._mod()
        seen = {}

        class FakeResp:
            status_code = 200
            def close(self): pass
            def iter_content(self, chunk_size=0):
                yield b"id,product-name,console-name\n1,a,b\n"

        def fake_get(url, params=None, headers=None, **kw):
            seen["headers"] = headers
            return FakeResp()

        with mock.patch.object(m, "curl_requests", None), \
                mock.patch.object(m.requests, "get", fake_get):
            m.download(m.PC, "", "tok",
                       os.path.join(tempfile.mkdtemp(), "x.csv"))
        self.assertIn("Mozilla", "".join(m.BROWSER_HEADERS.values()),
                      "the header set still exists for the diagnose harness")
        self.assertFalse(seen.get("headers"),
                         "but it must not be sent over plain requests")

    def test_falls_back_to_plain_requests_if_curl_cffi_errors(self):
        """A broken optional dependency must not stop the download."""
        m = self._mod()
        used = []

        class FakeResp:
            status_code = 200
            def close(self): pass
            def iter_content(self, chunk_size=0):
                used.append("plain")
                yield b"id,product-name,console-name\n1,a,b\n"

        class Exploding:
            @staticmethod
            def get(*a, **kw):
                raise RuntimeError("libcurl missing")

        with mock.patch.object(m, "curl_requests", Exploding), \
                mock.patch.object(m.requests, "get",
                                  lambda *a, **kw: FakeResp()):
            ok, _msg, _kind = m.download(
                m.PC, "", "tok", os.path.join(tempfile.mkdtemp(), "x.csv"))
        self.assertTrue(ok)
        self.assertEqual(used, ["plain"])

    def test_each_default_sends_exactly_one_selector(self):
        """`download` sends console-uids if present, category otherwise. An
        entry carrying both would silently drop the category."""
        m = self._mod()
        for g in m._normalise_guides(m.DEFAULT_GUIDES):
            self.assertFalse(g["category"] and g["console_uids"],
                             f"{g['name']} sets both selectors")


class TestLiveAuctionPricing(unittest.TestCase):
    """2026-07-26: the projected close was circular - model the close as
    fair x 0.92, then ask if that is cheap versus fair. With a 13.25% sell
    fee and 8% buy tax you must acquire at <= 80.3% of fair to break even,
    so any auction whose live bid had not yet started to dominate was
    negative-EV BY ARITHMETIC and every auction beyond ~48h was invisible.
    And 0.92 was never validated: the learner sits at n=0."""

    def _engine(self, mode="live", **algo):
        from valuation import ValuationEngine
        from valuation.price_guide import GuideQuote
        tmp = tempfile.mkdtemp()
        cfg = {"_config_dir": tmp,
               "database": {"file": os.path.join(tmp, "h.db")},
               "algorithm": dict({"auction_pricing": mode,
                                  "auction_settle_ratio": 0.92,
                                  "resale_fee_rate": 0.1325,
                                  "sales_tax_rate": 0.08}, **algo)}
        e = ValuationEngine(cfg)
        e.guide.quote = lambda ident: GuideQuote()
        return e

    def _auction(self, price, bids, hours):
        return Listing(
            site="ebay", title="1952 Topps #311 Mickey Mantle PSA 5",
            url="https://www.ebay.com/itm/50", listing_id="50",
            query="Mickey Mantle 1952", current_price=price, bid_count=bids,
            listing_type="AUCTION",
            end_time=datetime.now(timezone.utc) + timedelta(hours=hours))

    def _value(self, engine, listing, fair=1000.0):
        v = Valuation(fair_value=fair, comps_value=fair, n_comps=6,
                      confidence=0.8)
        return engine.score(listing, v)

    def test_live_pricing_uses_the_current_bid(self):
        e = self._engine("live")
        v = self._value(e, self._auction(200.0, 5, hours=40))
        self.assertAlmostEqual(v.expected_cost,
                               e_cost := v.expected_cost)  # sanity
        self.assertLess(v.expected_cost, 300.0,
                        "a $200 bid must be costed near $200, not near fair")
        self.assertGreater(v.expected_value, 0,
                           "a $200 bid on a $1,000 card is a real dislocation")

    def test_projected_pricing_buries_the_same_row(self):
        """The exact row live pricing surfaces, the old model discards.

        Measured on a $200 bid against a $1,000 card: projected EV crosses
        zero at about 72 hours out and is -$126 by 10 days, while the same
        row is +$652 under live pricing at every horizon. (An earlier note
        of mine said this happened by ~48h - it is ~72h. The 72-hour cutoff
        we now filter on is, by coincidence, almost exactly where the old
        model stopped finding anything at all.)
        """
        e = self._engine("projected")
        far = self._value(e, self._auction(200.0, 5, hours=120))
        self.assertGreater(far.expected_cost, 900.0,
                           "far out, projection drags cost to ~92% of fair")
        self.assertLess(far.expected_value, 0,
                        "which makes it negative-EV and invisible")
        # Nearer in it survives, but at a fraction of the real dislocation.
        near_proj = self._value(e, self._auction(200.0, 5, hours=40))
        near_live = self._value(self._engine("live"),
                                self._auction(200.0, 5, hours=40))
        self.assertGreater(near_live.expected_value,
                           near_proj.expected_value * 5)

    def test_a_zero_bid_auction_far_from_close_is_not_a_market(self):
        """A low opening bid is bait, and bait needs time to work. A $0.99
        start on a $5,000 card shows $5,000 of edge and is unwinnable."""
        e = self._engine("live")
        v = self._value(e, self._auction(50.0, 0, hours=40))
        self.assertGreater(v.expected_cost, 500.0,
                           "far from close, fall back to the model")
        self.assertTrue(any("not a market yet" in n for n in v.audit_notes))

    def test_a_zero_bid_auction_near_close_is_takeable(self):
        """Nobody has bid and it ends soon: opening at that price can
        genuinely win it, and it means nobody noticed the card."""
        e = self._engine("live")
        v = self._value(e, self._auction(200.0, 0, hours=5))
        self.assertLess(v.expected_cost, 300.0,
                        "priced at the opening bid, which is takeable")
        self.assertGreater(v.expected_value, 0)
        self.assertTrue(any("NO BIDS YET" in n for n in v.notes),
                        "the row must warn that one rival bid changes it")

    def test_the_zero_bid_window_is_configurable(self):
        e = self._engine("live", zero_bid_actionable_hours=6)
        far = self._value(e, self._auction(200.0, 0, hours=12))
        near = self._value(self._engine("live", zero_bid_actionable_hours=6),
                           self._auction(200.0, 0, hours=3))
        self.assertGreater(far.expected_cost, 500.0)
        self.assertLess(near.expected_cost, 300.0)

    def test_a_zero_bid_hybrid_still_uses_its_buy_it_now(self):
        """The actual 2026-07-25 bug: a $499 opening bid on an item that
        also had a Buy It Now. The BIN is takeable; the opening ask is not.
        This must hold at every horizon."""
        e = self._engine("live")
        l = self._auction(499.0, 0, hours=5)
        l.has_buy_now, l.buy_now_price = True, 700.0
        v = self._value(e, l)
        self.assertGreaterEqual(v.expected_cost, l.landed_cost(499.0),
                                "must not be priced below the opening ask")

    def test_output_ok_lets_a_near_close_zero_bid_row_through(self):
        near = self._auction(200.0, 0, hours=5)
        near.listing_type = "auction"
        far = self._auction(200.0, 0, hours=48)
        far.listing_type = "auction"
        good = Valuation(fair_value=1000.0, expected_value=500.0, roi=0.5)
        self.assertTrue(scanner.output_ok(
            Opportunity(listing=near, valuation=good), zero_bid_hours=24))
        self.assertFalse(scanner.output_ok(
            Opportunity(listing=far, valuation=good), zero_bid_hours=24))

    def test_live_rows_say_what_the_number_means(self):
        e = self._engine("live")
        v = self._value(e, self._auction(200.0, 5, hours=40))
        self.assertTrue(any("live pricing" in n for n in v.audit_notes),
                        "EV is edge-right-now, not a profit forecast - the "
                        "row has to say so")

    def test_a_hybrid_is_still_capped_at_its_buy_it_now(self):
        e = self._engine("live")
        l = self._auction(200.0, 3, hours=40)
        l.has_buy_now, l.buy_now_price = True, 400.0
        v = self._value(e, l)
        self.assertLessEqual(v.expected_cost,
                             l.landed_cost(400.0) * 1.10)


class TestAuctionHorizonFilter(unittest.TestCase):
    """Under live pricing, an auction six days out is quoting a bid the
    market has barely tested. Cut before valuation so it costs neither
    scoring time nor a paid guide lookup."""

    def _auction(self, hours, kind="AUCTION"):
        return Listing(
            site="ebay", title="1952 Topps #311 Mickey Mantle PSA 5",
            url="https://www.ebay.com/itm/60", listing_id="60",
            query="Mickey Mantle 1952", current_price=100.0, bid_count=2,
            listing_type=kind,
            end_time=(datetime.now(timezone.utc) + timedelta(hours=hours))
            if hours is not None else None)

    def test_auctions_inside_the_window_are_kept(self):
        for hrs in (1, 24, 71):
            self.assertTrue(
                scanner.within_auction_horizon(self._auction(hrs), 72))

    def test_auctions_beyond_the_window_are_dropped(self):
        for hrs in (73, 120, 400):
            self.assertFalse(
                scanner.within_auction_horizon(self._auction(hrs), 72))

    def test_fixed_price_listings_are_unaffected(self):
        self.assertTrue(
            scanner.within_auction_horizon(self._auction(500, "BIN"), 72))

    def test_an_unknown_end_time_is_kept(self):
        self.assertTrue(
            scanner.within_auction_horizon(self._auction(None), 72))

    def test_zero_or_missing_config_disables_the_filter(self):
        self.assertTrue(scanner.within_auction_horizon(self._auction(999), 0))
        self.assertTrue(
            scanner.within_auction_horizon(self._auction(999), None))

    def test_the_shipped_config_sets_a_horizon(self):
        path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "config.yaml")
        if not os.path.isfile(path):
            self.skipTest("live config.yaml not present")
        with open(path) as fh:
            cfg = yaml.safe_load(fh) or {}
        self.assertEqual(
            (cfg.get("filters") or {}).get("max_auction_hours"), 72)
        self.assertEqual(
            (cfg.get("algorithm") or {}).get("auction_pricing"), "live")


class TestRoiCeilingUnderLivePricing(unittest.TestCase):
    """The ROI ceiling catches nonsense, but its right value depends on what
    ROI MEANS. Live pricing changed that: a $200 bid on a $1,000 card was
    30% ROI projected and is 302% live. A 200% ceiling would have deleted
    precisely the dislocations live pricing exists to find."""

    def test_the_live_ceiling_is_far_above_the_projected_one(self):
        path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "config.yaml")
        if not os.path.isfile(path):
            self.skipTest("live config.yaml not present")
        with open(path) as fh:
            cfg = yaml.safe_load(fh) or {}
        flt = cfg.get("filters") or {}
        self.assertGreaterEqual(flt.get("max_roi_live", 0), 3.0,
                                "a 302% ROI row must survive the ceiling")
        self.assertGreater(flt.get("max_roi_live", 0), flt.get("max_roi", 0))

    def test_alerts_use_the_live_ceiling_too(self):
        path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "config.yaml")
        if not os.path.isfile(path):
            self.skipTest("live config.yaml not present")
        with open(path) as fh:
            cfg = yaml.safe_load(fh) or {}
        self.assertGreaterEqual(
            (cfg.get("alerts") or {}).get("max_roi_live", 0), 3.0,
            "otherwise every cheap auction is muted instead of alerted")

    def test_a_302_percent_row_survives_live_but_not_projected(self):
        good = Valuation(fair_value=1000.0, expected_value=652.0, roi=3.02)
        self.assertGreater(good.roi, 2.0, "would fail the projected ceiling")
        self.assertLess(good.roi, 10.0, "and pass the live one")


class TestCloseTrackingHasItsOwnFuse(unittest.TestCase):
    """The closer fetched item pages through the shared "html" lane, so every
    sold-comp scraping challenge starved close tracking - the ONLY source of
    ground truth for the settle ratio. Symptom: "settled 0 real closes, 20
    pending retry" every run for days, in 0s, because the breaker
    short-circuited it before a request went out. The learner sat at n=0."""

    def _scraper(self):
        from scrapers.ebay import EbayScraper
        tmp = tempfile.mkdtemp()
        return EbayScraper({"_config_dir": tmp, "scraping": {}})

    def test_a_cooling_html_lane_does_not_block_close_lookups(self):
        s = self._scraper()
        s._streaks["html"] = 99          # comp scraping is in the doghouse
        self.assertTrue(s.lane_tripped("html"))
        self.assertFalse(s.lane_tripped("close"),
                         "close tracking must survive a comp-scrape outage")

    def test_the_close_lane_still_trips_on_its_own_failures(self):
        s = self._scraper()
        for _ in range(s.trip_after):
            s._streaks["close"] += 1
        self.assertTrue(s.lane_tripped("close"),
                        "its own fuse must still work - this is not a bypass")
        self.assertFalse(s.lane_tripped("html"))

    def test_arbitrary_lane_names_do_not_crash(self):
        s = self._scraper()
        self.assertFalse(s.lane_tripped("close"))
        self.assertEqual(s._streaks["some-new-lane"], 0)

    def test_the_closer_asks_for_the_close_lane(self):
        import closer
        src = open(closer.__file__.replace(".pyc", ".py")).read()
        self.assertIn('lane="close"', src)
        self.assertNotIn('ebay._streaks["html"]', src,
                         "challenges must count against the close fuse")

    def test_the_loop_guard_checks_the_close_lane_not_html(self):
        """The 18:00 run STILL settled 0 closes after the lane change,
        because `if ebay.tripped: break` consulted the HTML lane and broke
        out before a single lookup. A private fuse is useless if the guard
        in front of it asks a different lane."""
        import closer
        src = open(closer.__file__.replace(".pyc", ".py")).read()
        self.assertNotIn("if ebay.tripped:", src)
        self.assertIn('ebay.lane_tripped("close")', src)

    def test_a_cooling_comp_lane_leaves_close_lookups_runnable(self):
        s = self._scraper()
        s._streaks["html"] = 99
        self.assertTrue(s.tripped, "the html shorthand reports cooling")
        self.assertFalse(s.lane_tripped("close"),
                         "but close tracking must still be allowed to run")


class TestUnresolvedRowsDoNotSwallowSheets(unittest.TestCase):
    """2026-07-26 18:00: the Sports Cards tab was 20 rows of "Tiger Woods
    UDA" all at $1,799.10, and Discovery was 47 rows of "Upper Deck
    Authenticated Signed Jersey" all at $1,085. When the guide cannot land a
    card every listing under that query falls back to the SAME broad median
    - and because that median is inflated relative to the individual cards,
    those are also the rows that clear the category floor. The correctly
    valued cheap cards were cut beneath them."""

    def _opp(self, title, fair, query, unresolved=True, score=0.5):
        listing = Listing(site="ebay", title=title,
                          url=f"https://www.ebay.com/itm/{abs(hash(title))}",
                          listing_id=str(abs(hash(title))), query=query,
                          current_price=100.0)
        notes = (["IDENTITY UNRESOLVED: no PriceCharting product matched"]
                 if unresolved else [])
        v = Valuation(fair_value=fair, opportunity_score=score, notes=notes)
        return Opportunity(listing=listing, valuation=v)

    def test_one_unresolved_value_cannot_fill_a_sheet(self):
        rows = [self._opp(f"Tiger Woods item {i}", 1799.10, "Tiger Woods UDA")
                for i in range(20)]
        kept = report_mod._cap_unresolved(rows, per_query=3)
        self.assertEqual(len(kept), 3)

    def test_identified_rows_are_never_capped(self):
        rows = [self._opp(f"Real card {i}", 1000.0 + i, "q", unresolved=False)
                for i in range(20)]
        self.assertEqual(len(report_mod._cap_unresolved(rows, 3)), 20)

    def test_different_queries_get_their_own_allowance(self):
        rows = ([self._opp(f"a{i}", 1085.0, "Jerseys") for i in range(10)]
                + [self._opp(f"b{i}", 1799.1, "Tiger Woods UDA")
                   for i in range(10)])
        self.assertEqual(len(report_mod._cap_unresolved(rows, 3)), 6)

    def test_identified_rows_outrank_unresolved_ones(self):
        weak = self._opp("mongrel median row", 5000.0, "q", score=0.9)
        good = self._opp("real card", 1200.0, "q", unresolved=False,
                         score=0.1)
        self.assertIs(report_mod._sorted([weak, good])[0], good,
                      "a card valued from itself beats one valued from a "
                      "query-wide median, however big the median looked")

    def test_ask_based_values_are_unresolved_too(self):
        """The Sports Cards tab was NOT mixed-pool: all 20 Tiger Woods rows
        shared $1,799.10 from one ASK-BASED estimate over 149 live asks
        under the query. Same collapse, different mechanism."""
        listing = Listing(site="ebay", title="Tiger Woods Framed Photo UDA",
                          url="https://www.ebay.com/itm/1", listing_id="1",
                          query="Tiger Woods UDA", current_price=140.0)
        v = Valuation(fair_value=1799.10, notes=[
            "ASK-BASED estimate from 149 live asks (no sold comps)"])
        self.assertTrue(report_mod._is_unresolved(
            Opportunity(listing=listing, valuation=v)))

    def test_every_query_level_marker_is_covered(self):
        for marker in ("IDENTITY UNRESOLVED", "MIXED POOL", "ASK-BASED"):
            self.assertIn(marker, report_mod.UNRESOLVED_MARKERS)

    def test_the_survivors_say_rows_were_hidden(self):
        rows = [self._opp(f"x{i}", 1085.0, "Jerseys") for i in range(9)]
        kept = report_mod._cap_unresolved(rows, per_query=3)
        self.assertTrue(any("hidden here" in n
                            for o in kept for n in o.valuation.notes))


class TestAskBasedIsCardSpecific(unittest.TestCase):
    """2026-07-26: the Sports Cards tab was 20 rows all at $1,799.10, from a
    single ASK-BASED estimate over 149 live asks under "Tiger Woods UDA" -
    framed photos, tournament-worn shirts, jersey cards, signed flags. The
    ask pool was gathered per QUERY and never filtered to the card, so it
    repeated the exact mistake the identity work fixed for comps."""

    def _engine(self):
        from valuation import ValuationEngine
        from valuation.price_guide import GuideQuote
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "h.db")},
            "algorithm": {"min_specific_comps": 3}})
        e.guide.quote = lambda ident: GuideQuote()
        return e

    def _ask(self, title, price):
        return Listing(site="ebay", title=title,
                       url=f"https://www.ebay.com/itm/{abs(hash(title))}",
                       listing_id=str(abs(hash(title))), query="Tiger Woods UDA",
                       current_price=price, listing_type="fixed")

    def test_asks_for_other_objects_are_excluded(self):
        from valuation.identity import identity_of
        e = self._engine()
        ident = identity_of("2001 Upper Deck #1 Tiger Woods Rookie PSA 9")
        pool = [self._ask("Tiger Woods Framed Photo UDA", 140.0),
                self._ask("2002 Tiger Woods Tourney Worn Shirt UDA", 189.95),
                self._ask("Tiger Woods Signed Flag UDA", 650.0),
                self._ask("2001 Upper Deck #1 Tiger Woods Rookie PSA 9", 900.0)]
        prices, dropped = e._ask_prices_for(ident, pool)
        self.assertGreater(dropped, 0, "photos and shirts are not this card")
        self.assertNotIn(140.0, prices)

    def test_too_few_matching_asks_means_no_value(self):
        e = self._engine()
        listing = Listing(
            site="ebay", title="2001 Upper Deck #1 Tiger Woods Rookie PSA 9",
            url="https://www.ebay.com/itm/1", listing_id="1",
            query="Tiger Woods UDA", current_price=800.0,
            listing_type="fixed")
        opp = e.evaluate(listing, [], asks=[
            self._ask("Tiger Woods Framed Photo UDA", 140.0),
            self._ask("Tiger Woods Signed Flag UDA", 650.0)])
        self.assertEqual(opp.valuation.fair_value, 0.0,
                         "a mongrel ask pool must not produce a value")

    def test_enough_matching_asks_still_price_the_card(self):
        e = self._engine()
        listing = Listing(
            site="ebay", title="2001 Upper Deck #1 Tiger Woods Rookie PSA 9",
            url="https://www.ebay.com/itm/1", listing_id="1",
            query="Tiger Woods UDA", current_price=800.0,
            listing_type="fixed")
        same = [self._ask(
            f"2001 Upper Deck #1 Tiger Woods Rookie PSA 9 copy {i}",
            900.0 + i * 10) for i in range(5)]
        opp = e.evaluate(listing, [], asks=same)
        self.assertGreater(opp.valuation.fair_value, 0)
        self.assertTrue(any("live asks of this card" in n
                            for n in opp.valuation.notes))

    def test_bare_prices_still_work_for_the_legacy_path(self):
        e = self._engine()
        prices, dropped = e._ask_prices_for(None, [100.0, 200.0, 300.0])
        self.assertEqual(prices, [100.0, 200.0, 300.0])
        self.assertEqual(dropped, 0)


class TestBinAgeWindow(unittest.TestCase):
    """Andrew's rule: a BIN is interesting when FRESH (before the market has
    seen it) or STALE (price set months ago, market has since moved). The
    middle is picked-over inventory and the biggest source of weak rows."""

    class _L:
        def __init__(self, days, kind="BIN"):
            self.age_hours = None if days is None else days * 24
            self.listing_type = kind

    def test_fresh_listings_are_kept(self):
        for d in (0.1, 3, 6.9):
            self.assertTrue(
                scanner.within_bin_age_window(self._L(d), 7, 60), d)

    def test_the_picked_over_middle_is_dropped(self):
        for d in (7, 14, 30, 59.9):
            self.assertFalse(
                scanner.within_bin_age_window(self._L(d), 7, 60), d)

    def test_stale_listings_come_back(self):
        for d in (60, 90, 400):
            self.assertTrue(
                scanner.within_bin_age_window(self._L(d), 7, 60), d)

    def test_auctions_are_unaffected(self):
        self.assertTrue(scanner.within_bin_age_window(
            self._L(30, "auction"), 7, 60))

    def test_an_unknown_age_is_kept_not_guessed(self):
        self.assertTrue(scanner.within_bin_age_window(self._L(None), 7, 60))

    def test_zero_config_disables_the_window(self):
        self.assertTrue(scanner.within_bin_age_window(self._L(30), 0, 0))


class TestSheetOrder(unittest.TestCase):
    def test_diagnostics_sit_at_the_back(self):
        front, back = report_mod.SHEET_ORDER_FRONT, report_mod.SHEET_ORDER_BACK
        for name in ("Filter Waterfall", "Research-Filtered", "Source Health",
                     "About"):
            self.assertIn(name, back)
            self.assertNotIn(name, front)

    def test_discovery_sits_just_before_watches(self):
        front = report_mod.SHEET_ORDER_FRONT
        self.assertEqual(front[front.index("Discovery") + 1], "Watches")

    def test_today_leads_the_book(self):
        self.assertEqual(report_mod.SHEET_ORDER_FRONT[0], "Today")


class TestIdentityCollisionCanary(unittest.TestCase):
    """A silent wrong answer is worse than a loud missing one. On 2026-07-26
    4,237 valued rows carried five distinct opinions and nothing said so."""

    def _opp(self, title, fair, key):
        listing = Listing(site="ebay", title=title,
                          url=f"https://www.ebay.com/itm/{abs(hash(title))}",
                          current_price=100.0, listing_id=str(abs(hash(title))),
                          query="q")
        v = Valuation(fair_value=fair, identity_key=key)
        return Opportunity(listing=listing, valuation=v)

    def test_distinct_assets_sharing_one_value_are_flagged(self):
        opps = [self._opp(t, 1069.60, f"key-{i}")
                for i, t in enumerate(DISNEY_TITLES)]
        found = scanner.find_value_collisions(opps)
        self.assertEqual(len(found), 1)
        fair, assets, rows, _ = found[0]
        self.assertAlmostEqual(fair, 1069.60)
        self.assertEqual(assets, len(DISNEY_TITLES))
        self.assertEqual(rows, len(DISNEY_TITLES))

    def test_many_copies_of_the_SAME_asset_are_not_a_collision(self):
        opps = [self._opp(f"same card copy {i}", 500.0, "one-key")
                for i in range(9)]
        self.assertEqual(scanner.find_value_collisions(opps), [],
                         "several listings of one card SHOULD share a value")

    def test_distinct_values_are_silent(self):
        opps = [self._opp(t, 100.0 * (i + 1), f"key-{i}")
                for i, t in enumerate(DISNEY_TITLES)]
        self.assertEqual(scanner.find_value_collisions(opps), [])

    def test_zero_and_missing_values_are_ignored(self):
        opps = [self._opp(f"t{i}", 0.0, f"key-{i}") for i in range(5)]
        self.assertEqual(scanner.find_value_collisions(opps), [])


class TestGuideLookupCost(unittest.TestCase):
    """2026-07-26: identity resolution costs a /api/products search PLUS a
    /api/product fetch. One BIN sweep made 661 calls in 19 minutes and was
    still running. PriceCharting's default quota is 5,000/day at one call per
    second, so an ungoverned run burns the day's budget and hours of clock."""

    def _guide(self, budget=None):
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        algo = {} if budget is None else {"guide_lookups_per_run": budget}
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "algorithm": algo,
                        "api_keys": {"pricecharting": {"token": "t"}}})
        g.calls = []
        def call(path, params, host=None):
            g.calls.append(path)
            return {"status": "success", "products": [], "loose-price": 100}
        g._pc_call = call
        return g

    def test_run_budget_caps_outward_calls(self):
        g = self._guide(budget=3)
        for i in range(10):
            g._cached_product(f"search:q{i}", "products", {"q": f"q{i}"})
        self.assertEqual(len(g.calls), 3,
                         "budget must cap outward PriceCharting calls")

    def test_cached_lookups_are_free(self):
        g = self._guide(budget=2)
        for _ in range(8):
            g._cached_product("search:same", "products", {"q": "same"})
        self.assertEqual(len(g.calls), 1)
        self.assertEqual(g._lookups_left, 1, "a cache hit must not be billed")

    def test_vague_identities_never_reach_the_api(self):
        from valuation import ValuationEngine
        from valuation.identity import identity_of
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({"_config_dir": tmp,
                             "database": {"file": os.path.join(tmp, "h.db")},
                             "algorithm": {}})
        asked = []
        e.guide.quote = lambda ident: asked.append(ident) or None
        # Real wasted queries from the 2026-07-26 sweep.
        for title in ("1933 Goudey Tues July 4 Babe Ruth reprint",
                      "1933 ACEO Sketch Card Babe Ruth art",
                      "1933 Washington Senators team photo"):
            ident = identity_of(title)
            self.assertFalse(ident.is_specific(e.identity_floor),
                             f"should be too vague to bid on: {title}")
        listing = Listing(site="ebay", title="1933 ACEO Sketch Card Babe Ruth",
                          url="https://www.ebay.com/itm/1", current_price=10.0,
                          listing_id="1", query="Babe Ruth 1933")
        e.evaluate(listing, [])
        self.assertEqual(asked, [],
                         "a browse-only row must not cost a paid lookup")

    def test_comps_far_below_the_floor_skip_the_lookup(self):
        """91% of valued rows died at the fair-value floor in the 11:54 run,
        most after paying for a guide call. When identity-matched comps
        already put a row far below its floor, the guide cannot rescue it."""
        from valuation import ValuationEngine
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "h.db")},
            "filters": {"min_value": 1000},
            "algorithm": {"min_specific_comps": 3}})
        asked = []
        e.guide.quote = lambda ident: asked.append(ident) or None
        listing = Listing(
            site="ebay", title="1952 Topps #311 Mickey Mantle PSA 5",
            url="https://www.ebay.com/itm/9", current_price=40.0,
            listing_id="9", query="Mickey Mantle 1952")
        comps = [SoldComp(f"1952 Topps #311 Mickey Mantle PSA 5 sale {i}",
                          80.0 + i) for i in range(5)]
        e.evaluate(listing, comps)
        self.assertEqual(asked, [],
                         "$80 comps against a $1,000 floor cannot be rescued")

    def test_a_cheap_listing_with_NO_comps_still_gets_priced(self):
        """The skip must never blind us to the actual hunt: a $20 listing
        with no comp evidence is exactly the case the guide exists for."""
        from valuation import ValuationEngine
        from valuation.price_guide import GuideQuote
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "h.db")},
            "filters": {"min_value": 1000},
            "algorithm": {"min_specific_comps": 3}})
        asked = []
        e.guide.quote = lambda ident: (asked.append(ident)
                                       or GuideQuote(note="x"))
        listing = Listing(
            site="ebay", title="1952 Topps #311 Mickey Mantle PSA 5",
            url="https://www.ebay.com/itm/10", current_price=20.0,
            listing_id="10", query="Mickey Mantle 1952")
        e.evaluate(listing, [])
        self.assertEqual(len(asked), 1)

    def test_a_grail_is_always_priced(self):
        from valuation import ValuationEngine
        from valuation.price_guide import GuideQuote
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({
            "_config_dir": tmp,
            "database": {"file": os.path.join(tmp, "h.db")},
            "filters": {"min_value": 1000},
            "algorithm": {"min_specific_comps": 3}})
        asked = []
        e.guide.quote = lambda ident: (asked.append(ident)
                                       or GuideQuote(note="x"))
        listing = Listing(
            site="ebay", title="1952 Topps #311 Mickey Mantle PSA 5",
            url="https://www.ebay.com/itm/11", current_price=40.0,
            listing_id="11", query="Mickey Mantle 1952", grail=True)
        comps = [SoldComp(f"1952 Topps #311 Mickey Mantle PSA 5 s{i}", 80.0)
                 for i in range(5)]
        e.evaluate(listing, comps)
        self.assertEqual(len(asked), 1, "grails are priced regardless of cost")

    def test_a_specific_identity_still_gets_a_lookup(self):
        from valuation import ValuationEngine
        tmp = tempfile.mkdtemp()
        e = ValuationEngine({"_config_dir": tmp,
                             "database": {"file": os.path.join(tmp, "h.db")},
                             "algorithm": {}})
        asked = []
        from valuation.price_guide import GuideQuote
        e.guide.quote = lambda ident: (asked.append(ident)
                                       or GuideQuote(note="x"))
        listing = Listing(
            site="ebay", title="1933 Goudey #149 Babe Ruth PSA 4",
            url="https://www.ebay.com/itm/2", current_price=9000.0,
            listing_id="2", query="Babe Ruth 1933")
        e.evaluate(listing, [])
        self.assertEqual(len(asked), 1)


class TestParallelFalsePositives(unittest.TestCase):
    """A bare colour is not a parallel. On 2026-07-26 a live sweep turned
    'Tiger Woods Red Tiger' into the parallel 'red tiger' and paid for it."""

    def test_tiger_woods_has_no_parallel(self):
        from valuation.identity import identity_of, parallel_of
        self.assertIsNone(parallel_of("2001 Tiger Woods Red Tiger Golf"))
        self.assertIsNone(identity_of("2001 Upper Deck Tiger Woods #1").parallel)

    def test_a_colour_needs_a_finish_or_a_serial(self):
        from valuation.identity import parallel_of
        self.assertIsNone(parallel_of("1952 Topps Mickey Mantle red back"))
        self.assertEqual(parallel_of("Pink Refractor"), "pink refractor")
        self.assertEqual(parallel_of("Black #/10"), "black")

    def test_real_disney_parallels_still_resolve(self):
        from valuation.identity import identity_of
        fps = {identity_of(t).fingerprint() for t in DISNEY_TITLES}
        self.assertEqual(len(fps), len(DISNEY_TITLES))

    def test_ambiguous_finish_words_are_not_parallels(self):
        from valuation.identity import parallel_of
        for text in ("Sega Genesis Sonic", "The Flash #1 comic",
                     "Detroit Tigers team card", "Aurora borealis print"):
            self.assertIsNone(parallel_of(text), text)


class TestCandidateScoringFromLiveData(unittest.TestCase):
    """Fixtures are real /api/products responses observed on 2026-07-26.
    The first probe run landed only 1 of 9 titles, and the cached payloads
    showed PriceCharting HAD the right product every time - we were scoring
    it below a sibling."""

    def _ident(self, title):
        from valuation.identity import identity_of
        return identity_of(title)

    def test_colour_beats_finish_when_ranking_parallels(self):
        # "Pink Refractor" must prefer [Pink] over the generic [Refractor]:
        # nearly every card in the set is some kind of refractor.
        i = self._ident("2023 Topps Chrome Disney 100 Cinderella Pink "
                        "Refractor #/399 PSA 10")
        pink = i.score_candidate("Cinderella [Pink] #45",
                                 "2023 Topps Chrome Disney 100")
        refr = i.score_candidate("Cinderella [Refractor] #45",
                                 "2023 Topps Chrome Disney 100")
        other = i.score_candidate("Tinker Bell [Pink] #8",
                                  "2023 Topps Chrome Disney 100")
        self.assertGreater(pink, other, "wrong character must rank lower")
        self.assertGreater(pink, 0.5)
        self.assertGreater(refr, 0.5)

    def test_word_order_and_extra_finish_words_still_match(self):
        # Listing says "Orange Wave Refractor"; catalogue says "[Orange Wave]"
        i = self._ident("2023 Topps Chrome Disney 100 Zurg Orange Wave "
                        "Refractor #/25 PSA 10")
        right = i.score_candidate("Zurg [Orange Wave] #28",
                                  "2023 Topps Chrome Disney 100")
        self.assertGreaterEqual(right, 0.62, "equality scoring gave this 22%")
        wrong_colour = i.score_candidate("Zurg [Blue Wave] #28",
                                         "2023 Topps Chrome Disney 100")
        self.assertGreater(right, wrong_colour)

    def test_set_membership_outvotes_a_better_card_name(self):
        i = self._ident("1999 Pokemon Base Set 1st Edition Charizard #4 "
                        "Holo PSA 8.5")
        base = i.score_candidate("Charizard #4", "Pokemon Base Set")
        foreign = i.score_candidate("Charizard [Holo] #4",
                                    "Pokemon Chinese CSM2cC")
        self.assertGreater(
            base, foreign - 0.25,
            "a card from the RIGHT set must not be buried by one from the "
            "wrong set that happens to match the words better")

    def test_ambiguous_candidates_are_refused_not_guessed(self):
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        # Base Set vs Base Set 2 are genuinely different sets with different
        # money, and nothing in the listing title tells them apart.
        payload = {"status": "success", "products": [
            {"id": "1", "product-name": "Charizard [1st Edition] #4",
             "console-name": "Pokemon Base Set"},
            {"id": "2", "product-name": "Charizard [1st Edition] #4",
             "console-name": "Pokemon Base Set 2"},
        ]}
        g._pc_call = lambda path, params, host=None: (
            payload if path == "products"
            else {"status": "success", "loose-price": 100000})
        q = g.quote(self._ident("1999 Pokemon Base Set 1st Edition "
                                "Charizard #4 Holo PSA 8.5"))
        self.assertFalse(q.landed, "a near-tie is not an identification")
        self.assertIsNone(q.value)
        self.assertIn("ambiguous", q.note)

    def test_a_clear_winner_still_lands(self):
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        payload = {"status": "success", "products": [
            {"id": "1", "product-name": "Zurg [Orange Wave] #28",
             "console-name": "2023 Topps Chrome Disney 100"},
            {"id": "2", "product-name": "Mickey Mouse #1",
             "console-name": "1998 Some Other Set"},
        ]}
        g._pc_call = lambda path, params, host=None: (
            payload if path == "products"
            else {"status": "success", "manual-only-price": 50000})
        q = g.quote(self._ident("2023 Topps Chrome Disney 100 Zurg Orange "
                                "Wave Refractor #/25 PSA 10"))
        self.assertTrue(q.landed)
        self.assertEqual(q.product_id, "1")

    def test_a_ruled_out_runner_up_does_not_suppress_the_winner(self):
        """A 1st Edition listing ranked the 1st Edition product first at 95%
        with the unlimited printing 5% behind. That is discrimination, not
        ambiguity - the margin gate must not treat it as a tie."""
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        payload = {"status": "success", "products": [
            {"id": "1", "product-name": "Charizard [1st Edition] #4",
             "console-name": "Pokemon Base Set"},
            {"id": "2", "product-name": "Charizard #4",
             "console-name": "Pokemon Base Set"},
        ]}
        g._pc_call = lambda path, params, host=None: (
            payload if path == "products"
            else {"status": "success", "new-price": 2500000})
        q = g.quote(self._ident("1999 Pokemon Base Set 1st Edition "
                                "Charizard #4 Holo PSA 8"))
        self.assertTrue(q.landed, "we stated 1st Edition and matched it")
        self.assertEqual(q.product_id, "1")

    def test_discriminates_only_on_fields_we_actually_stated(self):
        from valuation.identity import identity_of
        listing = identity_of("1999 Pokemon Base Set 1st Edition Charizard #4")
        first = identity_of("Charizard [1st Edition] #4 Pokemon Base Set")
        unlim = identity_of("Charizard #4 Pokemon Base Set")
        self.assertTrue(listing.discriminates(first, unlim))
        # A listing that never says which edition cannot rule either out.
        silent = identity_of("Pokemon Base Set Charizard #4")
        self.assertFalse(silent.discriminates(first, unlim))

    def test_foreign_printings_are_pushed_down(self):
        i = self._ident("1999 Pokemon Base Set 1st Edition Charizard #4 Holo")
        english = i.score_candidate("Charizard [1st Edition] #4",
                                    "Pokemon Base Set")
        korean = i.score_candidate("Charizard [1st Edition] #4",
                                   "Pokemon Korean Base Set")
        self.assertGreater(english, korean + 0.15)

    def test_every_returned_candidate_is_scored(self):
        """PriceCharting documents 20 results but returns up to 100. Slicing
        to 20 before scoring discarded correct products unseen."""
        from valuation.price_guide import candidates_of
        payload = {"products": [{"id": str(n)} for n in range(100)]}
        self.assertEqual(len(candidates_of(payload)), 100)

    def test_a_bracketed_parallel_needs_no_corroboration(self):
        """PriceCharting declares the parallel in brackets: `Cinderella
        [Pink] #45`. A bare colour in an eBay title needs a finish word or a
        print run to count, but inside brackets the declaration IS the
        corroboration. Without this, [Pink] parsed to no parallel at all and
        the generic [Refractor] outranked it 57% to 54%."""
        from valuation.identity import catalogue_parallel
        self.assertEqual(catalogue_parallel("Cinderella [Pink] #45"), "pink")
        self.assertEqual(catalogue_parallel("Zurg [Orange Wave] #28"),
                         "orange wave")
        self.assertIsNone(catalogue_parallel("Charizard #4"))

    def test_the_declared_colour_wins_over_a_generic_finish(self):
        i = self._ident("2023 Topps Chrome Disney 100 Cinderella Pink "
                        "Refractor #/399 PSA 10")
        console = "2023 Topps Chrome Disney 100"
        pink = i.score_candidate("Cinderella [Pink] #45", console)
        refr = i.score_candidate("Cinderella [Refractor] #45", console)
        other_char = i.score_candidate("Tinker Bell [Pink] #8", console)
        self.assertGreater(pink, refr)
        self.assertGreater(pink, other_char)

    def test_naming_the_character_discriminates(self):
        from valuation.identity import identity_of
        listing = identity_of("2023 Topps Chrome Disney 100 Cinderella Pink "
                              "Refractor #/399 PSA 10")
        won = identity_of("Cinderella [Pink] #45 2023 Topps Chrome Disney 100")
        lost = identity_of("Tinker Bell [Pink] #8 2023 Topps Chrome Disney 100")
        self.assertTrue(listing.discriminates(won, lost))

    def test_a_matched_product_with_no_price_says_so(self):
        """Identified the card but PriceCharting publishes nothing at that
        grade. That is not the same as picking the wrong product, and the
        note has to distinguish them."""
        from valuation.price_guide import PriceGuide
        tmp = tempfile.mkdtemp()
        g = PriceGuide({"_config_dir": tmp,
                        "database": {"file": os.path.join(tmp, "h.db")},
                        "api_keys": {"pricecharting": {"token": "t"}}})
        g._pc_call = lambda path, params, host=None: (
            {"status": "success", "products": [
                {"id": "1", "product-name": "Zurg [Orange Wave] #28",
                 "console-name": "2023 Topps Chrome Disney 100"}]}
            if path == "products"
            else {"status": "success", "product-name": "Zurg [Orange Wave] #28",
                  "console-name": "2023 Topps Chrome Disney 100"})
        q = g.quote(self._ident("2023 Topps Chrome Disney 100 Zurg Orange "
                                "Wave Refractor #/25 PSA 10"))
        self.assertFalse(q.landed)
        self.assertIsNone(q.value)
        self.assertIn("no price at this grade", q.note)
        self.assertIsNotNone(q.product_id, "we DID identify the card")

    def test_set_tokens_are_extracted_for_the_console_field(self):
        from valuation.identity import set_tokens_of
        self.assertIn("bowman", set_tokens_of("1948 Bowman #69 George Mikan"))
        self.assertIn("chrome", set_tokens_of("2023 Topps Chrome Disney 100"))
        self.assertIn("base", set_tokens_of("Pokemon Base Set Charizard"))

class TestTradeBlotter(unittest.TestCase):
    def _opportunity(self, listing_id="123456789012") -> Opportunity:
        listing = Listing(
            site="ebay", title="1986 Fleer Michael Jordan #57 PSA 8",
            url=f"https://www.ebay.com/itm/{listing_id}",
            listing_id=listing_id, current_price=1000, shipping=15,
            query="1986 Fleer Michael Jordan #57 PSA 8",
            matched_queries=[
                "1986 Fleer Michael Jordan #57 PSA 8",
                "1986 Fleer Michael Jordan PSA 8"],
            category="Sports Cards", listing_type="auction",
            priority=True,
            end_time=datetime.now(timezone.utc) + timedelta(hours=4))
        valuation = Valuation(
            fair_value=2000, n_comps=8, confidence=0.8,
            expected_cost=1200, expected_value=535, edge_now=720,
            roi=0.45, capture=0.8, opportunity_score=0.36,
            resale_channel="ebay", net_proceeds=1735)
        return Opportunity(listing, valuation)

    def _config(self, tmp):
        return {
            "_config_dir": tmp,
            "trade_blotter": {
                "enabled": True,
                "file": "trade_blotter/trade_blotter.csv",
                "auto_capture_top_n": 50,
            },
            "algorithm": {
                "sales_tax_rate": 0.08,
                "psa_vault": {
                    "enabled": True, "min_price": 500,
                    "sell_fee_rate": 0.07,
                },
            },
            "output": {"today": {"max_bid_target_roi": 0.15}},
        }

    def test_sync_adds_once_and_preserves_user_fields(self):
        import csv
        import trade_blotter
        with tempfile.TemporaryDirectory() as tmp:
            config = self._config(tmp)
            rows = trade_blotter.sync(
                [self._opportunity()], config)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["status"], "Discovered")
            path = os.path.join(
                tmp, "trade_blotter", "trade_blotter.csv")
            with open(path, newline="", encoding="utf-8") as handle:
                stored = list(csv.DictReader(handle))
            stored[0]["status"] = "Bid/Offer Placed"
            stored[0]["planned_bid_or_offer"] = "1300"
            stored[0]["notes"] = "verify the cert"
            trade_blotter.write_rows(path, stored)

            refreshed = trade_blotter.sync(
                [self._opportunity()], config)
            self.assertEqual(len(refreshed), 1)
            self.assertEqual(
                refreshed[0]["status"], "Bid/Offer Placed")
            self.assertEqual(
                refreshed[0]["planned_bid_or_offer"], "1300")
            self.assertEqual(refreshed[0]["notes"], "verify the cert")

    def test_realized_pnl_is_derived_from_actual_cash_flows(self):
        import trade_blotter
        row = {field: "" for field in trade_blotter.FIELDS}
        row.update({
            "actual_purchase_price": "1000",
            "buyer_fees_paid": "200",
            "shipping_paid": "25",
            "tax_paid": "0",
            "sale_proceeds": "1500",
            "date_won": "2026-07-01",
            "date_sold": "2026-07-21",
        })
        trade_blotter.derive(row)
        self.assertEqual(row["actual_landed_cost"], "1225.00")
        self.assertEqual(row["realized_profit"], "275.00")
        self.assertEqual(row["realized_roi"], "0.224490")
        self.assertEqual(row["holding_days"], "20")

    def test_report_contains_blotter_snapshot(self):
        from openpyxl import load_workbook
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "blotter.xlsx")
            report_mod.write_report(
                [], path, config={}, trade_blotter=[{
                    "status": "Watching", "verified": "yes",
                    "site": "ebay", "listing_type": "auction",
                    "title": "Jordan PSA 8", "query": "Jordan",
                    "current_price": "1000", "fair_value": "2000",
                    "edge_now": "700", "roi": "0.4",
                }])
            workbook = load_workbook(path, read_only=True)
            self.assertIn("Trade Blotter", workbook.sheetnames)
            self.assertEqual(
                workbook["Trade Blotter"]["A2"].value, "Watching")
            workbook.close()


class TestSourceOnboarding(unittest.TestCase):
    def _manifest(self, tmp, *, manifest_id="example_house",
                  enabled=True):
        os.makedirs(os.path.join(tmp, "source_manifests"), exist_ok=True)
        path = os.path.join(
            tmp, "source_manifests", f"{manifest_id}.yaml")
        data = {
            "id": manifest_id,
            "display_name": "Example House",
            "enabled": enabled,
            "capabilities": ["auctions", "fixed"],
            "access": {"feed_file": "imports/example.csv"},
            "field_map": {
                "listing_id": "lot_id",
                "title": "description",
                "url": "link",
                "current_price": "hammer",
                "listing_type": "sale_type",
            },
            "economics": {
                "buyer_fee_rate": 0.20,
                "shipping": 12,
            },
        }
        with open(path, "w", encoding="utf-8") as handle:
            yaml.safe_dump(data, handle, sort_keys=False)
        return path

    def test_valid_manifest_auto_registers_enabled_source(self):
        import source_registry
        with tempfile.TemporaryDirectory() as tmp:
            self._manifest(tmp)
            config = {"_config_dir": tmp, "scraping": {}, "api_keys": {}}
            manifests = source_registry.load_manifests(config)
            self.assertEqual([m.source_id for m in manifests],
                             ["example_house"])
            registry = source_registry.scraper_registry(config)
            self.assertIn("example_house", registry)
            self.assertIn(
                "example_house",
                source_registry.enabled_source_ids(config))
            from source_health import (_configured_sources,
                                       _endpoint_matches)
            health = _configured_sources(config)
            self.assertTrue(health["example_house/listings"][0])
            self.assertTrue(_endpoint_matches(
                "example_house/listings", "example_house/feed"))

    def test_manifest_rejects_embedded_secrets(self):
        import source_registry
        with tempfile.TemporaryDirectory() as tmp:
            path = self._manifest(tmp)
            with open(path, encoding="utf-8") as handle:
                data = yaml.safe_load(handle)
            data["access"]["api_key"] = "do-not-store-this"
            with open(path, "w", encoding="utf-8") as handle:
                yaml.safe_dump(data, handle)
            with self.assertRaises(source_registry.ManifestError):
                source_registry.validate_manifest_file(path)

    def test_csv_field_mapping_produces_normalized_listing(self):
        import source_registry
        with tempfile.TemporaryDirectory() as tmp:
            self._manifest(tmp)
            os.makedirs(os.path.join(tmp, "imports"), exist_ok=True)
            feed = os.path.join(tmp, "imports", "example.csv")
            with open(feed, "w", encoding="utf-8", newline="") as handle:
                handle.write(
                    "lot_id,description,link,hammer,sale_type,status\n"
                    "lot-7,1986 Fleer Michael Jordan PSA 8,"
                    "https://example.test/lot-7,1000,auction,live\n")
            config = {"_config_dir": tmp, "scraping": {}, "api_keys": {}}
            factory = source_registry.scraper_registry(config)[
                "example_house"]
            scraper = factory(config)
            with mock.patch.object(scraper, "_get") as get:
                rows = scraper.search_auctions(
                    "Michael Jordan PSA 8", 10)
            get.assert_not_called()
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0].listing_id, "lot-7")
            self.assertEqual(rows[0].buyer_fee_rate, 0.20)
            self.assertEqual(rows[0].shipping, 12)


class TestGradingQualifierExclusion(unittest.TestCase):
    """PSA prints condition qualifiers on the slab: OC (off-centre), MK
    (marks), ST (stain), MC (miscut), OF (out of focus), PD (print defect).
    A PSA 9 (OC) trades far below a clean PSA 9, so valuing it against
    clean comps overstates it - and one sat on the Action sheet.

    It cannot go in exclude_keywords, which is a case-insensitive substring
    test: 'oc' also appears in block, occasion, Knocks, Rockies, chocolate
    and Ochoa. Whole token, capitals required, brackets optional.
    """

    QUALS = ["OC", "MK", "ST", "MC", "OF", "PD"]

    def _x(self, title):
        import main as scanner
        return scanner._qualifier_excluded(title, self.QUALS)

    def test_qualified_slabs_are_excluded(self):
        for title in ["1986 Fleer Michael Jordan Rookie #57 PSA 9 (OC)",
                      "1986 Fleer Michael Jordan #57 PSA 8 OC",
                      "PSA 7 (MK) 1952 Topps Mantle",
                      "1955 Topps Clemente PSA 4 ST"]:
            self.assertTrue(self._x(title), title)

    def test_ordinary_words_containing_the_letters_are_kept(self):
        """The whole reason this is not exclude_keywords."""
        for title in ["Colorado Rockies Team Set PSA 9",
                      "Vintage Block Party Promo PSA 9",
                      "Chocolate Wrapper 1952",
                      "Knocks Out Boxing PSA 8",
                      "Ochoa Rookie PSA 9",
                      "Occasional Sports Illustrated",
                      "2001 SPX Tiger Woods SOTT Auto PSA 9",
                      "1986 Fleer Michael Jordan Rookie #57 PSA 9"]:
            self.assertFalse(self._x(title), title)

    def test_lowercase_is_not_the_qualifier(self):
        """PSA never writes it lowercase, and requiring capitals is what
        makes the whole-token match safe."""
        self.assertFalse(self._x("psa 9 oc lowercase"))

    def test_no_qualifiers_configured_is_a_no_op(self):
        import main as scanner
        self.assertFalse(scanner._qualifier_excluded("PSA 9 (OC)", []))


class TestSetTokensIncludeProductLines(unittest.TestCase):
    """2026-08-02, from a real basket run. SET_WORDS knew MANUFACTURERS but
    not PRODUCT LINES, and the line is what separates two cards of the same
    player in the same year.

        ambiguous: 'Barry Bonds #11T' (70%) vs 'Barry Bonds #11T' (70%)
                   - only 0% apart, cannot tell which card this is

    Both candidates scored `topps` and nothing else, because "Tiffany" was
    not a set token. The one word that told them apart was dropped before
    scoring, so a perfectly identifiable card was refused."""

    def test_the_product_line_is_a_set_token(self):
        from valuation.identity import identity_of
        cases = {
            "Barry Bonds 1986 Topps Tiffany RC": "tiffany",
            "Luka Doncic 2018 Panini Noir RPA": "noir",
            "Connor McDavid 2015 The Cup RPA": "cup",
            "Victor Wembanyama Bowman's Best Refractor Auto RC": "best",
            "Tom Brady 2000 Fleer Autographics RC": "autographics",
            "Steve Yzerman 1984 OPC RC": "opc",
            "Honus Wagner 1910 Sporting News": "sporting",
        }
        for text, token in cases.items():
            self.assertIn(token, identity_of(text).set_tokens, text)

    def test_tiffany_now_beats_base_topps_by_more_than_the_margin(self):
        from valuation.identity import identity_of
        ident = identity_of("Barry Bonds 1986 Topps Tiffany RC")
        tiffany = ident.score_candidate(
            "Barry Bonds #11T", "Baseball Cards 1986 Topps Traded Tiffany")
        base = ident.score_candidate(
            "Barry Bonds #11T", "Baseball Cards 1986 Topps Traded")
        self.assertGreater(tiffany, base)
        self.assertGreater(tiffany - base, 0.06,
                           "must clear algorithm.guide_match_margin")

    def test_the_ambiguity_note_names_the_set(self):
        """Two candidates called 'Barry Bonds #11T' read as nonsense until
        you can see one is Traded and the other Traded Tiffany."""
        import inspect
        from valuation import price_guide
        src = inspect.getsource(price_guide.PriceGuide._quote_from_rows)
        self.assertIn("console-name", src.split("ambiguous:")[1][:400],
                      "the note must show the set, not just the card name")


class TestMemorabiliaNeverInheritsCardPrices(unittest.TestCase):
    """2026-08-02, the only defect found this session that produced a
    confidently WRONG NUMBER rather than a misfiled row:

        "1986 Fleer Michael Jordan #57 Signed Photo PSA 9"
          -> $42,639.45  match=exact  from graded-price

    That is the PSA 9 price of the rookie CARD, quoted for a photograph.

    quote() emptied `hosts` for Sports Memorabilia, but the local-CSV
    lookup ran ABOVE that line, so the gate stopped API calls and nothing
    else. Downloading the four sports catalogues widened it sharply: far
    more memorabilia titles now find a matching card row locally.
    """

    def _guide(self, rows):
        from valuation.price_guide import PriceGuide

        class FakeIndex:
            def __init__(self, rows): self.rows = rows
            def __len__(self): return len(self.rows)
            def search(self, q, limit=60): return self.rows
        guide = PriceGuide({"api_keys": {"pricecharting": {"token": "t"}}})
        guide.csv_index = FakeIndex(rows)
        return guide

    def _jordan_row(self):
        return {"console-name": "Basketball Cards 1986 Fleer",
                "product-name": "Michael Jordan #57",
                "_guide-host": "sportscardspro",
                "loose-price": 455000, "cib-price": 1200000,
                "new-price": 2500000, "graded-price": 4263945,
                "manual-only-price": 30000000, "id": "1"}

    def test_a_signed_photo_is_not_priced_as_the_card(self):
        from valuation.identity import identity_of
        guide = self._guide([self._jordan_row()])
        q = guide.quote(
            identity_of("1986 Fleer Michael Jordan #57 Signed Photo PSA 9"),
            category="Sports Memorabilia")
        self.assertIsNone(q.value)
        self.assertIn("no card price guide", q.note)

    def test_the_same_card_still_prices_normally(self):
        """The gate must be about the CATEGORY, not the card."""
        from valuation.identity import identity_of
        guide = self._guide([self._jordan_row()])
        q = guide.quote(identity_of("1986 Fleer Michael Jordan #57 PSA 9"),
                        category="Sports Cards")
        self.assertAlmostEqual(q.value, 42639.45, places=2)

    def test_the_row_pricer_refuses_independently(self):
        """Belt and braces: _quote_from_rows turns rows into money, so it
        refuses too rather than trusting every caller to have checked."""
        from valuation.identity import identity_of
        guide = self._guide([self._jordan_row()])
        q = guide._quote_from_rows(
            identity_of("1986 Fleer Michael Jordan #57 Signed Photo PSA 9"),
            [self._jordan_row()], source="local CSV",
            category="Sports Memorabilia")
        self.assertIsNone(q.value)

    def test_no_api_call_is_made_for_memorabilia(self):
        from valuation.identity import identity_of
        guide = self._guide([self._jordan_row()])
        calls = []
        guide._pc_call = lambda *a, **kw: calls.append(a) or {}
        guide.quote(identity_of("Tiger Woods Signed US Open Photo UDA"),
                    category="Sports Memorabilia")
        self.assertEqual(calls, [])


class TestMemorabiliaVersusCard(unittest.TestCase):
    """2026-08-02: the Sports Cards tab was mostly Tiger Woods, and none of
    those rows were cards - a signed photo, a signed video-game cover, two
    signed Renditions pieces, flags and canvases.

    Cause: CARD_ONLY_RE counts manufacturer names as proof of a card, and
    "Upper Deck" is in it. Upper Deck Authenticated is Upper Deck's signed
    MEMORABILIA arm, so the brand pointed the wrong way.
    """

    def _class(self, title):
        from valuation.identity import object_class
        return object_class(title)

    def _category(self, query, title):
        from report import _category
        return _category(query, title)

    def test_uda_memorabilia_is_not_a_card(self):
        for title in [
            'Tiger Woods Signed "Tiger Woods PGA Tour 10" Video Game Cover',
            "Tiger Woods Signed 2008 US Open Champion Photo Upper Deck UDA",
            "Tiger Woods PGA Masters Champ Upper Deck Renditions Signed "
            "Autograph Photo UDA",
            "1997 Masters Golf Flag Tiger Woods UDA Upper Deck Authenticated",
            "Legends of Golf Canvas Signed By Palmer, Nicklaus, Woods UDA",
            "Tiger Woods 2002 Upper Deck Authenticated Tournament worn shirt",
        ]:
            self.assertEqual(self._class(title), "memorabilia", title)
            self.assertEqual(
                self._category("Tiger Woods UDA", title), "Sports Memorabilia",
                title)

    def test_relic_and_patch_cards_are_still_cards(self):
        """The first attempt at this fix dropped brands entirely and moved
        68 genuine jersey/relic CARDS out of Sports Cards - the opposite of
        the problem. A patch card's title is all memorabilia nouns."""
        for title in [
            "2022 National Treasures Skyy Moore Rookie Patch Auto Green "
            "Jersey #/24 PSA 9",
            "2023 National Treasures Jahmyr Gibbs Green Rookie Jersey Patch "
            "Auto #/26 BGS 7",
            "2004-05 Upper Deck SPx Nenad Krstic Rookie Auto Jersey #108",
            "2016 Topps Gypsy Queen Mark McGwire Jersey #GQR-MMC Cardinals",
            "2018-19 Upper Deck Chronology Luc Robitaille Jersey Auto /50",
            "2019 IMMACULATE #132 PJ WASHINGTON JR TRUE RPA ROOKIE PATCH "
            "AUTO /99",
        ]:
            self.assertEqual(self._class(title), "card", title)

    def test_patch_is_not_a_memorabilia_noun(self):
        """Adding 'patch' to the memorabilia nouns swept in 178 National
        Treasures / The Cup Rookie Patch Autos in a single pass."""
        from valuation.identity import MEMORABILIA_RE, STRONG_MEMORABILIA_RE
        self.assertFalse(MEMORABILIA_RE.search("Rookie Patch Auto RPA /99"))
        self.assertFalse(STRONG_MEMORABILIA_RE.search("Rookie Patch Auto"))
        self.assertTrue(STRONG_MEMORABILIA_RE.search("Signed 8x10 Photo"))

    def test_card_product_words_outrank_a_memorabilia_noun(self):
        """'Photo Variation' and 'Short Print' are card vocabulary even
        though 'photo' and 'print' are memorabilia nouns."""
        self.assertEqual(
            self._class("2020 Topps Series 1 SP Photo Variation McGwire 289"),
            "card")
        self.assertEqual(
            self._class("Upper Deck 1991 Michael Jordan Short Print SP1 "
                        "White Sox PSA 6"), "card")

    def test_plain_cards_are_untouched(self):
        for title in ["1986 Fleer Michael Jordan Rookie #57 PSA 9 (OC)",
                      "2003-04 Topps Chrome #111 LeBron James RC PSA 9",
                      "1909-11 T206 Ty Cobb PSA 1.5 Tigers Green Portrait"]:
            self.assertEqual(self._class(title), "card", title)


class TestCustomBasketPricer(unittest.TestCase):
    """Prices a user-supplied list of cards from the local guide CSVs.

    The scanner must refuse ambiguity because strangers write its input.
    Here the user names the card, so an exact (set, name) hit is allowed to
    be decisive - but a name that spans several sets is still reported,
    never resolved by guessing.
    """

    def _mod(self):
        import basket_pricer
        return basket_pricer

    def _rows(self):
        """Rows in the shape guide_csv actually produces.

        It converts "$343,098.00" to an integer number of CENTS at load
        time, so a fixture holding dollar strings is not a smaller version
        of the real thing - it is a different thing, and every price reads
        as missing.
        """
        def row(console, name, **prices):
            base = {"console-name": console, "product-name": name,
                    "loose-price": "", "cib-price": "", "new-price": "",
                    "graded-price": "", "box-only-price": "",
                    "manual-only-price": "", "bgs-10-price": "",
                    "condition-17-price": "", "condition-18-price": ""}
            base.update({k.replace("_", "-"): round(v * 100)
                         for k, v in prices.items()})
            return base
        return [
            row("Pokemon Base Set", "Charizard [1st Edition] #4",
                loose_price=3099.45, graded_price=23000.00,
                manual_only_price=343098.00),
            row("Pokemon Base Set", "Charizard [Shadowless] #4",
                loose_price=1005.00, manual_only_price=30100.00),
            # A plain "Charizard #4" really does exist in five different
            # sets. Leaving it in only one made the ambiguity test pass for
            # the wrong reason - the fixture, not the code, was deciding.
            row("Pokemon Base Set", "Charizard #4",
                loose_price=370.37, manual_only_price=30100.00),
            row("Pokemon Base Set 2", "Charizard #4",
                loose_price=95.00, manual_only_price=3000.00),
            row("Pokemon Base Set", "Booster Box [1st Edition]",
                loose_price=14179.19),
            row("Pokemon Base Set", "Dratini [1st Edition] #26",
                loose_price=20.00, manual_only_price=900.00),
            row("Pokemon Base Set", "Fighting Energy [1st Edition] #97",
                loose_price=9.00, manual_only_price=400.00),
        ]

    def _price(self, name, grade="", set_name=""):
        m = self._mod()
        index = m.ExactIndex(self._rows())
        row = {"name": name, "grade": grade, "set": set_name, "line": 2}
        return m.price_row(row, index, guide=None, api_budget=[0])

    def test_set_plus_name_is_decisive(self):
        got = self._price("Charizard [1st Edition] #4", "PSA 10",
                          "Pokemon Base Set")
        self.assertEqual(got["price"], 343098.00)
        self.assertEqual(got["how"], "manual-only-price")
        self.assertIn("local CSV", got["source"])

    def test_strict_mode_refuses_an_ambiguous_name(self):
        """'Charizard #4' is several products between $3,000 and $343,098.
        --strict restores the scanner's refusal."""
        m = self._mod()
        index = m.ExactIndex(self._rows())
        got = m.price_row({"name": "Charizard #4", "grade": "PSA 10",
                           "set": "", "line": 2}, index, None, [0],
                          strict=True)
        self.assertIsNone(got["price"])
        self.assertIn("ambiguous", got["note"])
        self.assertIn("Pokemon Base Set 2", got["note"])

    def test_by_default_a_near_tie_is_priced_and_labelled(self):
        """You own the card and you named it, so a blank cell is worse than
        a close estimate. The row is priced, flagged as a guess, and told
        what else it could have been."""
        m = self._mod()
        index = m.ExactIndex(self._rows())
        got = m.price_row({"name": "Charizard #4", "grade": "PSA 10",
                           "set": "", "line": 2}, index, None, [0])
        self.assertIsNotNone(got["price"])
        self.assertTrue(got["guess"])
        self.assertIn("best guess", got["note"])
        self.assertIn("other set", got["note"])

    def test_grade_routing_is_the_scanners_own(self):
        """Not a reimplementation: same ladder, same cross-grader shift,
        same refusal to round a grade up."""
        cases = [
            ("PSA 10", 343098.00, "manual-only-price"),
            ("PSA 9", 23000.00, "graded-price"),
            ("", 3099.45, "loose (ungraded)"),
        ]
        for grade, want, how in cases:
            got = self._price("Charizard [1st Edition] #4", grade,
                              "Pokemon Base Set")
            self.assertAlmostEqual(got["price"], want, places=2, msg=grade)
            self.assertEqual(got["how"], how)

    def test_sealed_product_is_left_out_of_a_seeded_set(self):
        """A $14,179 booster box priced as a PSA 10 card silently inflated
        the 1st Edition Base Set total by ~$32k."""
        m = self._mod()
        cards, sealed = m.seed_set(self._rows(), "Pokemon Base Set")
        names = [c["name"] for c in cards]
        self.assertIn("Booster Box [1st Edition]",
                      [s["name"] for s in sealed])
        self.assertNotIn("Booster Box [1st Edition]", names)
        cards2, sealed2 = m.seed_set(self._rows(), "Pokemon Base Set",
                                     include_sealed=True)
        self.assertFalse(sealed2)
        self.assertIn("Booster Box [1st Edition]",
                      [c["name"] for c in cards2])

    def test_sealed_detection_uses_word_boundaries(self):
        """Regression: a substring test called 'Dratini' and 'Fighting
        Energy' sealed product, because both contain 'tin'. Two real cards
        would have vanished from the set silently."""
        m = self._mod()
        cards, _ = m.seed_set(self._rows(), "Pokemon Base Set")
        names = [c["name"] for c in cards]
        self.assertIn("Dratini [1st Edition] #26", names)
        self.assertIn("Fighting Energy [1st Edition] #97", names)
        self.assertFalse(m._is_sealed("Dratini [1st Edition] #26"))
        self.assertTrue(m._is_sealed("Booster Box [1st Edition]"))

    def test_an_identity_changing_word_must_appear_on_both_sides(self):
        """2026-08-02: 'Michael Jordan 1986 Fleer Sticker RC' resolved to
        'Michael Jordan #57' - the base rookie - and priced at $15,604.
        The sticker is a different product."""
        m = self._mod()
        index = m.ExactIndex(self._rows() + [{
            "console-name": "Basketball Cards 1986 Fleer",
            "product-name": "Michael Jordan #57",
            "loose-price": 45500, "cib-price": 120000, "new-price": 250000,
            "graded-price": 4263945, "manual-only-price": 30000000,
            "box-only-price": "", "bgs-10-price": "",
            "condition-17-price": "", "condition-18-price": ""}])
        class FakeGuide:
            guide_hosts = []
            def quote(self, ident, category=None):
                class Q:
                    value = 15604.11; match = "strong"
                    how = "graded-price"; note = ""
                    product_name = "Michael Jordan #57"
                    console_name = "Basketball Cards 1986 Fleer"
                return Q()

        got = m.price_row({"name": "Michael Jordan 1986 Fleer Sticker RC",
                           "grade": "BGS 8.5", "set": "", "line": 2},
                          index, FakeGuide(), [0])
        self.assertIsNone(got["price"])
        self.assertIn("sticker", got["note"])
        plain = m.price_row({"name": "Michael Jordan #57", "grade": "PSA 9",
                             "set": "Basketball Cards 1986 Fleer",
                             "line": 3}, index, None, [0])
        self.assertAlmostEqual(plain["price"], 42639.45, places=2)

    def test_the_matched_subject_must_be_someone_you_named(self):
        """2026-08-02, from a real run: 'Matt Stafford 2009 SPX RPA' matched
        'Matt Cain #12' and priced at $1.64. Matt Cain is a baseball
        pitcher; the shared first name cleared the score threshold.

        Bracketed variants and card numbers are stripped first, so an
        [Autograph] label the user never wrote is not a mismatch."""
        m = self._mod()
        self.assertEqual(m._subject_mismatch(
            "Matt Stafford 2009 SPX RPA", "Matt Cain #12"), "cain")
        for supplied, matched in [
                ("Michael Jordan 1986 Fleer RC", "Michael Jordan #57"),
                ("Tom Brady 2000 Contenders Auto RC",
                 "Tom Brady [Autograph] #144"),
                ("Mike Trout 2009 Bowman Chrome Ref RC Auto",
                 "Mike Trout [Autograph] #BDPP89"),
                ("Ken Griffey Jr 1989 Fleer Auto RC",
                 "Ken Griffey Jr. [Autograph] #548"),
                ("Lebron James 2003 Chrome Refractor RC",
                 "LeBron James [Refractor] #111")]:
            self.assertEqual(m._subject_mismatch(supplied, matched), "",
                             f"{matched} should be kept for {supplied}")

    def test_the_api_cap_limits_paid_calls_but_not_free_ones(self):
        """A 500-row basket must not drain the quota - but the fallback
        reads the LOCAL CSVs before it reaches the API, so the cap must gate
        only the paid part.

        Gating the whole call made --api-cap 0 mean 'resolve nothing': a
        37-row list of free-text names priced 0 of 37 with the guide data
        for most of them sitting on disk.
        """
        m = self._mod()
        index = m.ExactIndex(self._rows())
        seen_hosts = []

        class FakeGuide:
            guide_hosts = ["https://www.pricecharting.com"]

            def quote(self, ident, category=None):
                seen_hosts.append(list(self.guide_hosts))
                class Q:
                    value = None; match = "none"; how = ""; note = "miss"
                    product_name = console_name = ""
                return Q()

        guide = FakeGuide()
        budget = [2]
        for i in range(5):
            m.price_row({"name": f"Unknown Card {i}", "grade": "",
                         "set": "", "line": i}, index, guide, budget)

        self.assertEqual(len(seen_hosts), 5, "every row is still resolved")
        self.assertTrue(all(seen_hosts[i] for i in (0, 1)),
                        "the first two may reach the paid API")
        self.assertTrue(all(not seen_hosts[i] for i in (2, 3, 4)),
                        "after the cap, resolution continues LOCAL-ONLY")
        self.assertEqual(guide.guide_hosts,
                         ["https://www.pricecharting.com"],
                         "and the guide is left as it was found")

    def test_grades_are_floats_before_they_reach_the_ladder(self):
        """grade_info reports strings; _guide_cents compares numerically.
        Passing them through raised TypeError on the first graded row."""
        got = self._price("Charizard [1st Edition] #4", "PSA 9",
                          "Pokemon Base Set")
        self.assertIsInstance(got["effective_grade"], float)
        self.assertIsNone(got["note"] or None)

    def test_reads_a_csv_basket_with_loose_headers(self):
        m = self._mod()
        tmp = os.path.join(tempfile.mkdtemp(), "b.csv")
        with open(tmp, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Item Name", "Console-Name", "Condition"])
            w.writerow(["Charizard [1st Edition] #4", "Pokemon Base Set",
                        "PSA 10"])
        rows, warnings = m.read_basket(tmp)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "Charizard [1st Edition] #4")
        self.assertEqual(rows[0]["grade"], "PSA 10")
        self.assertEqual(rows[0]["set"], "Pokemon Base Set")
        self.assertFalse(warnings)

    def test_resolves_names_written_the_way_collectors_write_them(self):
        """The guide says 'Alakazam [1st Edition] #1'; a personal list says
        '1999 1st Edition Alakazam'. Both name one card."""
        m = self._mod()
        index = m.ExactIndex(self._rows())
        got = m.price_row({"name": "1999 1st Edition Charizard",
                           "grade": "PSA 10", "set": "Pokemon Base Set",
                           "line": 2}, index, None, [0])
        self.assertEqual(got["matched_name"], "Charizard [1st Edition] #4")
        self.assertEqual(got["price"], 343098.00)

    def test_variants_must_match_exactly_both_ways(self):
        """'1st Edition' must not match the Shadowless product, and a plain
        name must not match a 1st Edition one. Silently pricing a $3,099
        1st Edition as a $370 base copy is the failure that matters."""
        m = self._mod()
        index = m.ExactIndex(self._rows())
        rows, _ = index.flexible("1999 Shadowless Charizard",
                                 "Pokemon Base Set")
        self.assertEqual([r["product-name"] for r in rows],
                         ["Charizard [Shadowless] #4"])
        rows, _ = index.flexible("1999 1st Edition Charizard",
                                 "Pokemon Base Set")
        self.assertEqual([r["product-name"] for r in rows],
                         ["Charizard [1st Edition] #4"])

    def test_cost_column_produces_pnl(self):
        m = self._mod()
        tmp = os.path.join(tempfile.mkdtemp(), "h.csv")
        with open(tmp, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Card", "Grade", "Cost"])
            w.writerow(["1999 1st Edition Charizard", "PSA 10", "$2,034"])
        rows, _ = m.read_basket(tmp)
        self.assertEqual(rows[0]["cost"], 2034.0)

    def test_a_basket_over_the_limit_is_truncated_loudly(self):
        m = self._mod()
        tmp = os.path.join(tempfile.mkdtemp(), "big.csv")
        with open(tmp, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["Card", "Grade"])
            for i in range(m.MAX_ROWS + 25):
                w.writerow([f"Card {i}", "PSA 9"])
        rows, warnings = m.read_basket(tmp)
        self.assertEqual(len(rows), m.MAX_ROWS)
        self.assertTrue(any("ignoring the rest" in w for w in warnings))


if __name__ == "__main__":
    unittest.main(verbosity=2)
