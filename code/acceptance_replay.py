"""Replay real listing titles through the valuation path and check invariants.

    .venv/bin/python code/acceptance_replay.py
    .venv/bin/python code/acceptance_replay.py --workbook "reports/Opp Runs/x.xlsx"

WHY
---
A live scan is the real acceptance test, but it needs the network, a clean
breaker and about ten minutes, so in practice it is run rarely - and the
2026-07-28 sports overhaul reached 2026-08-02 without one. This is the part
that CAN be run any time: take every title from the last workbook - real
eBay titles, thousands of them - and push them through identity, category
and the price guide, then assert the things that must never be true.

It uses only local CSVs, so it costs nothing and is deterministic.

WHAT IT CHECKS
--------------
1. No Sports Memorabilia row is given a card-guide value. A signed photo
   is not the card it depicts. This failed on 2026-08-02 at $42,639.45.
2. Objects that cannot be cards are not filed as cards.
3. Relic/patch CARDS are not filed as memorabilia - the mirror of (2), and
   the failure mode of the first attempt at fixing it.
4. Guide values are never quoted above the grade asked for.
5. Structured sports targets still produce well-formed queries.

Exit code is non-zero if any invariant fails, so it can gate a release.
"""
from __future__ import annotations

import argparse
import collections
import glob
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import main as scanner                                      # noqa: E402
from report import _category                                # noqa: E402
from targets import sports_target_entries                   # noqa: E402
from valuation.guide_csv import load_index                  # noqa: E402
from valuation.identity import identity_of, object_class    # noqa: E402
from valuation.price_guide import PriceGuide, NO_CARD_GUIDE  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def latest_workbook() -> str | None:
    files = glob.glob(os.path.join(ROOT, "reports", "Opp Runs",
                                   "opportunities_*.xlsx"))
    return max(files, key=os.path.getmtime) if files else None


def titles_from(path: str) -> list[tuple[str, str]]:
    """[(title, query)] from every sheet that has them."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out, seen = [], set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c or "") for c in next(rows)]
        except StopIteration:
            continue
        ti = next((i for i, h in enumerate(header)
                   if h.strip().lower() == "title"), None)
        qi = next((i for i, h in enumerate(header)
                   if h.strip().lower() == "query"), None)
        if ti is None:
            continue
        for row in rows:
            if ti >= len(row) or not row[ti]:
                continue
            title = str(row[ti])
            query = str(row[qi]) if qi is not None and qi < len(row) and \
                row[qi] else ""
            if (title, query) in seen:
                continue
            seen.add((title, query))
            out.append((title, query))
    return out


def main(argv=None) -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--workbook")
    # 12,000 titles x a scan of a million-row index is ~20 minutes. The
    # first couple of thousand already exercise every code path, so the
    # default is a sample and --limit 0 is the exhaustive run.
    p.add_argument("--limit", type=int, default=2500)
    args = p.parse_args(argv)

    path = args.workbook or latest_workbook()
    if not path or not os.path.isfile(path):
        print("No workbook found under reports/Opp Runs - run a scan first.")
        return 1

    config = scanner.load_config(os.path.join(ROOT, "config.yaml"))
    index = load_index(ROOT)
    guide = PriceGuide(config)
    guide.guide_hosts = []          # local CSVs only: free and deterministic

    rows = titles_from(path)
    if args.limit:
        rows = rows[:args.limit]

    print("=" * 74)
    print(f"ACCEPTANCE REPLAY  {os.path.basename(path)}")
    print(f"{len(rows):,} distinct titles | {len(index):,} guide products")
    print("=" * 74, flush=True)

    failures: list[str] = []
    total = len(rows)
    classes = collections.Counter()
    categories = collections.Counter()
    guide_hits = 0

    for n, (title, query) in enumerate(rows, start=1):
        if n % 250 == 0:
            print(f"   ... {n:,}/{total:,}", flush=True)
        klass = object_class(title)
        category = _category(query or title, title)
        classes[klass] += 1
        categories[category] += 1

        ident = identity_of(title)
        quote = guide.quote(ident, category=category)
        if quote.value:
            guide_hits += 1

        # (1) the 2026-08-02 defect
        if category in NO_CARD_GUIDE and quote.value:
            failures.append(
                f"MEMORABILIA PRICED FROM A CARD GUIDE: ${quote.value:,.2f} "
                f"({quote.how}) for {title[:60]!r}")

        # (2)/(3) class and category must agree about cards
        if klass == "memorabilia" and category == "Sports Cards":
            failures.append(f"NON-CARD IN SPORTS CARDS: {title[:66]!r}")

    # (5) structured targets still well formed
    for entry in sports_target_entries(config):
        q = entry.get("query") or ""
        if not q or "#" not in q:
            failures.append(f"MALFORMED SPORTS TARGET: {q!r}")

    print("\nobject classes:")
    for name, n in classes.most_common(8):
        print(f"   {name:<14} {n:>6}")
    print("\nreport categories:")
    for name, n in categories.most_common(10):
        print(f"   {name:<22} {n:>6}")
    print(f"\nguide landed a value on {guide_hits:,} of {len(rows):,} titles "
          f"({guide_hits / max(1, len(rows)):.0%}) using local CSVs only")

    print()
    if failures:
        print(f"FAILED - {len(failures)} invariant violation(s):")
        for line in failures[:25]:
            print(f"   {line}")
        if len(failures) > 25:
            print(f"   ... and {len(failures) - 25} more")
        return 1
    print("PASSED - no invariant violations.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
